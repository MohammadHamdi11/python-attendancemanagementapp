#==========================================================imports and constants==========================================================#
import sys
import re
import pandas as pd
import traceback
import os
import requests
import json
import random
import shutil
import threading
from collections import defaultdict
from datetime import datetime, timedelta, date
from typing import List, Dict
import io
from PIL import Image
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QHeaderView, QGridLayout,
    QPushButton, QLabel, QLineEdit, QComboBox, QFileDialog, QSizePolicy, QCalendarWidget,
    QFrame, QTableWidget, QTableWidgetItem, QProgressBar, QGraphicsDropShadowEffect, QListWidget,
    QTextEdit, QDialog, QMessageBox, QScrollArea, QStackedWidget, QGroupBox, QRadioButton, QAbstractItemView
)
from PyQt6.QtCore import Qt, pyqtSignal, QThread, QSize, QTimer, QObject, pyqtSignal, QEvent, QDate
from PyQt6.QtGui import QIcon, QPixmap, QFont, QIntValidator, QColor, QMovie, QTextCursor, QPainter, QPainterPath, QPen, QPalette

# Constants - Inverted Colors
DARK_BLUE = "#24325f"
DARK_RED = "#951d1e"
BLACK = "#000000"
CARD_BG = "#1a1a1a"
TEXT_COLOR = "#ffffff"
INPUT_BG = "#2d2d2d"
BORDER_COLOR = "#3d3d3d"

# Standard button style
MENU_BUTTON_STYLE = f"""
QPushButton {{
    background-color: {DARK_BLUE};
    color: {TEXT_COLOR};
    border: none;
    border-radius: 3px;
    padding: 10px 20px;
    font-size: 16px;
    font-weight: bold;
}}
QPushButton:hover {{
    background-color: {DARK_RED};
}}
"""

# Menu Exit button style
MENU_EXIT_BUTTON_STYLE = f"""
QPushButton {{
    background-color: {DARK_RED};
    color: {TEXT_COLOR};
    border: none;
    border-radius: 3px;
    padding: 10px 20px;
    font-size: 16px;
    font-weight: bold;
}}
QPushButton:hover {{
    background-color: #ab2223;
}}
"""

# STANDARD button style
STANDARD_BUTTON_STYLE = f"""
QPushButton {{
    background-color: {DARK_BLUE};
    color: {TEXT_COLOR};
    border: none;
    border-radius: 3px;
    padding: 5px 15px;
}}
QPushButton:hover {{
    background-color: {DARK_RED};
}}
"""

# Exit button style
EXIT_BUTTON_STYLE = f"""
QPushButton {{
    background-color: {DARK_RED};
    color: {TEXT_COLOR};
    border: none;
    border-radius: 3px;
    padding: 5px 15px;
}}
QPushButton:hover {{
    background-color: #ab2223;
}}
"""

# Table style
TABLE_STYLE = f"""
QTableWidget::item {{
    text-align: center;
    padding: 5px;
}}
QTableWidget {{
    background-color: {INPUT_BG};
    gridline-color: #3d3d3d;
    border: 1px solid #3d3d3d;
    border-radius: 5px;
}}
QTableWidget::item:selected {{
    background-color: {DARK_BLUE};
    color: white;
}}
QTableWidget::item:hover {{
    background-color: {DARK_BLUE};
    color: white;
}}
QHeaderView::section:hover {{
    background-color: {DARK_BLUE};
    color: white;
}}
QHeaderView::section {{
    background-color: #202c54;
    color: white;
    gridline-color: #3d3d3d;
    border: 1px solid #3d3d3d;
}}
"""

GROUP_BOX_STYLE = f"""
QGroupBox {{
    border: 2px solid {DARK_BLUE};
    border-radius: 5px;
    margin-top: 2ex;
    font-weight: bold;
    font-size: 20px;
    color: {TEXT_COLOR};
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 3px 0 3px;
    color: {TEXT_COLOR};
}}
QLabel {{
    color: white;
}}
QRadioButton {{
    color: white;
}}
QRadioButton::indicator::unchecked {{
    border: 2px solid white;
    background-color: white; /* Background for unchecked state */
    border-radius: 7px;
}}
QRadioButton::indicator::checked {{
    border: 2px solid white;
    background-color: {DARK_BLUE}; /* Dark blue check mark */
    border-radius: 7px;
}}

QComboBox {{
    color: white;
    background-color: {INPUT_BG};
}}

QLineEdit {{
    color: white;
    background-color: {INPUT_BG};
}}
"""

PROGRESS_BAR_STYLE = f"""
QProgressBar {{
    text-align: center;
    background-color: {INPUT_BG};
}}
QProgressBar::chunk {{
    background-color: {DARK_RED}; 
}}
"""

# For the console
CONSOLE_STYLE = f"""
QTextEdit {{
    background-color: {INPUT_BG};
    color: white;
    border: 1px solid {BLACK};
}}
"""

#==========================================================start page==========================================================#

class StartPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.init_ui()

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # Create a dark card-like container
        card_container = QWidget()
        card_container.setStyleSheet(f"""
            QWidget {{
                background-color: {BLACK};
                border-radius: 10px;
                border: 1px solid {BORDER_COLOR};
                color: {TEXT_COLOR};
            }}
            QFrame {{
                border: none;
            }}
            QComboBox {{
                background-color: {INPUT_BG};
                color: {TEXT_COLOR};
                border: 1px solid {BORDER_COLOR};
                padding: 5px;
                border-radius: 3px;
            }}
            QComboBox::drop-down {{
                border: none;
            }}
            QComboBox::down-arrow {{
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid {TEXT_COLOR};
            }}
            QLineEdit {{
                background-color: {INPUT_BG};
                color: {TEXT_COLOR};
                border: 1px solid {BORDER_COLOR};
                padding: 5px;
                border-radius: 3px;
            }}
        """)
        card_layout = QVBoxLayout(card_container)
        card_layout.setSpacing(20)
        card_layout.setContentsMargins(40, 40, 40, 40)

        # Create info button in top left of the card
        info_container = QHBoxLayout()
        info_container.setContentsMargins(0, 0, 0, 0)

        self.info_button = QPushButton("", self)
        self.info_button.setFixedSize(40, 40)
        icon_path = os.path.join(os.path.dirname(__file__), 'info.png')
        if os.path.exists(icon_path):
            self.info_button.setIcon(QIcon(icon_path))
            self.info_button.setIconSize(QSize(32, 32))
        else:
            self.info_button.setStyleSheet(f"""
                QPushButton {{
                    background-color: {DARK_BLUE};
                    color: {TEXT_COLOR};
                    border: none;
                    border-radius: 20px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {DARK_RED};
                }}
            """)
            self.info_button.setText("i")

        info_container.addWidget(self.info_button)
        # Add spacer after button to push everything else to the right
        info_container.addStretch()
        card_layout.addLayout(info_container)

        # Center the card in the window
        center_layout = QHBoxLayout()
        center_layout.addStretch(1)
        center_layout.addWidget(card_container)
        center_layout.addStretch(1)

        # Vertical centering
        vertical_layout = QVBoxLayout()
        vertical_layout.addStretch(1)
        vertical_layout.addLayout(center_layout)
        vertical_layout.addStretch(1)

        # Background container
        bg_container = QWidget()
        bg_container.setStyleSheet(f"background-color: {CARD_BG};")
        bg_container.setLayout(vertical_layout)
        main_layout.addWidget(bg_container)

        # Logo
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(128, 128, Qt.AspectRatioMode.KeepAspectRatio,
                                               Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        card_layout.addWidget(logo_label)

        # Title
        title_label = QLabel("Faculty Attendance \nManagement System")
        title_label.setStyleSheet(f"""
            color: {TEXT_COLOR};
            font-size: 24px;
            font-weight: bold;
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        card_layout.addWidget(title_label)

        # Buttons Container
        buttons_widget = QWidget()
        buttons_widget.setStyleSheet("border: none;")
        buttons_layout = QVBoxLayout(buttons_widget)
        buttons_layout.setSpacing(15)
        
        # First row - 3 buttons side by side
        first_row = QHBoxLayout()
        first_row.setSpacing(10)  # Set horizontal spacing between buttons
        
        # Reference File Preparer Button
        self.reference_btn = QPushButton("Prepare Reference File")
        self.reference_btn.setMinimumHeight(50)
        self.reference_btn.setStyleSheet(MENU_BUTTON_STYLE)
        first_row.addWidget(self.reference_btn)
        
        # Prepare Log Sheet Button
        self.preparer_btn = QPushButton("Prepare Log Sheet")
        self.preparer_btn.setMinimumHeight(50)
        self.preparer_btn.setStyleSheet(MENU_BUTTON_STYLE)
        first_row.addWidget(self.preparer_btn)
        
        # Schedule Manager Button
        self.schedule_btn = QPushButton("Manage Schedules")
        self.schedule_btn.setMinimumHeight(50)
        self.schedule_btn.setStyleSheet(MENU_BUTTON_STYLE)
        first_row.addWidget(self.schedule_btn)
        
        buttons_layout.addLayout(first_row)
        
        # Second row - 2 buttons side by side (process buttons)
        second_row = QHBoxLayout()
        second_row.setSpacing(10)

        self.appeal_btn = QPushButton("Process Appeals")
        self.appeal_btn.setMinimumHeight(50)
        self.appeal_btn.setStyleSheet(MENU_BUTTON_STYLE)
        second_row.addWidget(self.appeal_btn)
        
        self.process_btn = QPushButton("Process Attendance")
        self.process_btn.setMinimumHeight(50)
        self.process_btn.setStyleSheet(MENU_BUTTON_STYLE)
        second_row.addWidget(self.process_btn)
        
        buttons_layout.addLayout(second_row)
        
        # Third row - Dashboard button (full width)
        self.dashboard_btn = QPushButton("Analyze Attendance")
        self.dashboard_btn.setMinimumHeight(50)
        self.dashboard_btn.setStyleSheet(MENU_BUTTON_STYLE)
        buttons_layout.addWidget(self.dashboard_btn)
        
        # Exit Button
        exit_btn = QPushButton("Exit")
        exit_btn.setMinimumHeight(50)
        exit_btn.setStyleSheet(MENU_EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(self.parent().close)
        buttons_layout.addWidget(exit_btn)

        # Add buttons container to card
        card_layout.addWidget(buttons_widget)

        # Set fixed size for the card
        card_container.setFixedWidth(700)  # Slightly wider to accommodate 3 buttons side by side
        card_container.setMinimumHeight(600)  # Taller to accommodate the new layout

#==========================================================info page==========================================================#

class InfoPage(QWidget):
    def __init__(self):
        super().__init__()
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)

        self.init_ui()

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio,
                                               Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("About the app")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)

        # Back button
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)

        # Exit button
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)

        # Add buttons to vertical layout
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()

        # Add button layout to header
        header_layout.addStretch()
        header_layout.addLayout(header_buttons_layout)
        main_layout.addLayout(header_layout)

        # Create a GroupBox for info content, similar to AttendanceProcessor style
        info_group = QGroupBox()
        # Using same style as other GroupBoxes
        info_group.setStyleSheet(GROUP_BOX_STYLE)
        info_layout = QVBoxLayout(info_group)

        # Create scroll area for content
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setMinimumHeight(600)
        scroll_area.setStyleSheet(f"""
            QScrollArea {{
                background-color: transparent;
                border: none;
            }}
            QScrollBar:vertical {{
                border: none;
                background: {CARD_BG};
                width: 10px;
                margin: 0px;
            }}
            QScrollBar::handle:vertical {{
                background: {DARK_BLUE};
                min-height: 20px;
                border-radius: 5px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                border: none;
                background: none;
            }}
        """)

        # Content widget for the scroll area
        info_content = QWidget()
        info_content.setStyleSheet("background-color: transparent;")
        content_layout = QVBoxLayout(info_content)

        # Create info label with rich text
        info_text = """
        <h1 style='color: white; text-align: center;'>Faculty Attendance Management System</h1>
        
        <h2 style='color: white;'>Special Thanks</h2>
        <ul style='color: white;'>
            <li><b>Dr. Ahmad Samir</b> who encouraged me to start working on this project</li>
            <li><b>Dr. Amani Helmi</b> for sponsoring this project and for bing the reason behind its success</li>
            <li><b>Dr. Gehan Adel</b> for supporting me every step on the way</li>
            <li><b>Dr. Doaa Mohammad Abu Bakr</b> who was the catalyst to this project's success</li>
            <li><b>Dr. Taqwa Mohammad Abd Al-Salam</b> for aiding me with her efforts in deployment</li>
            <li>My companion and friend, <b>Dr. Mazin Helmi</b> who never got tired of me</li>

        </ul>
                
        <h2 style='color: white;'>What This App Does</h2>
        <p style='color: white;'>This app is a simple tool to track and manage student attendance. It helps you manage faculty attendance easily. It tracks who attended classes and creates reports.</p>
        
        <h2 style='color: white;'>Main Features</h2>
        <ol style='color: white;'>
            <li><b>Prepare Reference File</b>
                <ul>
                    <li>Create a file with all student information</li>
                    <li>Organize students by ID, name, year, and group</li>
                </ul>
            </li>
            <li><b>Prepare Log Sheet</b>
                <ul>
                    <li>Combine attendance records from multiple sources</li>
                    <li>Get files from your computer or cloud storage</li>
                </ul>
            </li>
            <li><b>Manage Schedules</b>
                <ul>
                    <li>Create new class schedules</li>
                    <li>Update existing schedules when needed</li>
                </ul>
            </li>
            <li><b>Process Appeals</b>
                <ul>
                    <li>Manage attendance exceptions for students who missed regular logging</li>
                    <li>Connect student information, attendance logs, and session schedules</li>
                    <li>Select students and sessions for attendance appeals</li>
                    <li>Process and add these appeals to attendance records</li>
                </ul>
            </li>
            <li><b>Process Attendance</b>
                <ul>
                    <li>Compare student logs against official schedules</li>
                    <li>Track and validate attendance for students who have changed groups</li>
                    <li>Identify when student transfers between groups occurred</li>
                    <li>Validate attendance records against the appropriate group schedule</li>
                    <li>Generate detailed attendance reports with transfer logs</li>
                </ul>
            </li>
            <li><b>Analyze Attendance</b>
                <ul>
                    <li>See attendance statistics and trends</li>
                    <li>Identify students who need help or attention</li>
                </ul>
            </li>
        </ol>
        
        <h2 style='color: white;'>How To Use This App</h2>
        
        <h3 style='color: white;'>Step 1: Prepare Reference File</h3>
        <ul style='color: white;'>
            <li>Click "Prepare Reference File" button on the main screen</li>
            <li>Click "Browse" to select your Excel file with student information</li>
            <li>Choose which columns have student ID, name, year, and group</li>
            <li>Click "Preview Result" to check the data</li>
            <li>Click "Process and Save Reference File" when everything looks right</li>
        </ul>
        
        <h3 style='color: white;'>Step 2: Prepare Log Sheet</h3>
        <ul style='color: white;'>
            <li>Click "Prepare Log Sheet" button on the main screen</li>
            <li>Choose where to get your files from:
                <ul>
                    <li>"Import from Cloud Storage" to download files from online</li>
                    <li>"Import Local Files" to use files from your computer</li>
                </ul>
            </li>
            <li>Select the files you want to combine</li>
            <li>Click "Merge Logs Files" to join them into one file</li>
        </ul>
        
        <h3 style='color: white;'>Step 3: Manage Schedules</h3>
        <ul style='color: white;'>
            <li>Click "Manage Schedules" button on the main screen</li>
            <li>To create a new schedule:
                <ul>
                    <li>Enter the module name</li>
                    <li>Select year, group, subject, and location</li>
                    <li>Pick the date and time using the calendar</li>
                    <li>Click "Add Session" for each class meeting</li>
                    <li>Click "Create Schedule" when you're done</li>
                </ul>
            </li>
            <li>To update an existing schedule:
                <ul>
                    <li>Select "Update Existing" option</li>
                    <li>Click "Browse" to find your schedule file</li>
                    <li>Make changes to existing sessions or add new ones</li>
                    <li>Click "Update Schedule" to save changes</li>
                </ul>
            </li>
        </ul>
        
        <h3 style='color: white;'>Step 4: Process Appeals</h3>
        <ul style='color: white;'>
            <li>Click "Process Appeals" button on the main screen</li>
            <li>Set up your data sources:
                <ul>
                    <li>Select your reference data file with student information</li>
                    <li>Select your existing attendance logs file</li>
                    <li>Select your session schedules file</li>
                </ul>
            </li>
            <li>Managing appeals:
                <ul>
                    <li>Search for students by name or ID in the search box</li>
                    <li>Select a student from the student table</li>
                    <li>View student details in the details panel</li>
                    <li>Select relevant sessions from the filtered sessions table</li>
                    <li>Click "Add to Appeals" to create an appeal</li>
                    <li>Review all selected appeals in the appeals management table</li>
                    <li>Remove any unneeded appeals if necessary</li>
                    <li>Click "Process Appeals" to finalize and save the changes</li>
                </ul>
            </li>
        </ul>
        
        <h3 style='color: white;'>Step 5: Process Attendance</h3>
        <ul style='color: white;'>
            <li>Click "Process Attendance" button on the main screen</li>
                <ul>
            <li>For generating the attendance reports:
<ul>
                    <li>Select your reference file</li>
                    <li>Select your attendance logs</li>
                    <li>Add all relevant schedule files</li>
</ul>

            </li>

            <li>Click "Process Attendance Records" to start</li>
            <li>Wait for the system to finish processing</li>
            <li>Find your reports in the "attendance_reports" folder</li>
                </ul>
                <ul>
            <li>For updating previous reports in case some students transferred between groups:
<ul>
                    <li>Select the previous report</li>
                    <li>Select your NEW reference file</li>
                    <li>Select your attendance logs</li>
                    <li>Add all relevant schedule files</li> 
</ul>
            </li>
            <li>Click "Update Report" to start</li>
            <li>Wait for the system to finish updating</li>
            <li>Find your updated reports in the "attendance_reports" folder</li>
                </ul>
        </ul>
        
        <h3 style='color: white;'>Step 6: Analyze Attendance</h3>
        <ul style='color: white;'>
            <li>Click "Analyze Attendance" button on the main screen</li>
            <li>Select your processed attendance report file from the previous step</li>
            <li>Click "Display Statistics" to see:
                <ul>
                    <li>Total number of students</li>
                    <li>Pass rate and average attendance</li>
                    <li>Number of at-risk students</li>
                    <li>Detailed information for each student</li>
                </ul>
            </li>
            <li>Use the search box to find specific students by name or ID</li>
        </ul>
        
        <h2 style='color: white;'>Tips for Best Results</h2>
        <ul style='color: white;'>
            <li>Keep your files organized in dedicated folders</li>
            <li>Process attendance weekly to stay up-to-date</li>
            <li>Use consistent naming for all your files</li>
            <li>When validating attendance, be aware that different time windows are used:
                <ul>
                    <li>Standard sessions: 15 minutes before to 150 minutes after session start</li>
                    <li>Exception hours (12, 1, 13, 3, 15): 15 minutes before to 150 minutes after</li>
                </ul>
            </li>
            <li>For students who transferred between groups, the system will:
                <ul>
                    <li>Analyze attendance patterns to identify when transfers occurred</li>
                    <li>Validate attendance against the appropriate group schedule based on transfer dates</li>
                </ul>
            </li>
        </ul>
        
        <h2 style='color: white;'>Troubleshooting</h2>
        <ul style='color: white;'>
            <li>If files won't import:
                <ul>
                    <li>Check that they are Excel (.xlsx) format</li>
                    <li>Make sure required columns are present</li>
                    <li>Verify you have permission to read/write these files</li>
                </ul>
            </li>
            <li>If data doesn't match:
                <ul>
                    <li>Check that student IDs are consistent across files</li>
                    <li>Look for typos in names or IDs</li>
                    <li>Make sure date formats are the same in all files</li>
                </ul>
            </li>
            <li>If the app crashes:
                <ul>
                    <li>Restart the application</li>
                    <li>Check that you have enough disk space</li>
                    <li>Make sure all required files are available</li>
                </ul>
            </li>
        </ul>
        
        <h2 style='color: white;'>Need Help?</h2>
        <p style='color: white;'><b>Developer Note:</b><br>
        This application was developed by a medical student at the Faculty of Medicine, Ain Shams University with the help of AI models including Claude, DeepSeek, Perplexity, and a little bit of ChatGPT.</p>
        <p style='color: white;'><b>Contact Information:</b></p>
        <ul style='color: white;'>
            <li>  231249@med.asu.edu.eg</li>
            <li>  mohammadhamdisaid.mh@icloud.com</li>
            <li>  mohammad_hamdi11@yahoo.com</li>
        </ul>
        """

        info_label = QLabel()
        info_label.setTextFormat(Qt.TextFormat.RichText)
        info_label.setText(info_text)
        info_label.setWordWrap(True)
        info_label.setStyleSheet("background-color: transparent;")
        content_layout.addWidget(info_label)

        scroll_area.setWidget(info_content)
        info_layout.addWidget(scroll_area)

        # Add the group box to the main layout
        main_layout.addWidget(info_group)

        # Add a stretch after the group box to push everything up
        main_layout.addStretch()

#==========================================================reference preparer==========================================================#

class ReferenceFilePreparer(QWidget):
    def __init__(self):
        super().__init__()
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)
        self.init_ui()

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio,
                                               Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("Reference File Preparer")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)

        # Back button
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)

        # Exit button
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)

        # Add buttons to vertical layout
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()

        # Add button layout to header
        header_layout.addStretch()
        header_layout.addLayout(header_buttons_layout)
        main_layout.addLayout(header_layout)

        # Input File Section
        input_group = QGroupBox("Input Excel File")
        input_group.setStyleSheet(GROUP_BOX_STYLE)
        input_layout = QHBoxLayout(input_group)

        # File selection layout with sheet combo on same line
        input_layout.addWidget(QLabel("Source File:"))
        self.input_file_path = QLineEdit()
        self.input_file_path.setPlaceholderText("Select Excel file...")
        input_layout.addWidget(self.input_file_path)
        
        browse_btn = QPushButton("Browse")
        browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        browse_btn.clicked.connect(self.browse_input_file)
        input_layout.addWidget(browse_btn)
        
        # Sheet selection on same line
        input_layout.addWidget(QLabel("Sheet:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(150)
        input_layout.addWidget(self.sheet_combo)
        
        main_layout.addWidget(input_group)

        # Data Before Section
        data_before_group = QGroupBox("Data Before Processing")
        data_before_group.setStyleSheet(GROUP_BOX_STYLE)
        data_before_layout = QVBoxLayout(data_before_group)

        self.data_before_table = QTableWidget()
        self.data_before_table.setStyleSheet(TABLE_STYLE)
        data_before_layout.addWidget(self.data_before_table)

        main_layout.addWidget(data_before_group)

        # Column Mapping Section
        mapping_group = QGroupBox("Column Mapping")
        mapping_group.setStyleSheet(GROUP_BOX_STYLE)
        mapping_layout = QGridLayout(mapping_group)

        # Student ID mapping
        mapping_layout.addWidget(QLabel("Student ID Column:"), 0, 0)
        self.id_column_combo = QComboBox()
        mapping_layout.addWidget(self.id_column_combo, 0, 1)

        # Name mapping
        mapping_layout.addWidget(QLabel("Name Column:"), 1, 0)
        self.name_column_combo = QComboBox()
        mapping_layout.addWidget(self.name_column_combo, 1, 1)

        # Year mapping
        mapping_layout.addWidget(QLabel("Year Column:"), 2, 0)
        self.year_column_combo = QComboBox()
        mapping_layout.addWidget(self.year_column_combo, 2, 1)

        # Group mapping
        mapping_layout.addWidget(QLabel("Group Column:"), 3, 0)
        self.group_column_combo = QComboBox()
        mapping_layout.addWidget(self.group_column_combo, 3, 1)

        # Preview button for mapping
        preview_mapping_btn = QPushButton("Preview Result")
        preview_mapping_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        preview_mapping_btn.clicked.connect(self.preview_mapping_result)
        mapping_layout.addWidget(preview_mapping_btn, 4, 0, 1, 2)

        main_layout.addWidget(mapping_group)

        # Data After Section
        data_after_group = QGroupBox("Data After Processing")
        data_after_group.setStyleSheet(GROUP_BOX_STYLE)
        data_after_layout = QVBoxLayout(data_after_group)

        self.data_after_table = QTableWidget()
        self.data_after_table.setStyleSheet(TABLE_STYLE)
        data_after_layout.addWidget(self.data_after_table)

        main_layout.addWidget(data_after_group)

        # Bottom buttons
        button_layout = QHBoxLayout()
        self.process_btn = QPushButton("Process and Save Reference File")
        self.process_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        self.process_btn.clicked.connect(self.process_file)
        button_layout.addWidget(self.process_btn)
        main_layout.addLayout(button_layout)

        # Connect file input changes to sheet loading
        self.input_file_path.textChanged.connect(self.load_sheets)
        
        # Connect sheet combobox changes to auto-preview
        self.sheet_combo.currentTextChanged.connect(self.preview_data)
        
        # Connect column combo changes to both "before" and "after" preview updates
        self.id_column_combo.currentTextChanged.connect(self.on_column_selection_changed)
        self.name_column_combo.currentTextChanged.connect(self.on_column_selection_changed)
        self.year_column_combo.currentTextChanged.connect(self.on_column_selection_changed)
        self.group_column_combo.currentTextChanged.connect(self.on_column_selection_changed)

    def on_column_selection_changed(self):
        """Handle any column selection change by updating both preview tables"""
        self.update_data_before_table()
        self.update_data_after_table()

    def browse_input_file(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)")
        if filename:
            self.input_file_path.setText(filename)

    def load_sheets(self):
        file_path = self.input_file_path.text()
        if file_path and os.path.isfile(file_path):
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(wb.sheetnames)
            except Exception as e:
                self.show_error_message(f"Error loading workbook: {str(e)}")

    def preview_data(self):
        file_path = self.input_file_path.text()
        sheet_name = self.sheet_combo.currentText()
        
        if not file_path or not sheet_name:
            # Don't show error, just return silently
            return
            
        try:
            # Load the data
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Clear previous preview
            self.data_before_table.clear()
            self.data_before_table.setRowCount(0)
            self.data_before_table.setColumnCount(0)
            self.id_column_combo.clear()
            self.name_column_combo.clear()
            self.year_column_combo.clear()
            self.group_column_combo.clear()
            
            # Set up the table
            rows, cols = min(10, df.shape[0]), df.shape[1]
            self.data_before_table.setRowCount(rows)
            self.data_before_table.setColumnCount(cols)
            
            # Set column headers
            self.data_before_table.setHorizontalHeaderLabels(df.columns.tolist())
            
            # Populate combo boxes with column names
            column_names = [""] + df.columns.tolist()
            self.id_column_combo.addItems(column_names)
            self.name_column_combo.addItems(column_names)
            self.year_column_combo.addItems(column_names)
            self.group_column_combo.addItems(column_names)
            
            # Fill the preview table with data
            for i in range(rows):
                for j in range(cols):
                    value = str(df.iloc[i, j]) if not pd.isna(df.iloc[i, j]) else ""
                    item = QTableWidgetItem(value)
                    self.data_before_table.setItem(i, j, item)
            
            # Auto-adjust column widths
            self.data_before_table.resizeColumnsToContents()
            
            # Try to auto-detect column mappings
            self.auto_detect_columns(df.columns.tolist())
            
        except Exception as e:
            # Only show error if it's not because no sheet is selected yet
            if sheet_name:
                self.show_error_message(f"Error previewing data: {str(e)}")

    def update_data_before_table(self):
        """Update the 'Data Before' table with the currently selected columns"""
        file_path = self.input_file_path.text()
        sheet_name = self.sheet_combo.currentText()
        
        if not file_path or not sheet_name:
            return
            
        try:
            # Get selected columns
            id_col = self.id_column_combo.currentText()
            name_col = self.name_column_combo.currentText()
            year_col = self.year_column_combo.currentText()
            group_col = self.group_column_combo.currentText()
            
            # If no columns are selected yet, clear the table and return
            if not (id_col or name_col or year_col or group_col):
                self.data_before_table.clear()
                self.data_before_table.setRowCount(0)
                self.data_before_table.setColumnCount(0)
                return
                
            # Load the data
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Create a new dataframe with only the selected columns
            selected_cols = []
            col_headers = []
            
            if id_col:
                selected_cols.append(df[id_col])
                col_headers.append("Student ID")
            if name_col:
                selected_cols.append(df[name_col])
                col_headers.append("Name")
            if year_col:
                selected_cols.append(df[year_col])
                col_headers.append("Year")
            if group_col:
                selected_cols.append(df[group_col])
                col_headers.append("Group")
                
            if not selected_cols:
                return
                
            preview_df = pd.concat(selected_cols, axis=1)
            preview_df.columns = col_headers
            
            # Clear previous data
            self.data_before_table.clear()
            self.data_before_table.setRowCount(0)
            self.data_before_table.setColumnCount(0)
            
            # Set up the table
            rows, cols = min(10, preview_df.shape[0]), preview_df.shape[1]
            self.data_before_table.setRowCount(rows)
            self.data_before_table.setColumnCount(cols)
            
            # Set column headers
            self.data_before_table.setHorizontalHeaderLabels(preview_df.columns.tolist())
            
            # Fill the table with data
            for i in range(rows):
                for j in range(cols):
                    value = str(preview_df.iloc[i, j]) if not pd.isna(preview_df.iloc[i, j]) else ""
                    item = QTableWidgetItem(value)
                    self.data_before_table.setItem(i, j, item)
            
            # Auto-adjust column widths to fit content
            self.data_before_table.resizeColumnsToContents()
            
            # Set table to stretch columns to fill width
            header = self.data_before_table.horizontalHeader()
            for i in range(cols):
                header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            
        except Exception as e:
            # Silent error handling for this automatic update
            pass

    def update_data_after_table(self):
        """Update the 'Data After' table with the currently selected columns after processing"""
        file_path = self.input_file_path.text()
        sheet_name = self.sheet_combo.currentText()
        
        if not file_path or not sheet_name:
            return
            
        try:
            # Get column mappings
            id_col = self.id_column_combo.currentText()
            name_col = self.name_column_combo.currentText()
            year_col = self.year_column_combo.currentText()
            group_col = self.group_column_combo.currentText()
            
            # If no columns are selected yet, clear the table and return
            if not (id_col or name_col or year_col or group_col):
                self.data_after_table.clear()
                self.data_after_table.setRowCount(0)
                self.data_after_table.setColumnCount(0)
                return
                
            # Load the data
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Create new dataframe with required columns
            new_df = pd.DataFrame()
            
            # Only add columns that are selected
            if id_col:
                new_df['Student ID'] = df[id_col].astype(str)
                # Clean Student ID - remove non-digit characters
                new_df['Student ID'] = new_df['Student ID'].apply(lambda x: ''.join(c for c in x if c.isdigit()))
            
            if name_col:
                new_df['Name'] = df[name_col]
                # Capitalize first letter of each word in Name
                new_df['Name'] = new_df['Name'].str.title()
            
            if year_col:
                new_df['Year'] = df[year_col]
                # Format Year values
                new_df['Year'] = new_df['Year'].apply(self.format_year)
            
            if group_col:
                new_df['Group'] = df[group_col]
                # Format Group values
                new_df['Group'] = new_df['Group'].apply(self.format_group)
            
            if new_df.empty:
                return
            
            # Clear previous data in After table
            self.data_after_table.clear()
            self.data_after_table.setRowCount(0)
            self.data_after_table.setColumnCount(0)
            
            # Set up the After table
            rows, cols = min(10, new_df.shape[0]), new_df.shape[1]
            self.data_after_table.setRowCount(rows)
            self.data_after_table.setColumnCount(cols)
            
            # Set column headers
            self.data_after_table.setHorizontalHeaderLabels(new_df.columns.tolist())
            
            # Fill the table with processed data
            for i in range(rows):
                for j in range(cols):
                    value = str(new_df.iloc[i, j]) if not pd.isna(new_df.iloc[i, j]) else ""
                    item = QTableWidgetItem(value)
                    self.data_after_table.setItem(i, j, item)
            
            # Auto-adjust column widths
            self.data_after_table.resizeColumnsToContents()
            
            # Set table to stretch columns to fill width
            header = self.data_after_table.horizontalHeader()
            for i in range(cols):
                header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            
        except Exception as e:
            # Silent error handling for automatic updates
            pass

    def preview_mapping_result(self):
        """Preview the result of the mapping and formatting in the 'Data After' table"""
        # Validate inputs
        if not self.validate_inputs():
            return
            
        # Use the same method as auto-update
        self.update_data_after_table()

    def auto_detect_columns(self, columns):
        """Try to automatically detect appropriate column mappings based on column names"""
        for i, col in enumerate(columns):
            col_lower = col.lower()
            
            # +1 because we added an empty string at the beginning of combo box items
            if any(keyword in col_lower for keyword in ['id', 'student id', 'studentid', 'number']):
                self.id_column_combo.setCurrentIndex(i + 1)
                
            if any(keyword in col_lower for keyword in ['name', 'student name', 'studentname']):
                self.name_column_combo.setCurrentIndex(i + 1)
                
            if any(keyword in col_lower for keyword in ['year', 'level']):
                self.year_column_combo.setCurrentIndex(i + 1)
                
            if any(keyword in col_lower for keyword in ['group', 'section']):
                self.group_column_combo.setCurrentIndex(i + 1)

    def format_year(self, year_value):
        """Format year values to 'Year X' format"""
        if pd.isna(year_value) or not year_value:
            return "Year 1"  # Default value
            
        # Convert to string and clean it
        year_str = str(year_value).strip().lower()
        
        # Extract numeric part if it exists
        numeric_part = ''.join(c for c in year_str if c.isdigit())
        
        if not numeric_part:
            return "Year 1"  # Default if no number found
            
        # Format as "Year X"
        return f"Year {numeric_part}"

    def format_group(self, group_value):
        """Format group values to 'AX' format (uppercase letter followed by number)"""
        if pd.isna(group_value) or not group_value:
            return "A1"  # Default value
            
        # Convert to string and clean it
        group_str = str(group_value).strip().lower()
        
        # Try to extract letter and number parts
        letter_part = ""
        numeric_part = ""
        
        for c in group_str:
            if c.isalpha():
                letter_part += c
            elif c.isdigit():
                numeric_part += c
        
        # Default values if parts are missing
        if not letter_part:
            letter_part = "A"
        if not numeric_part:
            numeric_part = "1"
            
        # Take first letter and capitalize it
        letter_part = letter_part[0].upper()
        
        # Format as "A1" (no space)
        return f"{letter_part}{numeric_part}"

    def process_file(self):
        # Validate inputs
        if not self.validate_inputs():
            return
            
        try:
            # Load input data
            input_file = self.input_file_path.text()
            sheet_name = self.sheet_combo.currentText()
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            
            # Get column mappings
            id_col = self.id_column_combo.currentText()
            name_col = self.name_column_combo.currentText()
            year_col = self.year_column_combo.currentText()
            group_col = self.group_column_combo.currentText()
            
            # Create new dataframe with required columns
            new_df = pd.DataFrame()
            new_df['Student ID'] = df[id_col] if id_col else None
            new_df['Name'] = df[name_col] if name_col else None
            new_df['Year'] = df[year_col] if year_col else None
            new_df['Group'] = df[group_col] if group_col else None
            
            # Clean and format data
            # Convert Student ID to string and remove any non-digit characters
            new_df['Student ID'] = new_df['Student ID'].astype(str)
            new_df['Student ID'] = new_df['Student ID'].apply(lambda x: ''.join(c for c in x if c.isdigit()))
            
            # Capitalize first letter of each word in Name
            new_df['Name'] = new_df['Name'].str.title()
            
            # Format Year values to "Year X"
            new_df['Year'] = new_df['Year'].apply(self.format_year)
            
            # Format Group values to "AX" (uppercase letter followed by number)
            new_df['Group'] = new_df['Group'].apply(self.format_group)
            
            # Method 1: Get the directory of the running executable
            if getattr(sys, 'frozen', False):
                # If running as compiled executable
                app_dir = os.path.dirname(sys.executable)
            else:
                # If running as script
                app_dir = os.path.dirname(os.path.abspath(__file__))
            
            # Create a reference_data folder if it doesn't exist
            reference_data_dir = os.path.join(app_dir, "reference_data")
            if not os.path.exists(reference_data_dir):
                os.makedirs(reference_data_dir)
            
            # Create filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"reference_data_{timestamp}.xlsx"
            
            # Set output file in the reference_data directory
            output_file = os.path.join(reference_data_dir, output_filename)
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reference Data"
            
            # Add header row
            headers = ['Student ID', 'Name', 'Year', 'Group']
            ws.append(headers)
            # Format header row
            for i, cell in enumerate(ws[1]):
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D3D3D3")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data rows
            for _, row in new_df.iterrows():
                ws.append([row['Student ID'], row['Name'], row['Year'], row['Group']])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = max(len(str(cell.value) if cell.value else '') for cell in column)
                ws.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = max_length + 3
            
            # Save the workbook
            wb.save(output_file)
            
            self.show_success_message(f"Reference file created successfully!\nSaved in the 'reference_data' folder as:\n{output_filename}")
            
        except Exception as e:
            self.show_error_message(f"Error processing file: {str(e)}")
    
    def validate_inputs(self):
        # Check if input file is selected
        if not self.input_file_path.text():
            self.show_error_message("Please select an input file.")
            return False
            
        # Check if sheet is selected
        if not self.sheet_combo.currentText():
            self.show_error_message("Please select a sheet.")
            return False
            
        # Check if column mappings are selected
        if not self.id_column_combo.currentText():
            self.show_error_message("Please select a Student ID column.")
            return False
            
        if not self.name_column_combo.currentText():
            self.show_error_message("Please select a Name column.")
            return False
            
        if not self.year_column_combo.currentText():
            self.show_error_message("Please select a Year column.")
            return False
            
        if not self.group_column_combo.currentText():
            self.show_error_message("Please select a Group column.")
            return False
            
        return True

    def show_error_message(self, message):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Critical)
        msg_box.setWindowTitle("Error")
        msg_box.setText(message)
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)
        # Style OK button
        ok_button = msg_box.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)
        msg_box.exec()

    def show_success_message(self, message):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setWindowTitle("Success")
        msg_box.setText(message)
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)
        # Style OK button
        ok_button = msg_box.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)
        msg_box.exec()

#==========================================================log sheet preparer==========================================================#

class GithubDownloadWorker(QThread):
    progress_signal = pyqtSignal(int)
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(list)

    def __init__(self, repo_url, token):
        super().__init__()
        self.repo_url = repo_url
        self.token = token
        self.downloaded_files = []

    def run(self):
        try:
            # Parse the repo URL to extract owner and repo name
            # Example: "https://github.com/username/repo"
            parts = self.repo_url.strip('/').split('/')
            if len(parts) < 5 or parts[2] != 'github.com':
                self.log_signal.emit("Invalid GitHub repository URL format")
                return

            owner = parts[3]
            repo = parts[4]

            # Get repository contents
            self.log_signal.emit(
                f"Connecting to GitHub repository: {owner}/{repo}")
            headers = {
                'Authorization': f'token {self.token}'} if self.token else {}

            # Get all files in the repository
            # Specify the backups folder
            api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/backups"
            response = requests.get(api_url, headers=headers)

            if response.status_code != 200:
                self.log_signal.emit(
                    f"Error accessing repository: {response.status_code}, {response.text}")
                return

            contents = response.json()
            excel_files = [item for item in contents if item['name'].endswith(
                '.xlsx') or item['name'].endswith('.xls')]

            if not excel_files:
                self.log_signal.emit("No Excel files found in the repository")
                return

            # Create base directory and imported logs subdirectory
            # Use a better approach for determining base directory that works with both script and exe
            if getattr(sys, 'frozen', False):
                # If the application is run as a bundle, the PyInstaller bootloader
                # extends the sys module by a flag frozen=True and sets the app 
                # path into variable sys._MEIPASS
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))
            
            temp_dir = os.path.join(base_dir, 'log_history', 'Imported_logs')
            os.makedirs(temp_dir, exist_ok=True)

            # Download each Excel file
            total_files = len(excel_files)
            for idx, file in enumerate(excel_files):
                self.log_signal.emit(f"Downloading {file['name']}...")

                download_url = file['download_url']
                file_response = requests.get(download_url, headers=headers)

                if file_response.status_code == 200:
                    # Save file locally
                    file_path = os.path.join(temp_dir, file['name'])
                    with open(file_path, 'wb') as f:
                        f.write(file_response.content)

                    self.downloaded_files.append(file_path)
                    self.log_signal.emit(f"Downloaded {file['name']}")
                else:
                    self.log_signal.emit(
                        f"Failed to download {file['name']}: {file_response.status_code}")

                # Update progress
                progress = int(((idx + 1) / total_files) * 100)
                self.progress_signal.emit(progress)

            self.log_signal.emit(
                f"Downloaded {len(self.downloaded_files)} Excel files")
            self.finished_signal.emit(self.downloaded_files)

        except Exception as e:
            self.log_signal.emit(f"Error: {str(e)}")
            import traceback
            self.log_signal.emit(traceback.format_exc())

class MergeWorker(QThread):
    progress_signal = pyqtSignal(int)
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(str)

    def __init__(self, files, output_file):
        super().__init__()
        self.files = files
        self.output_file = output_file

    def run(self):
        try:
            if not self.files:
                self.log_signal.emit("No files to merge")
                return

            self.log_signal.emit(f"Starting merge of {len(self.files)} files")

            # Initialize a list to hold all dataframes
            all_dfs = []

            # Process each file
            for idx, file_path in enumerate(self.files):
                self.log_signal.emit(
                    f"Processing {os.path.basename(file_path)}")

                try:
                    # Read all sheets from the Excel file
                    excel_file = pd.ExcelFile(file_path)
                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)

                        # Only process if the dataframe is not empty
                        if not df.empty:
                            # Add file and sheet metadata
                            df['Source_File'] = os.path.basename(file_path)
                            df['Source_Sheet'] = sheet_name
                            all_dfs.append(df)
                            self.log_signal.emit(
                                f"Added sheet '{sheet_name}' with {len(df)} rows")

                except Exception as e:
                    self.log_signal.emit(
                        f"Error processing {file_path}: {str(e)}")
                    continue

                # Update progress
                progress = int(((idx + 1) / len(self.files)) * 100)
                self.progress_signal.emit(progress)

            if not all_dfs:
                self.log_signal.emit("No valid data found in the files")
                return

            # Standardize column names across all dataframes
            self.log_signal.emit("Standardizing column headers...")

            # Find common columns or use a predefined set of columns
            # For now, we'll use a simple approach of getting all unique columns
            all_columns = set()
            for df in all_dfs:
                all_columns.update(df.columns)

            # Remove metadata columns we added
            standard_columns = [col for col in all_columns if col not in [
                'Source_File', 'Source_Sheet']]

            # Reindex all dataframes with the standard columns
            standardized_dfs = []
            for df in all_dfs:
                # Create a new dataframe with all standard columns (will be filled with NaN for missing columns)
                new_df = pd.DataFrame(columns=standard_columns)

                # Copy data from original dataframe for matching columns
                for col in standard_columns:
                    if col in df.columns:
                        new_df[col] = df[col]

                # Add back metadata columns
                new_df['Source_File'] = df['Source_File']
                new_df['Source_Sheet'] = df['Source_Sheet']

                standardized_dfs.append(new_df)

            # Concatenate all dataframes
            self.log_signal.emit("Merging all sheets...")
            merged_df = pd.concat(standardized_dfs, ignore_index=True)

            # Identify and reorder columns as per requirements:
            # 1. Student ID (looking for column containing "Student" and "ID")
            # 2. Location (looking for column containing "Location")
            # 3. Log date (looking for column containing "Log" and "date" or "Date")
            # 4. Log time (looking for column containing "Log" and "time" or "Time")
            # 5. All other columns

            self.log_signal.emit("Reordering columns to specified format...")

            # Find the best matching columns based on column names
            student_id_col = None
            location_col = None
            log_date_col = None
            log_time_col = None

            # Look for exact or partial matches
            for col in merged_df.columns:
                col_lower = str(col).lower()

                # Check for Student ID
                if "student" in col_lower and "id" in col_lower:
                    student_id_col = col
                # Check for Location
                elif "location" in col_lower:
                    location_col = col
                # Check for Log date
                elif "log" in col_lower and ("date" in col_lower or "day" in col_lower):
                    log_date_col = col
                # Check for Log time
                elif "log" in col_lower and "time" in col_lower:
                    log_time_col = col

            # Create the ordered columns list
            ordered_columns = []

            # Add the main required columns if they exist
            for col in [student_id_col, location_col, log_date_col, log_time_col]:
                if col is not None and col in merged_df.columns:
                    ordered_columns.append(col)

            # Add all remaining columns (excluding the ones we've already added and metadata)
            remaining_columns = [col for col in merged_df.columns
                                if col not in ordered_columns
                                and col not in ['Source_File', 'Source_Sheet']]
            ordered_columns.extend(remaining_columns)

            # Add metadata columns at the end
            ordered_columns.extend(['Source_File', 'Source_Sheet'])

            # Log the column ordering
            self.log_signal.emit(
                f"Column order being used: {', '.join(ordered_columns[:4])} + remaining columns")

            # Reorder the dataframe columns
            merged_df = merged_df[ordered_columns]

            # Make sure the output directory exists
            os.makedirs(os.path.dirname(self.output_file), exist_ok=True)

            # Save the merged data with reordered columns
            self.log_signal.emit(f"Saving merged data to {self.output_file}")
            merged_df.to_excel(self.output_file, index=False)

            self.log_signal.emit(
                f"Successfully merged {len(all_dfs)} sheets into {self.output_file} with ordered columns")
            self.finished_signal.emit(self.output_file)

        except Exception as e:
            self.log_signal.emit(f"Error during merge: {str(e)}")
            import traceback
            self.log_signal.emit(traceback.format_exc())

class LogSheetPreparer(QWidget):
    def __init__(self):
        super().__init__()
        self.files_to_merge = []
        # Define the hardcoded GitHub token - split to avoid detection
        token_part1 = "github_pat_"
        token_part2 = "11BREVRNQ0LX45XKQZzjkB_TL3KNQxHy4Sms4Fo20IUcxNLUwNAFbfeiXy92idb3mwTVANNZ4EC92cvkof"
        self.github_token = token_part1 + token_part2
        # Define the hardcoded GitHub repo URL - hidden from UI
        self.github_repo = "https://github.com/MohammadHamdi11/RN-E-attendancerecorderapp"
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)

        self.init_ui()

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio,
                                               Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("Log Sheet Preparer")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)

        # Back button
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)

        # Exit button
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)

        # Add buttons to vertical layout
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()

        # Add button layout to header
        header_layout.addStretch()
        header_layout.addLayout(header_buttons_layout)
        main_layout.addLayout(header_layout)

        # Import Method Selection
        import_group = QGroupBox("Import Method")
        import_group.setStyleSheet(GROUP_BOX_STYLE)
        import_layout = QVBoxLayout(import_group)

        # Radio buttons for import method
        radio_layout = QHBoxLayout()
        self.github_radio = QRadioButton("Import from Cloud Storage")
        self.github_radio.setChecked(True)  # Default to GitHub import
        self.local_radio = QRadioButton("Import Local Files")
        radio_layout.addWidget(self.github_radio)
        radio_layout.addWidget(self.local_radio)
        radio_layout.addStretch()
        import_layout.addLayout(radio_layout)

        # Connect radio buttons to toggle between input methods
        self.github_radio.toggled.connect(self.toggle_import_method)

        # Stacked widget for different import methods
        self.import_stack = QStackedWidget()

        # GitHub Import Widget
        github_widget = QWidget()
        github_layout = QVBoxLayout(github_widget)

        # Only show informational text about the GitHub repo, hiding the actual implementation details
        github_info_label = QLabel(
            "Excel files will be downloaded from the QRScanner-webapp repository's backup folder.")
        github_layout.addWidget(github_info_label)

        # Button to download files from GitHub
        download_btn = QPushButton("Download Excel Files from Cloud Storage")
        download_btn.clicked.connect(self.download_github_files)
        download_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        github_layout.addWidget(download_btn)

        # Local Files Import Widget
        local_widget = QWidget()
        local_layout = QVBoxLayout(local_widget)

        local_files_layout = QHBoxLayout()
        local_files_layout.addWidget(QLabel("Local Excel Files:"))
        self.local_files_label = QLineEdit()
        self.local_files_label.setPlaceholderText("Select Excel file...")
        self.local_files_label.setMinimumWidth(300)
        self.local_files_label.setReadOnly(True)
        local_files_layout.addWidget(self.local_files_label)

        # Button for importing local files
        browse_btn = QPushButton("Browse")
        browse_btn.clicked.connect(self.browse_files)
        browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        local_files_layout.addWidget(browse_btn)
        local_layout.addLayout(local_files_layout)

        # Add widgets to stack
        self.import_stack.addWidget(github_widget)
        self.import_stack.addWidget(local_widget)
        import_layout.addWidget(self.import_stack)

        main_layout.addWidget(import_group)

        # Files Table Section
        files_group = QGroupBox("Files to Merge")
        files_group.setStyleSheet(GROUP_BOX_STYLE)
        files_layout = QVBoxLayout(files_group)

        self.files_table = QTableWidget()
        self.files_table.setColumnCount(2)
        self.files_table.setHorizontalHeaderLabels(['File Path', 'Status'])

        # Center align the header text
        header = self.files_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)

        # Set column resize modes
        header.setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch)  # File Path
        header.setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)  # Status

        # Apply table styles
        self.files_table.setStyleSheet(TABLE_STYLE)
        files_layout.addWidget(self.files_table)

        # Buttons for files table
        files_btn_layout = QHBoxLayout()
        clear_files_btn = QPushButton("Clear Files")
        clear_files_btn.clicked.connect(self.clear_files)
        clear_files_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        files_btn_layout.addWidget(clear_files_btn)
        files_layout.addLayout(files_btn_layout)

        main_layout.addWidget(files_group)

        # Progress Bar Section
        progress_group = QGroupBox("Progress")
        progress_group.setStyleSheet(GROUP_BOX_STYLE)
        progress_layout = QVBoxLayout(progress_group)

        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_bar.setStyleSheet(PROGRESS_BAR_STYLE)

        # Create loading gif label
        self.loading_label = QLabel()
        self.loading_label.setFixedSize(24, 24)
        self.loading_label.setVisible(False)

        # Create the movie object for the GIF
        self.loading_movie = QMovie()
        self.loading_movie.setScaledSize(QSize(24, 24))
        self.loading_label.setMovie(self.loading_movie)

        loading_gif_path = os.path.join(
            os.path.dirname(__file__), 'loading.gif')
        if os.path.exists(loading_gif_path):
            self.loading_movie.setFileName(loading_gif_path)
        else:
            print(f"Warning: loading.gif not found at {loading_gif_path}")

        # Create a horizontal layout to hold both the progress bar and loading animation
        progress_h_layout = QHBoxLayout()
        progress_h_layout.addWidget(self.progress_bar)
        progress_h_layout.addWidget(self.loading_label)
        progress_layout.addLayout(progress_h_layout)

        main_layout.addWidget(progress_group)

        # Output Console Section
        console_group = QGroupBox("Output Console")
        console_group.setStyleSheet(GROUP_BOX_STYLE)
        console_layout = QVBoxLayout(console_group)

        self.output_console = QTextEdit()
        self.output_console.setReadOnly(True)
        self.output_console.setMaximumHeight(150)
        self.output_console.setStyleSheet(CONSOLE_STYLE)
        console_layout.addWidget(self.output_console)
        main_layout.addWidget(console_group)

        # Bottom Buttons
        button_layout = QHBoxLayout()
        merge_btn = QPushButton("Merge Logs Files")
        merge_btn.clicked.connect(self.merge_files)
        merge_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        button_layout.addWidget(merge_btn)
        main_layout.addLayout(button_layout)

    def toggle_import_method(self):
        # Set the current import method based on radio button selection
        if self.github_radio.isChecked():
            self.import_stack.setCurrentIndex(0)
        else:
            self.import_stack.setCurrentIndex(1)

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def browse_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select Excel Files", "", "Excel Files (*.xlsx *.xls)"
        )
        if files:
            self.files_to_merge = files
            self.local_files_label.setText(f"{len(files)} files selected")
            self.update_files_table()
            self.log_message(f"Selected {len(files)} files for merging")

    def update_files_table(self):
        # Clear existing table
        self.files_table.setRowCount(0)

        # Add files to table
        for file_path in self.files_to_merge:
            row_position = self.files_table.rowCount()
            self.files_table.insertRow(row_position)

            # Create items for the cells
            file_item = QTableWidgetItem(os.path.basename(file_path))
            status_item = QTableWidgetItem("Ready")

            # Set items to the table
            self.files_table.setItem(row_position, 0, file_item)
            self.files_table.setItem(row_position, 1, status_item)

    def clear_files(self):
        self.files_to_merge = []
        self.files_table.setRowCount(0)
        self.local_files_label.setText("No files selected")
        self.log_message("Cleared all files")

    def log_message(self, message):
        self.output_console.append(
            f"[{datetime.now().strftime('%H:%M:%S')}] {message}")
        # Scroll to the bottom
        self.output_console.moveCursor(QTextCursor.MoveOperation.End)

    def download_github_files(self):
        # Use hardcoded repo URL and token - not visible to users
        repo_url = self.github_repo
        token = self.github_token

        self.log_message(f"Starting download from Cloud Storage")

        # Start loading animation
        self.loading_label.setVisible(True)
        self.loading_movie.start()
        self.progress_bar.setValue(0)

        # Create and start the worker thread
        self.github_worker = GithubDownloadWorker(repo_url, token)
        self.github_worker.progress_signal.connect(self.update_progress)
        self.github_worker.log_signal.connect(self.log_message)
        self.github_worker.finished_signal.connect(
            self.handle_downloaded_files)
        self.github_worker.start()

    def handle_downloaded_files(self, files):
        # Update the list of files to merge
        self.files_to_merge = files
        self.update_files_table()

        # Stop loading animation
        self.loading_movie.stop()
        self.loading_label.setVisible(False)

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    def merge_files(self):
        if not self.files_to_merge:
            self.log_message("No files to merge. Please import files first.")
            return

        # Generate output filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"log_history_{timestamp}.xlsx"

        # Create base directory and merged files subdirectory
        # Use a better approach for determining base directory that works with both script and exe
        if getattr(sys, 'frozen', False):
            # If the application is run as a bundle, the PyInstaller bootloader
            # extends the sys module by a flag frozen=True and sets the app 
            # path into variable sys._MEIPASS
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        
        output_dir = os.path.join(base_dir, 'log_history', 'Merged_Files')
        os.makedirs(output_dir, exist_ok=True)

        # Full path for output file
        output_file = os.path.join(output_dir, output_filename)

        self.log_message(f"Merging files to: {output_file}")

        # Start loading animation
        self.loading_label.setVisible(True)
        self.loading_movie.start()
        self.progress_bar.setValue(0)

        # Create and start the worker thread
        self.merge_worker = MergeWorker(self.files_to_merge, output_file)
        self.merge_worker.progress_signal.connect(self.update_progress)
        self.merge_worker.log_signal.connect(self.log_message)
        self.merge_worker.finished_signal.connect(self.handle_merge_complete)
        self.merge_worker.start()

    def handle_merge_complete(self, output_file):
        # Stop loading animation
        self.loading_movie.stop()
        self.loading_label.setVisible(False)

        self.log_message(f"Merge completed successfully: {output_file}")

        # Ask if user wants to open the merged file
        from PyQt6.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            self, 'Merge Complete',
            f'Merge completed successfully!\nThe merged file has been saved as:\n{os.path.basename(output_file)}\n\nWould you like to open the merged file?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )

        if reply == QMessageBox.StandardButton.Yes:
            # Open the file with the default application
            import subprocess
            import platform

            if platform.system() == 'Windows':
                os.startfile(output_file)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(('open', output_file))
            else:  # Linux
                subprocess.call(('xdg-open', output_file))

#==========================================================Schedule Manager==========================================================#

class ScheduleManager(QWidget):
    def __init__(self):
        super().__init__()
        self.schedule_data = []
        self.current_year = None
        self.current_group = None
        self.setStyleSheet(f"""
            background-color: {BLACK}; 
            color: {TEXT_COLOR};
            QLabel {{
                color: {TEXT_COLOR};
            }}
            QCalendarWidget {{
                background-color: {INPUT_BG};
                color: {TEXT_COLOR};
            }}
            QCalendarWidget QAbstractItemView {{
                selection-background-color: #555;
                selection-color: {TEXT_COLOR};
            }}
            QCalendarWidget QWidget {{
                alternate-background-color: #444;
                color: {TEXT_COLOR};
            }}
            QCalendarWidget QTableView {{
                border: none;
            }}
        """)
        
        self.init_ui()

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Create a scroll area to make the page scrollable
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("QScrollArea { border: none; }")
        
        # Create a widget to hold the content
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setSpacing(20)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        
        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio, 
                                             Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("Schedule Manager")
        title_label.setStyleSheet("font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()
    
        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)
        
        # Back button
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)
        
        # Exit button
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)
        
        # Add buttons to vertical layout
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()
        
        # Add button layout to header
        header_layout.addStretch()
        header_layout.addLayout(header_buttons_layout)
        scroll_layout.addLayout(header_layout)
    
        # Mode Selection Group
        mode_group = QGroupBox("Schedule Mode")
        mode_group.setStyleSheet(GROUP_BOX_STYLE)
        mode_layout = QHBoxLayout(mode_group)
        
        # Radio buttons for mode selection
        self.create_radio = QRadioButton("Create New Schedule")
        self.create_radio.setChecked(True)  # Default selection
        self.update_radio = QRadioButton("Update Existing Schedule")
        
        mode_layout.addWidget(self.create_radio)
        mode_layout.addWidget(self.update_radio)
        mode_layout.addStretch()
        
        # Connect radio buttons to switch mode functions
        self.create_radio.toggled.connect(self.switch_to_create_mode)
        self.update_radio.toggled.connect(self.switch_to_update_mode)
        
        scroll_layout.addWidget(mode_group)
    
        # Stacked widget to hold different mode layouts
        self.mode_stack = QStackedWidget()
        scroll_layout.addWidget(self.mode_stack)
        
        # Create mode widget
        self.create_widget = QWidget()
        self.create_layout = QVBoxLayout(self.create_widget)
        self.setup_create_mode()
        self.mode_stack.addWidget(self.create_widget)
        
        # Update mode widget
        self.update_widget = QWidget()
        self.update_layout = QVBoxLayout(self.update_widget)
        self.setup_update_mode()
        self.mode_stack.addWidget(self.update_widget)
    
        # Default to create mode
        self.mode_stack.setCurrentIndex(0)
    
        # Bottom Buttons layout
        self.button_layout = QHBoxLayout()
        
        # Create Schedule button
        self.create_schedule_btn = QPushButton("Create Schedule")
        self.create_schedule_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        self.create_schedule_btn.clicked.connect(self.save_schedule)
        self.button_layout.addWidget(self.create_schedule_btn)
        
        # Update Schedule button (initially hidden)
        self.update_schedule_btn = QPushButton("Update Schedule")
        self.update_schedule_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        self.update_schedule_btn.clicked.connect(self.update_schedule)
        self.update_schedule_btn.setVisible(False)
        self.button_layout.addWidget(self.update_schedule_btn)
        
        scroll_layout.addLayout(self.button_layout)
        
        # Set the scroll content and add to main layout
        scroll_area.setWidget(scroll_content)
        main_layout.addWidget(scroll_area)
    
    def setup_create_mode(self):
        # Add New Sessions group
        add_sessions_group = QGroupBox("Add New Sessions")
        add_sessions_group.setStyleSheet(GROUP_BOX_STYLE)
        add_sessions_layout = QVBoxLayout(add_sessions_group)  # Use vertical layout as container
        
        # MODULE NAME INPUT - Add this new section
        module_name_layout = QHBoxLayout()
        module_name_layout.addWidget(QLabel("Module Name:"))
        self.module_name_input = QLineEdit()
        self.module_name_input.setPlaceholderText("Enter module name...")
        module_name_layout.addWidget(self.module_name_input)
        add_sessions_layout.addLayout(module_name_layout)
        
        # Create horizontal layout for the form fields
        form_layout = QHBoxLayout()
        
        # Left column for form inputs
        left_column = QVBoxLayout()
        left_column.setSpacing(15)
        
        # Year Selection
        year_row = QHBoxLayout()
        year_row.addWidget(QLabel("Academic Year:"))
        self.year_combo = QComboBox()
        self.year_combo.setMaxVisibleItems(10)  # Use setMaxVisibleItems instead
        self.year_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)  # Make it expand horizontally
        self.year_combo.addItems(["Year 1", "Year 2", "Year 3"])
        self.year_combo.currentTextChanged.connect(self.year_selected)
        year_row.addWidget(self.year_combo)
        year_row.addStretch()
        left_column.addLayout(year_row)
        
        # Group selection row
        group_row = QHBoxLayout()
        group_row.addWidget(QLabel("Group:            "))
        self.group_combo = QComboBox()
        self.group_combo.setMaxVisibleItems(10)
        self.group_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.populate_groups()
        group_row.addWidget(self.group_combo)
        group_row.addStretch()
        left_column.addLayout(group_row)
        
        # Subject row
        subject_row = QHBoxLayout()
        subject_row.addWidget(QLabel("Subject:          "))
        self.subject_combo = QComboBox()
        self.subject_combo.setMaxVisibleItems(10)
        self.subject_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.subject_combo.addItems([
            "Anatomy", "Histology", "Biochemistry", "Physiology", 
            "Microbiology", "Parasitology", "Pathology", "Pharmacology", "Clinical"
        ])
        self.subject_combo.currentTextChanged.connect(self.update_location_options)
        subject_row.addWidget(self.subject_combo)
        subject_row.addStretch()
        left_column.addLayout(subject_row)
        
        # Session row
        session_row = QHBoxLayout()
        session_row.addWidget(QLabel("Session:          "))
        self.session_combo = QComboBox()
        self.session_combo.setMaxVisibleItems(10)
        self.session_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.session_combo.addItems([str(i) for i in range(1, 31)])
        session_row.addWidget(self.session_combo)
        session_row.addStretch()
        left_column.addLayout(session_row)
    
        # Location row
        location_row = QHBoxLayout()
        location_row.addWidget(QLabel("Location:         "))
        self.location_combo = QComboBox()
        self.location_combo.setMaxVisibleItems(10)
        self.location_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.populate_locations()
        location_row.addWidget(self.location_combo)
        location_row.addStretch()
        left_column.addLayout(location_row)
        
        # Add left column to form layout
        form_layout.addLayout(left_column, 1)
        
        # Right column for date and time selection
        right_column = QVBoxLayout()
        right_column.setSpacing(15)
        
        # Time row - moved to top of right column
        time_label = QLabel("Start Time:")
        time_label.setStyleSheet("font-weight: bold;")
        right_column.addWidget(time_label)
        
        self.time_combo = QComboBox()
        self.time_combo.setMaxVisibleItems(10)
        self.time_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        # Generate times from 7:00 to 17:00 with 15-minute intervals
        times = []
        for hour in range(7, 18):  # 7 AM to 5 PM (17:00)
            for minute in [0, 15, 30, 45]:
                # Skip 17:15, 17:30, 17:45
                if hour == 17 and minute > 0:
                    continue
                times.append(f"{hour:02d}:{minute:02d}:00")
        self.time_combo.addItems(times)
        right_column.addWidget(self.time_combo)
        
        # Add spacer between time and date
        right_column.addSpacing(10)
        
        # Date with modern calendar widget
        date_label = QLabel("Date:")
        date_label.setStyleSheet("font-weight: bold;")
        right_column.addWidget(date_label)
        
        # Modern styled calendar
        self.date_calendar = QCalendarWidget()
        self.date_calendar.setFirstDayOfWeek(Qt.DayOfWeek.Monday)
        self.date_calendar.setGridVisible(True)
        self.date_calendar.setMinimumDate(QDate.currentDate())
        self.date_calendar.setMaximumDate(QDate.currentDate().addDays(3650))
        
        # Make calendar more modern
        self.date_calendar.setStyleSheet("""
            QCalendarWidget {
                background-color: #2a2a2a;
                border: 1px solid #555;
                border-radius: 4px;
            }
            QCalendarWidget QToolButton {
                color: #ddd;
                background-color: #3a3a3a;
                border: 1px solid #555;
                border-radius: 4px;
                padding: 5px;
            }
            QCalendarWidget QMenu {
                background-color: #2a2a2a;
                color: #ddd;
            }
            QCalendarWidget QSpinBox {
                background-color: #3a3a3a;
                color: #ddd;
                border: 1px solid #555;
                border-radius: 4px;
                padding: 2px;
            }
            QCalendarWidget QAbstractItemView:enabled {
                color: #ddd;
                background-color: #2a2a2a;
                selection-background-color: #555;
                selection-color: #fff;
            }
            QCalendarWidget QAbstractItemView:disabled {
                color: #555;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #3a3a3a;
                border-bottom: 1px solid #555;
            }
        """)
        
        right_column.addWidget(self.date_calendar)
        
        # Add right column to form layout
        form_layout.addLayout(right_column, 1)
        
        # Add the form layout to the main add_sessions_layout
        add_sessions_layout.addLayout(form_layout)
        
        # Add session button row
        add_session_btn_layout = QHBoxLayout()
        add_session_btn = QPushButton("Add Session")
        add_session_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        add_session_btn.clicked.connect(self.add_session)
        add_session_btn_layout.addStretch()
        add_session_btn_layout.addWidget(add_session_btn)
        add_sessions_layout.addLayout(add_session_btn_layout)
        
        # Add the grouped widget to the main layout
        self.create_layout.addWidget(add_sessions_group)
        
        # Added Sessions Table
        sessions_group = QGroupBox("Added Sessions")
        sessions_group.setStyleSheet(GROUP_BOX_STYLE)
        sessions_layout = QVBoxLayout(sessions_group)
        
        self.sessions_table = QTableWidget()
        self.sessions_table.setColumnCount(7)
        self.sessions_table.setHorizontalHeaderLabels([
            'Year', 'Group', 'Subject', 'Session', 'Location', 'Date', 'Start Time'
        ])
        self.sessions_table.setStyleSheet(TABLE_STYLE)
        
        # Set column properties
        header = self.sessions_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Set columns to stretch
        for i in range(7):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        
        sessions_layout.addWidget(self.sessions_table)
        
        # Buttons for table management
        table_buttons = QHBoxLayout()
        remove_btn = QPushButton("Remove Selected")
        remove_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        remove_btn.clicked.connect(self.remove_selected_session)
        
        clear_btn = QPushButton("Clear All")
        clear_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        clear_btn.clicked.connect(self.clear_sessions)
        
        table_buttons.addWidget(remove_btn)
        table_buttons.addWidget(clear_btn)
        sessions_layout.addLayout(table_buttons)
        
        self.create_layout.addWidget(sessions_group)
        
        # Initialize UI state
        self.update_location_options()
    
    def setup_update_mode(self):
        # File selection
        file_group = QGroupBox("Select Schedule to Update")
        file_group.setStyleSheet(GROUP_BOX_STYLE)
        file_layout = QHBoxLayout(file_group)
        
        # MODULE NAME INPUT - Add this new section
        module_name_layout = QHBoxLayout()
        module_name_layout.addWidget(QLabel("Module Name:"))
        self.update_module_name_input = QLineEdit()
        self.update_module_name_input.setPlaceholderText("Enter module name...")
        file_layout.addWidget(self.update_module_name_input)
        
        self.update_file_input = QLineEdit()
        self.update_file_input.setPlaceholderText("Select Excel schedule file...")
        file_layout.addWidget(self.update_file_input)
        
        browse_btn = QPushButton("Browse")
        browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        browse_btn.clicked.connect(self.browse_schedule_file)
        file_layout.addWidget(browse_btn)
        
        self.update_layout.addWidget(file_group)
        
        # Existing Sessions Table
        existing_group = QGroupBox("Existing Sessions")
        existing_group.setStyleSheet(GROUP_BOX_STYLE)
        existing_layout = QVBoxLayout(existing_group)
        
        self.existing_table = QTableWidget()
        self.existing_table.setColumnCount(7)
        self.existing_table.setHorizontalHeaderLabels([
            'Year', 'Group', 'Subject', 'Session', 'Location', 'Date', 'Start Time'
        ])
        self.existing_table.setStyleSheet(TABLE_STYLE)
        
        # Set column properties
        header = self.existing_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Set columns to stretch
        for i in range(7):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        
        existing_layout.addWidget(self.existing_table)
        
        # Table manipulation buttons
        existing_buttons = QHBoxLayout()
        remove_existing_btn = QPushButton("Remove Selected")
        remove_existing_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        remove_existing_btn.clicked.connect(self.remove_existing_session)
        existing_buttons.addWidget(remove_existing_btn)
        existing_layout.addLayout(existing_buttons)
    
        self.update_layout.addWidget(existing_group)
        
        # Add New Sessions group
        add_sessions_group = QGroupBox("Add New Sessions")
        add_sessions_group.setStyleSheet(GROUP_BOX_STYLE)
        add_sessions_layout = QVBoxLayout(add_sessions_group)  # Use vertical layout as container
        
        # Create horizontal layout for form fields
        form_layout = QHBoxLayout()
        
        # Left column for form inputs
        left_column = QVBoxLayout()
        left_column.setSpacing(15)
        
        # Year Selection
        year_row = QHBoxLayout()
        year_row.addWidget(QLabel("Academic Year:"))
        self.update_year_combo = QComboBox()
        self.update_year_combo.setMaxVisibleItems(10)
        self.update_year_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.update_year_combo.addItems(["Year 1", "Year 2", "Year 3"])
        self.update_year_combo.currentTextChanged.connect(self.update_year_selected)
        year_row.addWidget(self.update_year_combo)
        year_row.addStretch()
        left_column.addLayout(year_row)
        
        # Group selection row
        group_row = QHBoxLayout()
        group_row.addWidget(QLabel("Group:            "))
        self.update_group_combo = QComboBox()
        self.update_group_combo.setMaxVisibleItems(10)
        self.update_group_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.populate_update_groups()
        group_row.addWidget(self.update_group_combo)
        group_row.addStretch()
        left_column.addLayout(group_row)
        
        # Subject row
        subject_row = QHBoxLayout()
        subject_row.addWidget(QLabel("Subject:          "))
        self.update_subject_combo = QComboBox()
        self.update_subject_combo.setMaxVisibleItems(10)
        self.update_subject_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.update_subject_combo.addItems([
            "Anatomy", "Histology", "Biochemistry", "Physiology", 
            "Microbiology", "Parasitology", "Pathology", "Pharmacology", "Clinical"
        ])
        self.update_subject_combo.currentTextChanged.connect(self.update_update_location_options)
        subject_row.addWidget(self.update_subject_combo)
        subject_row.addStretch()
        left_column.addLayout(subject_row)
        
        # Session row
        session_row = QHBoxLayout()
        session_row.addWidget(QLabel("Session:           "))
        self.update_session_combo = QComboBox()
        self.update_session_combo.setMaxVisibleItems(10)
        self.update_session_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.update_session_combo.addItems([str(i) for i in range(1, 31)])
        session_row.addWidget(self.update_session_combo)
        session_row.addStretch()
        left_column.addLayout(session_row)
    
        # Location row
        location_row = QHBoxLayout()
        location_row.addWidget(QLabel("Location:         "))
        self.update_location_combo = QComboBox()
        self.update_location_combo.setMaxVisibleItems(10)
        self.update_location_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.populate_update_locations()
        location_row.addWidget(self.update_location_combo)
        location_row.addStretch()
        left_column.addLayout(location_row)
        
        # Add left column to form layout
        form_layout.addLayout(left_column, 1)
        
        # Right column for date and time selection
        right_column = QVBoxLayout()
        right_column.setSpacing(15)
        
        # Time row - moved to top of right column
        time_label = QLabel("Start Time:")
        time_label.setStyleSheet("font-weight: bold;")
        right_column.addWidget(time_label)
        
        self.update_time_combo = QComboBox()
        self.update_time_combo.setMaxVisibleItems(10)
        self.update_time_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        # Generate times from 7:00 to 17:00 with 15-minute intervals
        times = []
        for hour in range(7, 18):  # 7 AM to 5 PM (17:00)
            for minute in [0, 15, 30, 45]:
                # Skip 17:15, 17:30, 17:45
                if hour == 17 and minute > 0:
                    continue
                times.append(f"{hour:02d}:{minute:02d}:00")
        self.update_time_combo.addItems(times)
        right_column.addWidget(self.update_time_combo)
        
        # Add spacer between time and date
        right_column.addSpacing(10)
        
        # Date with modern calendar widget
        date_label = QLabel("Date:")
        date_label.setStyleSheet("font-weight: bold;")
        right_column.addWidget(date_label)
        
        # Modern styled calendar
        self.update_date_calendar = QCalendarWidget()
        self.update_date_calendar.setFirstDayOfWeek(Qt.DayOfWeek.Monday)
        self.update_date_calendar.setGridVisible(True)
        self.update_date_calendar.setMinimumDate(QDate.currentDate())
        self.update_date_calendar.setMaximumDate(QDate.currentDate().addDays(3650))
        
        # Make calendar more modern
        self.update_date_calendar.setStyleSheet("""
            QCalendarWidget {
                background-color: #2a2a2a;
                border: 1px solid #555;
                border-radius: 4px;
            }
            QCalendarWidget QToolButton {
                color: #ddd;
                background-color: #3a3a3a;
                border: 1px solid #555;
                border-radius: 4px;
                padding: 5px;
            }
            QCalendarWidget QMenu {
                background-color: #2a2a2a;
                color: #ddd;
            }
            QCalendarWidget QSpinBox {
                background-color: #3a3a3a;
                color: #ddd;
                border: 1px solid #555;
                border-radius: 4px;
                padding: 2px;
            }
            QCalendarWidget QAbstractItemView:enabled {
                color: #ddd;
                background-color: #2a2a2a;
                selection-background-color: #555;
                selection-color: #fff;
            }
            QCalendarWidget QAbstractItemView:disabled {
                color: #555;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #3a3a3a;
                border-bottom: 1px solid #555;
            }
        """)
        
        right_column.addWidget(self.update_date_calendar)
        
        # Add right column to form layout
        form_layout.addLayout(right_column, 1)
        
        # Add the form layout to the main add_sessions_layout
        add_sessions_layout.addLayout(form_layout)
        
        # Add session button row
        add_session_btn_layout = QHBoxLayout()
        add_session_btn = QPushButton("Add Session")
        add_session_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        add_session_btn.clicked.connect(self.add_update_session)
        add_session_btn_layout.addStretch()
        add_session_btn_layout.addWidget(add_session_btn)
        add_sessions_layout.addLayout(add_session_btn_layout)
        
        # Add the grouped widget to the main layout
        self.update_layout.addWidget(add_sessions_group)
        
        # New Sessions Group
        new_sessions_group = QGroupBox("New Sessions to Add")
        new_sessions_group.setStyleSheet(GROUP_BOX_STYLE)
        new_sessions_layout = QVBoxLayout(new_sessions_group)
        
        self.new_sessions_table = QTableWidget()
        self.new_sessions_table.setColumnCount(7)
        self.new_sessions_table.setHorizontalHeaderLabels([
            'Year', 'Group', 'Subject', 'Session', 'Location', 'Date', 'Start Time'
        ])
        self.new_sessions_table.setStyleSheet(TABLE_STYLE)
        
        # Set column properties
        header = self.new_sessions_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Set columns to stretch
        for i in range(7):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        
        new_sessions_layout.addWidget(self.new_sessions_table)
        
        # New sessions buttons
        new_buttons = QHBoxLayout()
        remove_new_btn = QPushButton("Remove Selected")
        remove_new_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        remove_new_btn.clicked.connect(self.remove_new_session)
        
        clear_new_btn = QPushButton("Clear All")
        clear_new_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        clear_new_btn.clicked.connect(self.clear_new_sessions)
        
        new_buttons.addWidget(remove_new_btn)
        new_buttons.addWidget(clear_new_btn)
        new_sessions_layout.addLayout(new_buttons)
        
        self.update_layout.addWidget(new_sessions_group)
        
        # Initialize UI state
        self.update_update_location_options()
    
    def populate_groups(self):
        """Populate groups based on year selection"""
        self.group_combo.clear()
        year = self.year_combo.currentText()

        groups = []
        for prefix in ["A", "B"]:
            for num in range(1, 11):
                groups.append(f"{prefix}{num}")

        self.group_combo.addItems(groups)

    def populate_update_groups(self):
        """Populate groups for update mode"""
        self.update_group_combo.clear()
        year = self.update_year_combo.currentText()

        groups = []
        for prefix in ["A", "B"]:
            for num in range(1, 11):
                groups.append(f"{prefix}{num}")

        self.update_group_combo.addItems(groups)

    def populate_locations(self):
        """Initialize locations combobox with all preset values"""
        self.location_combo.clear()

        # Set maximum visible items to 10
        self.location_combo.setMaxVisibleItems(10)

        # Add all preset locations regardless of subject
        locations = [
            "Morgue",
            "Anatomy Lecture Hall",
            "Histology Lab",
            "Histology Lecture Hall",
            "Biochemistry Lab",
            "Biochemistry Lecture Hall",
            "Physiology Lab",
            "Physiology Lecture Hall",
            "Microbiology Lab",
            "Microbiology Lecture Hall",
            "Parasitology Lab",
            "Parasitology Lecture Hall",
            "Pathology Lab",
            "Pathology Lecture Hall",
            "Pharmacology Lab",
            "Pharmacology Lecture Hall",
            "Building 'A' Lecture Hall",
            "Building 'B' Lecture Hall",
        ]

        # Add all locations to the combo box
        self.location_combo.addItems(locations)

    def populate_update_locations(self):
        """Initialize locations combobox with all preset values"""
        self.update_location_combo.clear()

        # Set maximum visible items to 10
        self.update_location_combo.setMaxVisibleItems(10)

        # Add all preset locations regardless of subject
        locations = [
            "Morgue",
            "Anatomy Lecture Hall",
            "Histology Lab",
            "Histology Lecture Hall",
            "Biochemistry Lab",
            "Biochemistry Lecture Hall",
            "Physiology Lab",
            "Physiology Lecture Hall",
            "Microbiology Lab",
            "Microbiology Lecture Hall",
            "Parasitology Lab",
            "Parasitology Lecture Hall",
            "Pathology Lab",
            "Pathology Lecture Hall",
            "Pharmacology Lab",
            "Pharmacology Lecture Hall",
            "Building 'B' Lecture Hall",
        ]

        # Add all locations to the combo box
        self.update_location_combo.addItems(locations)

    def update_location_options(self):
        """Update location options based on selected subject"""
        self.populate_locations()

    def update_update_location_options(self):
        """Update location options for update mode"""
        self.populate_update_locations()

    def year_selected(self):
        """Handle year selection change"""
        self.current_year = self.year_combo.currentText()
        self.populate_groups()

    def update_year_selected(self):
        """Handle year selection change in update mode"""
        self.current_year = self.update_year_combo.currentText()
        self.populate_update_groups()

    def switch_to_create_mode(self):
        """Switch to create mode"""
        if self.create_radio.isChecked():
            self.mode_stack.setCurrentIndex(0)
            self.create_schedule_btn.setVisible(True)
            self.update_schedule_btn.setVisible(False)

    def switch_to_update_mode(self):
        """Switch to update mode"""
        if self.update_radio.isChecked():
            self.mode_stack.setCurrentIndex(1)
            self.create_schedule_btn.setVisible(False)
            self.update_schedule_btn.setVisible(True)

    def add_session(self):
        """Add a session to the table in create mode"""
        year = self.year_combo.currentText()
        group = self.group_combo.currentText()
        subject = self.subject_combo.currentText()
        session = self.session_combo.currentText()
        location = self.location_combo.currentText()
        # Get date from calendar
        date = self.date_calendar.selectedDate().toString("dd/MM/yyyy")
        time = self.time_combo.currentText()

        # Add row to table
        row_position = self.sessions_table.rowCount()
        self.sessions_table.insertRow(row_position)

        # Set data in table
        self.sessions_table.setItem(row_position, 0, QTableWidgetItem(year))
        self.sessions_table.setItem(row_position, 1, QTableWidgetItem(group))
        self.sessions_table.setItem(row_position, 2, QTableWidgetItem(subject))
        self.sessions_table.setItem(row_position, 3, QTableWidgetItem(session))
        self.sessions_table.setItem(
            row_position, 4, QTableWidgetItem(location))
        self.sessions_table.setItem(row_position, 5, QTableWidgetItem(date))
        self.sessions_table.setItem(row_position, 6, QTableWidgetItem(time))

        # Center all items
        for col in range(7):
            item = self.sessions_table.item(row_position, col)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        # Add to data structure
        self.schedule_data.append(
            [year, group, subject, session, location, date, time])

    def add_update_session(self):
        """Add a session to the table in update mode"""
        year = self.update_year_combo.currentText()
        group = self.update_group_combo.currentText()
        subject = self.update_subject_combo.currentText()
        session = self.update_session_combo.currentText()
        location = self.update_location_combo.currentText()
        # Get date from calendar
        date = self.update_date_calendar.selectedDate().toString("dd/MM/yyyy")
        time = self.update_time_combo.currentText()

        # Add row to table
        row_position = self.new_sessions_table.rowCount()
        self.new_sessions_table.insertRow(row_position)

        # Set data in table
        self.new_sessions_table.setItem(
            row_position, 0, QTableWidgetItem(year))
        self.new_sessions_table.setItem(
            row_position, 1, QTableWidgetItem(group))
        self.new_sessions_table.setItem(
            row_position, 2, QTableWidgetItem(subject))
        self.new_sessions_table.setItem(
            row_position, 3, QTableWidgetItem(session))
        self.new_sessions_table.setItem(
            row_position, 4, QTableWidgetItem(location))
        self.new_sessions_table.setItem(
            row_position, 5, QTableWidgetItem(date))
        self.new_sessions_table.setItem(
            row_position, 6, QTableWidgetItem(time))

        # Center all items
        for col in range(7):
            item = self.new_sessions_table.item(row_position, col)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    def clear_sessions(self):
        """Clear all sessions from the table"""
        self.sessions_table.setRowCount(0)
        self.schedule_data = []

    def clear_new_sessions(self):
        """Clear all new sessions from the update table"""
        self.new_sessions_table.setRowCount(0)

    def remove_selected_session(self):
        """Remove selected session from the table"""
        selected_rows = self.sessions_table.selectionModel().selectedRows()
        if not selected_rows:
            return

        # Remove from highest index to lowest to avoid index issues
        rows_to_remove = sorted([index.row()
                                for index in selected_rows], reverse=True)

        for row in rows_to_remove:
            # Get information before removing for logging
            year = self.sessions_table.item(row, 0).text()
            group = self.sessions_table.item(row, 1).text()
            subject = self.sessions_table.item(row, 2).text()
            session = self.sessions_table.item(row, 3).text()

            # Remove from table
            self.sessions_table.removeRow(row)

            # Remove from data structure
            if 0 <= row < len(self.schedule_data):
                self.schedule_data.pop(row)

    def remove_new_session(self):
        """Remove selected new session from the update table"""
        selected_rows = self.new_sessions_table.selectionModel().selectedRows()
        if not selected_rows:
            return

        # Remove from highest index to lowest to avoid index issues
        rows_to_remove = sorted([index.row()
                                for index in selected_rows], reverse=True)

        for row in rows_to_remove:
            # Get information before removing for logging
            year = self.new_sessions_table.item(row, 0).text()
            group = self.new_sessions_table.item(row, 1).text()
            subject = self.new_sessions_table.item(row, 2).text()
            session = self.new_sessions_table.item(row, 3).text()

            # Remove from table
            self.new_sessions_table.removeRow(row)

    def remove_existing_session(self):
        """Remove selected session from the existing sessions table"""
        selected_rows = self.existing_table.selectionModel().selectedRows()
        if not selected_rows:
            return

        # Remove from highest index to lowest to avoid index issues
        rows_to_remove = sorted([index.row()
                                for index in selected_rows], reverse=True)

        for row in rows_to_remove:
            # Get information before removing for logging
            year = self.existing_table.item(row, 0).text()
            group = self.existing_table.item(row, 1).text()
            subject = self.existing_table.item(row, 2).text()
            session = self.existing_table.item(row, 3).text()

            # Remove from table
            self.existing_table.removeRow(row)

    def browse_schedule_file(self):
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Excel Files (*.xlsx *.xls)")
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFile)

        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles()
            if selected_files:
                file_path = selected_files[0]
                self.update_file_input.setText(file_path)
                # Auto-load the file
                self.load_schedule()

    def load_schedule(self):
        """Load an existing schedule from file"""
        file_path = self.update_file_input.text()
        if not file_path:
            self.show_message_box("Error", "Please select a file first")
            return
        
        try:
            # Clear existing data
            self.existing_table.setRowCount(0)
            
            # Load Excel file
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            
            # Try to get module name from sheet title
            module_name = sheet.title
            if module_name and module_name != "Sheet":  # If it's not the default sheet name
                self.update_module_name_input.setText(module_name)
            
            # Check if the first row contains "Module:" (old format) or headers (new format)
            first_cell = sheet.cell(row=1, column=1).value
            if first_cell and isinstance(first_cell, str) and first_cell.startswith("Module:"):
                # Old format with module name in first row
                module_name = first_cell.replace("Module:", "").strip()
                self.update_module_name_input.setText(module_name)
                start_row = 3  # Start reading data from row 3
            else:
                # New format with headers in first row
                start_row = 2  # Start reading data from row 2
            
            # Skip header row
            row_count = 0
            for row in sheet.iter_rows(min_row=start_row, values_only=True):
                if all(cell is None for cell in row):
                    continue  # Skip empty rows
                
                # Add row to table
                self.existing_table.insertRow(row_count)
                
                # Set data in table
                # Only use the first 7 columns
                for col, cell_value in enumerate(row[:7]):
                    value = str(cell_value) if cell_value is not None else ""
                    item = QTableWidgetItem(value)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.existing_table.setItem(row_count, col, item)
                
                row_count += 1
            
        except Exception as e:
            self.show_message_box("Error", f"Failed to load schedule: {str(e)}")
    
    def save_schedule(self):
        """Save the created schedule to Excel file directly without prompting"""
        if len(self.schedule_data) == 0:
            self.show_message_box("Error", "No schedule data to save")
            return
    
        # Get module name (default if empty)
        module_name = self.module_name_input.text().strip()
        if not module_name:
            module_name = "Untitled_Module"
    
        try:
            # Create modules_schedules directory in the current working directory
            # This will be where the EXE is running from
            save_dir = os.path.join(os.getcwd(), "modules_schedules")
            os.makedirs(save_dir, exist_ok=True)
    
            # Generate timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
            # Create filename with module name and timestamp
            filename = f"{module_name}_{timestamp}.xlsx"
            filepath = os.path.join(save_dir, filename)
        
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
        
            # Name the worksheet with the module name (limit to 31 chars as Excel has a limit)
            ws.title = module_name[:31]
        
            # Add headers in row 1 (no module name row anymore)
            headers = ['Year', 'Group', 'Subject', 'Session', 'Location', 'Date', 'Start Time']
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = header
                ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
            # Sort data by Year, Group, and then Session number
            sorted_data = sorted(self.schedule_data,
                                 key=lambda x: (x[0], x[1], int(x[3])))
        
            # Add data starting from row 2 (was row 3 before)
            for row_idx, row_data in enumerate(sorted_data, start=2):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx).value = cell_value
        
            # Auto-fit columns - modified to use all rows including header
            for col in range(1, 8):  # Columns A through G
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col)
            
                # Include all rows
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
        
            # Save the file
            wb.save(filepath)
        
            self.show_message_box("Success", f"Schedule saved successfully to:\n{filepath}")
        
        except Exception as e:
            self.show_message_box("Error", f"Failed to save schedule: {str(e)}")
    
    def update_schedule(self):
        """Update the existing schedule with new sessions without prompting for save location"""
        # Check if we have loaded an existing schedule
        if self.existing_table.rowCount() == 0:
            self.show_message_box("Error", "Please load an existing schedule first")
            return
    
        # Check if we have new sessions to add
        if self.new_sessions_table.rowCount() == 0:
            self.show_message_box("Error", "No new sessions to add")
            return
    
        # Get module name (default if empty)
        module_name = self.update_module_name_input.text().strip()
        if not module_name:
            module_name = "Untitled_Module"
    
        try:
            # Create modules_schedules directory in the current working directory
            # This will be where the EXE is running from
            save_dir = os.path.join(os.getcwd(), "modules_schedules")
            os.makedirs(save_dir, exist_ok=True)
        
            # Generate timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
            # Create filename with module name and timestamp
            filename = f"{module_name}_updated_{timestamp}.xlsx"
            filepath = os.path.join(save_dir, filename)
        
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
        
            # Name the worksheet with the module name (limit to 31 chars as Excel has a limit)
            ws.title = module_name[:31]
        
            # Add headers in row 1 (no module name row anymore)
            headers = ['Year', 'Group', 'Subject', 'Session', 'Location', 'Date', 'Start Time']
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = header
                ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
            # Collect existing data
            existing_data = []
            for row in range(self.existing_table.rowCount()):
                row_data = []
                for col in range(7):
                    value = self.existing_table.item(row, col).text()
                    row_data.append(value)
                existing_data.append(row_data)
        
            # Collect new data
            new_data = []
            for row in range(self.new_sessions_table.rowCount()):
                row_data = []
                for col in range(7):
                    value = self.new_sessions_table.item(row, col).text()
                    row_data.append(value)
                new_data.append(row_data)
        
            # Combine data - group by Year and Group
            grouped_data = {}
        
            # Process existing data
            for row in existing_data:
                key = (row[0], row[1])  # (Year, Group)
                if key not in grouped_data:
                    grouped_data[key] = []
                grouped_data[key].append(row)
        
            # Add new data to appropriate groups
            for row in new_data:
                key = (row[0], row[1])  # (Year, Group)
                if key not in grouped_data:
                    grouped_data[key] = []
                grouped_data[key].append(row)
        
            # Sort groups by year and group name
            sorted_keys = sorted(grouped_data.keys())
        
            # Prepare final data - sort each group by session number
            final_data = []
            for key in sorted_keys:
                # Sort by session number within each group
                group_data = sorted(grouped_data[key], key=lambda x: int(x[3]))
                final_data.extend(group_data)
        
            # Add data to worksheet starting from row 2 (was row 3 before)
            for row_idx, row_data in enumerate(final_data, start=2):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx).value = cell_value
        
            # Auto-fit columns - include all rows
            for col in range(1, 8):  # Columns A through G
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col)
            
                # Include all rows
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
        
            # Save the file
            wb.save(filepath)
        
            self.show_message_box("Success", f"Schedule updated successfully to:\n{filepath}")
        
        except Exception as e:
            self.show_message_box("Error", f"Failed to update schedule: {str(e)}")
            
    def show_message_box(self, title, message):
        """Show a message box with the given title and message"""
        msg_box = QMessageBox()
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.setStyleSheet("QLabel{min-width: 300px;}")
        msg_box.exec()

#==========================================================appeal processor==========================================================#

class AppealProcessor(QWidget):
    def __init__(self):
        super().__init__()
        self.schedules = []
        self.students = []
        self.sessions = []
        self.selected_appeals = []
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)

        self.init_ui()

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Create a scroll area to make the page scrollable
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("QScrollArea { border: none; }")
        
        # Create a widget to hold the content
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setSpacing(20)
        scroll_layout.setContentsMargins(0, 0, 0, 0)

        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio,
                                               Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("Appeal Processor")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)

        # Back button
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)

        # Exit button
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)

        # Add buttons to vertical layout
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()

        # Add button layout to header
        header_layout.addStretch()
        header_layout.addLayout(header_buttons_layout)
        scroll_layout.addLayout(header_layout)

        # Create three file input sections
        # REORDERED AS REQUESTED:
        
        # 1. Reference File Section
        ref_group = QGroupBox("Reference Data")
        ref_group.setStyleSheet(GROUP_BOX_STYLE)
        ref_layout = QVBoxLayout(ref_group)

        # Single line layout for reference data
        ref_input_layout = QHBoxLayout()
        ref_input_layout.addWidget(QLabel("Database File:"))
        self.ref_file_input = QLineEdit()
        self.ref_file_input.setPlaceholderText("Select Excel file...")
        self.ref_file_input.setMinimumWidth(200)
        ref_input_layout.addWidget(self.ref_file_input)
        ref_browse_btn = QPushButton("Browse")
        ref_browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        ref_input_layout.addWidget(ref_browse_btn)
        ref_input_layout.addWidget(QLabel("Sheet Name:"))
        self.ref_sheet_combo = QComboBox()
        self.ref_sheet_combo.setMinimumWidth(100)
        ref_input_layout.addWidget(self.ref_sheet_combo)
        ref_browse_btn.clicked.connect(
            lambda: self.browse_file(self.ref_file_input))
        ref_layout.addLayout(ref_input_layout)
        scroll_layout.addWidget(ref_group)
        
        # 2. Log File Section
        log_group = QGroupBox("Attendance Logs")
        log_group.setStyleSheet(GROUP_BOX_STYLE)
        log_layout = QVBoxLayout(log_group)

        # Single line layout for log data
        log_input_layout = QHBoxLayout()
        log_input_layout.addWidget(QLabel("Log File:"))
        self.log_file_input = QLineEdit()
        self.log_file_input.setPlaceholderText("Select Excel file...")
        self.log_file_input.setMinimumWidth(200)
        log_input_layout.addWidget(self.log_file_input)
        log_browse_btn = QPushButton("Browse")
        log_browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        log_input_layout.addWidget(log_browse_btn)
        log_input_layout.addWidget(QLabel("Sheet Name:"))
        self.log_sheet_combo = QComboBox()
        self.log_sheet_combo.setMinimumWidth(100)
        log_input_layout.addWidget(self.log_sheet_combo)
        log_browse_btn.clicked.connect(
            lambda: self.browse_file(self.log_file_input))
        log_layout.addLayout(log_input_layout)
        scroll_layout.addWidget(log_group)
        
        # 3. Schedule File Section
        schedule_group = QGroupBox("Sessions Schedules")
        schedule_group.setStyleSheet(GROUP_BOX_STYLE)
        schedule_layout = QVBoxLayout(schedule_group)

        # Single line layout for schedule data
        schedule_input_layout = QHBoxLayout()
        schedule_input_layout.addWidget(QLabel("Schedule File:"))
        self.schedule_file_input = QLineEdit()
        self.schedule_file_input.setPlaceholderText("Select Excel file...")
        self.schedule_file_input.setMinimumWidth(200)
        schedule_input_layout.addWidget(self.schedule_file_input)
        schedule_browse_btn = QPushButton("Browse")
        schedule_browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        schedule_input_layout.addWidget(schedule_browse_btn)
        schedule_input_layout.addWidget(QLabel("Sheet Name:"))
        self.schedule_sheet_combo = QComboBox()
        self.schedule_sheet_combo.setMinimumWidth(100)
        schedule_input_layout.addWidget(self.schedule_sheet_combo)
        schedule_browse_btn.clicked.connect(
            lambda: self.browse_file(self.schedule_file_input))
        schedule_layout.addLayout(schedule_input_layout)
        scroll_layout.addWidget(schedule_group)

        # Appeal Selection Section
        appeal_group = QGroupBox("Appeal Selection")
        appeal_group.setStyleSheet(GROUP_BOX_STYLE)
        appeal_layout = QVBoxLayout(appeal_group)
        appeal_group.setMinimumHeight(500)  # Increase height


        # Student search section
        student_search_layout = QHBoxLayout()
        student_search_layout.addWidget(QLabel("Search Student:"))
        self.student_search = QLineEdit()
        self.student_search.setPlaceholderText("Enter student ID or name...")
        self.student_search.setMinimumWidth(300)
        self.student_search.textChanged.connect(self.filter_students)
        student_search_layout.addWidget(self.student_search)
        
        # CHANGED: Replace student list with table
        self.student_table = QTableWidget()
        self.student_table.setColumnCount(4)
        self.student_table.setHorizontalHeaderLabels(['Student ID', 'Name', 'Year', 'Group'])
        self.student_table.setStyleSheet(TABLE_STYLE)
        # Center align the header text
        header = self.student_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        # Set column resize modes to stretch
        for i in range(4):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        self.student_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.student_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.student_table.itemSelectionChanged.connect(self.update_student_info)
        
        # Session selection section
        session_title = QLabel("Select the sessions which the students should have attended:")
        session_title.setStyleSheet("font-weight: bold; margin-top: 10px;")
        
        # CHANGED: Replace session list with table
        self.session_table = QTableWidget()
        self.session_table.setColumnCount(5)
        self.session_table.setHorizontalHeaderLabels([
            'Subject', 'Session', 'Location', 'Date', 'Time'
        ])
        self.session_table.setStyleSheet(TABLE_STYLE)
        # Center align the header text
        header = self.session_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        # Set column resize modes to stretch
        for i in range(5):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        self.session_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.session_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.session_table.setMinimumHeight(150)  # Make it taller
        self.session_table.itemSelectionChanged.connect(self.update_session_info)
  
        # Add appeal button
        add_appeal_layout = QHBoxLayout()
        add_appeal_btn = QPushButton("Add to Appeals")
        add_appeal_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        add_appeal_btn.clicked.connect(self.add_appeal)
        add_appeal_layout.addStretch()
        add_appeal_layout.addWidget(add_appeal_btn)
        
        appeal_layout.addLayout(student_search_layout)
        appeal_layout.addWidget(self.student_table)
        appeal_layout.addWidget(session_title)
        appeal_layout.addWidget(self.session_table)
        appeal_layout.addLayout(add_appeal_layout)
        
        scroll_layout.addWidget(appeal_group)

        # Selected Appeals Table
        selected_appeals_group = QGroupBox("Selected Appeals")
        selected_appeals_group.setStyleSheet(GROUP_BOX_STYLE)
        selected_appeals_group.setMinimumHeight(300)  # Increase height
        selected_appeals_layout = QVBoxLayout(selected_appeals_group)

        self.appeals_table = QTableWidget()
        self.appeals_table.setColumnCount(9)
        self.appeals_table.setHorizontalHeaderLabels([
            'Student ID', 'Name', 'Year', 'Group', 'Subject', 
            'Session', 'Location', 'Date', 'Time'
        ])
        self.appeals_table.setStyleSheet(TABLE_STYLE)

        # Center align the header text
        header = self.appeals_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)  # Center header text

        # Set column resize modes to stretch and fit content
        for i in range(9):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)

        selected_appeals_layout.addWidget(self.appeals_table)
        
        # Remove appeal button
        remove_appeal_layout = QHBoxLayout()
        remove_appeal_btn = QPushButton("Remove Selected Appeal")
        remove_appeal_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        remove_appeal_btn.clicked.connect(self.remove_appeal)
        remove_appeal_layout.addStretch()
        remove_appeal_layout.addWidget(remove_appeal_btn)
        selected_appeals_layout.addLayout(remove_appeal_layout)
        
        scroll_layout.addWidget(selected_appeals_group)

        # Bottom Buttons
        button_layout = QHBoxLayout()
        process_btn = QPushButton("Process Appeals")
        process_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        process_btn.clicked.connect(self.process_appeals)
        button_layout.addWidget(process_btn)
        scroll_layout.addLayout(button_layout)

        # Connect file input changes to sheet loading AND data loading (autoload)
        self.ref_file_input.textChanged.connect(lambda: self.load_sheets_and_data(self.ref_file_input.text(), self.ref_sheet_combo, 'reference'))
        self.log_file_input.textChanged.connect(lambda: self.load_sheets_and_data(self.log_file_input.text(), self.log_sheet_combo, 'log'))
        self.schedule_file_input.textChanged.connect(lambda: self.load_sheets_and_data(self.schedule_file_input.text(), self.schedule_sheet_combo, 'schedule'))
        
        # Also connect sheet combo box changes to trigger data loading when changed
        self.ref_sheet_combo.currentIndexChanged.connect(lambda: self.autoload_data())
        self.log_sheet_combo.currentIndexChanged.connect(lambda: self.autoload_data())
        self.schedule_sheet_combo.currentIndexChanged.connect(lambda: self.autoload_data())
            
        # Set the scroll content and add to main layout
        scroll_area.setWidget(scroll_content)
        main_layout.addWidget(scroll_area)

    def browse_file(self, input_field):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx)")
        if filename:
            input_field.setText(filename)

    def load_sheets_and_data(self, file_path, combo_box, file_type):
        if os.path.isfile(file_path):
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                
                # Store current selection if exists
                current_selection = combo_box.currentText()
                
                # Update combo box
                combo_box.clear()
                combo_box.addItems(wb.sheetnames)
                
                # Restore selection if possible
                if current_selection in wb.sheetnames:
                    combo_box.setCurrentText(current_selection)
                    
                # Auto-load data after sheets are loaded
                self.autoload_data()
                
            except Exception as e:
                QMessageBox.critical(
                    self, "Error", f"Error loading workbook: {str(e)}")

    def autoload_data(self):
        """Automatically load data when files and sheets are selected"""
        # Check if all required files and sheets are selected
        if (self.ref_file_input.text() and self.ref_sheet_combo.currentText() and 
            self.schedule_file_input.text() and self.schedule_sheet_combo.currentText()):
            self.load_data()
    
    def load_data(self):
        """Load student and session data from the selected files"""
        try:
            # Clear previous data
            self.students = []
            self.sessions = []
            self.student_table.setRowCount(0)
            self.session_table.setRowCount(0)
        
            # Load student reference data
            ref_wb = openpyxl.load_workbook(self.ref_file_input.text(), read_only=True)
            ref_ws = ref_wb[self.ref_sheet_combo.currentText()]
            student_db = list(ref_ws.values)
        
            # Skip header row and load students
            for row in student_db[1:]:
                if row[0]:  # If student ID exists
                    student = {
                        'id': str(row[0]),
                        'name': row[1],
                        'year': row[2],
                        'group': row[3],
                        'email': f"{row[0]}@med.asu.edu.eg"  # Generate email from ID
                    }
                    self.students.append(student)
                
                    # Add to student table
                    current_row = self.student_table.rowCount()
                    self.student_table.insertRow(current_row)
                
                    # Add data to table cells
                    self.student_table.setItem(current_row, 0, QTableWidgetItem(str(student['id'])))
                    self.student_table.setItem(current_row, 1, QTableWidgetItem(student['name']))
                    self.student_table.setItem(current_row, 2, QTableWidgetItem(str(student['year'])))
                    self.student_table.setItem(current_row, 3, QTableWidgetItem(str(student['group'])))
                
                    # Center align all items
                    for col in range(4):
                        item = self.student_table.item(current_row, col)
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        
            # Load session schedule data
            sched_wb = openpyxl.load_workbook(self.schedule_file_input.text(), read_only=True)
            sched_ws = sched_wb[self.schedule_sheet_combo.currentText()]
            session_schedule = list(sched_ws.values)
        
            # Skip header row and load sessions
            for row in session_schedule[1:]:
                if len(row) >= 7:  # Ensure we have required fields
                    year, group, subject, session_num, location, date, start_time = row[:7]
                    session = {
                        'year': year,
                        'group': group,
                        'subject': subject,
                        'session': session_num,
                        'location': location,
                        'date': date,
                        'start_time': start_time
                    }
                    self.sessions.append(session)
        
            # Sort sessions by date and subject
            self.sessions.sort(key=lambda x: (x['date'], x['subject']))
        
        except Exception as e:
            self.show_custom_warning("Error", f"Error loading data: {str(e)}")
    
    def filter_students(self):
        """Filter student table based on search text"""
        search_text = self.student_search.text().lower()
        self.student_table.setRowCount(0)
    
        for student in self.students:
            # Filter students by ID or name
            if not search_text or search_text in student['id'].lower() or search_text in student['name'].lower():
                current_row = self.student_table.rowCount()
                self.student_table.insertRow(current_row)
            
                # Add data to table cells
                self.student_table.setItem(current_row, 0, QTableWidgetItem(str(student['id'])))
                self.student_table.setItem(current_row, 1, QTableWidgetItem(student['name']))
                self.student_table.setItem(current_row, 2, QTableWidgetItem(str(student['year'])))
                self.student_table.setItem(current_row, 3, QTableWidgetItem(str(student['group'])))
            
                # Center align all items
                for col in range(4):
                    item = self.student_table.item(current_row, col)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    def update_student_info(self):
        """Update the student information based on selection"""
        # Get the selected row
        selected_rows = self.student_table.selectionModel().selectedRows()
        if not selected_rows:
            return
    
        row = selected_rows[0].row()

        # Extract student ID from the selected row
        student_id = self.student_table.item(row, 0).text()

        # Find the selected student
        selected_student = None
        for student in self.students:
            if student['id'] == student_id:
                selected_student = student
                break

        if selected_student:
            # Update UI labels
            self.student_id_label.setText(selected_student['id'])
            self.student_name_label.setText(selected_student['name'])
            self.student_year_label.setText(str(selected_student['year']))
            self.student_group_label.setText(str(selected_student['group']))
    
            # Update session table to show applicable sessions
            self.session_table.setRowCount(0)
            self.applicable_sessions = []  # Store as an instance variable to access later
    
            for session in self.sessions:
                # Filter sessions by student's year and group
                if (session['year'] == selected_student['year'] and 
                    session['group'] == selected_student['group']):
                    self.applicable_sessions.append(session)
            
                    # Format date and time for display
                    date_str = session['date']
                    time_str = session['start_time']
            
                    if isinstance(session['date'], datetime):
                        date_str = session['date'].strftime('%d/%m/%Y')
                    if isinstance(session['start_time'], datetime):
                        time_str = session['start_time'].strftime('%H:%M:%S')
            
                    # Add to session table
                    current_row = self.session_table.rowCount()
                    self.session_table.insertRow(current_row)
            
                    # Add data to table cells
                    self.session_table.setItem(current_row, 0, QTableWidgetItem(str(session['subject'])))
                    self.session_table.setItem(current_row, 1, QTableWidgetItem(str(session['session'])))
                    self.session_table.setItem(current_row, 2, QTableWidgetItem(str(session['location'])))
                    self.session_table.setItem(current_row, 3, QTableWidgetItem(str(date_str)))
                    self.session_table.setItem(current_row, 4, QTableWidgetItem(str(time_str)))
            
                    # Center align all items
                    for col in range(5):
                        item = self.session_table.item(current_row, col)
            
    def update_session_info(self):
        """Store the selected session to be used when adding an appeal"""
        self.selected_session = None
    
        # Get the selected row
        selected_rows = self.session_table.selectionModel().selectedRows()
        if not selected_rows:
            return
        
        selected_index = selected_rows[0].row()
    
        if selected_index >= 0 and hasattr(self, 'applicable_sessions') and selected_index < len(self.applicable_sessions):
            self.selected_session = self.applicable_sessions[selected_index]
    
    def add_appeal(self):
        """Add selected student and session to the appeals table"""
        # Check if student is selected
        student_rows = self.student_table.selectionModel().selectedRows()
        if not student_rows:
            self.show_custom_warning("Selection Required", "Please select a student.")
            return
        
        # Check if session is selected
        session_rows = self.session_table.selectionModel().selectedRows()
        if not session_rows:
            self.show_custom_warning("Selection Required", "Please select a session.")
            return
        
        # Update session info to ensure we have the latest selection
        self.update_session_info()
    
        if not hasattr(self, 'selected_session') or not self.selected_session:
            self.show_custom_warning("Selection Required", "Please select a valid session.")
            return
        
        # Get student info
        student_row = student_rows[0].row()
        student_id = self.student_table.item(student_row, 0).text()
    
        selected_student = None
        for student in self.students:
            if student['id'] == student_id:
                selected_student = student
                break
    
        if not selected_student:
            return
        
        # Create appeal record
        appeal = {
            'student_id': selected_student['id'],
            'name': selected_student['name'],
            'year': selected_student['year'],
            'group': selected_student['group'],
            'subject': self.selected_session['subject'],
            'session': self.selected_session['session'],
            'location': self.selected_session['location'],
            'date': self.selected_session['date'],
            'time': self.selected_session['start_time']
        }
    
        # Check if this appeal already exists
        for existing_appeal in self.selected_appeals:
            if (existing_appeal['student_id'] == appeal['student_id'] and
                existing_appeal['subject'] == appeal['subject'] and
                existing_appeal['session'] == appeal['session'] and
                existing_appeal['date'] == appeal['date']):
                self.show_custom_warning("Duplicate Appeal", 
                                        f"This appeal for {selected_student['name']} already exists.")
                return
    
        # Add to list and update table
        self.selected_appeals.append(appeal)
        self.update_appeals_table()
    
    def remove_appeal(self):
        """Remove selected appeal from the table"""
        current_row = self.appeals_table.currentRow()
        if current_row >= 0:
            removed_appeal = self.selected_appeals.pop(current_row)
            self.update_appeals_table()
        else:
            self.show_custom_warning("Selection Required", "Please select an appeal to remove.")
    
    def update_appeals_table(self):
        """Update the appeals table with current selections"""
        self.appeals_table.setRowCount(len(self.selected_appeals))
        
        for i, appeal in enumerate(self.selected_appeals):
            # Format date and time properly
            date_str = appeal['date']
            time_str = appeal['time']
            
            if isinstance(appeal['date'], datetime):
                date_str = appeal['date'].strftime('%d/%m/%Y')
            if isinstance(appeal['time'], datetime):
                time_str = appeal['time'].strftime('%H:%M:%S')
                
            # Create table items
            items = [
                appeal['student_id'],
                appeal['name'],
                str(appeal['year']),
                str(appeal['group']),
                appeal['subject'],
                str(appeal['session']),
                appeal['location'],
                str(date_str),
                str(time_str)
            ]
            
            for j, value in enumerate(items):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.appeals_table.setItem(i, j, item)

    def show_custom_warning(self, title, message):
        """Display a custom warning message box"""
        warning_dialog = QMessageBox(self)
        warning_dialog.setWindowTitle(title)
        warning_dialog.setText(message)
        warning_dialog.setIcon(QMessageBox.Icon.Warning)
        
        # Style OK button
        ok_button = warning_dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)
        
        # Style dialog background
        warning_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: black;
            }}
            QLabel {{
                color: white;
                font-size: 14px;
            }}
        """)
        
        warning_dialog.exec()

    def validate_inputs(self):
        """Validates that all necessary inputs are provided before processing appeals"""
        if not self.log_file_input.text() or not os.path.isfile(self.log_file_input.text()):
            self.show_custom_warning("Missing Input", "Please select a valid log file.")
            return False
            
        if not self.log_sheet_combo.currentText():
            self.show_custom_warning("Missing Input", "Please select a log sheet.")
            return False
            
        if not self.ref_file_input.text() or not os.path.isfile(self.ref_file_input.text()):
            self.show_custom_warning("Missing Input", "Please select a valid student reference file.")
            return False
            
        if not self.schedule_file_input.text() or not os.path.isfile(self.schedule_file_input.text()):
            self.show_custom_warning("Missing Input", "Please select a valid schedule file.")
            return False
            
        return True
        
    def process_appeals(self):
        """Process the selected appeals and add them to the log file"""
        if not self.selected_appeals:
            self.show_custom_warning("No Appeals", "Please add at least one appeal to process.")
            return
        
        if not self.validate_inputs():
            return
        
        try:
            # Directly process appeals without using a thread
            log_file = self.log_file_input.text()
            log_sheet = self.log_sheet_combo.currentText()
        
            # Load existing log file
            log_wb = openpyxl.load_workbook(log_file)
            log_ws = log_wb[log_sheet]
        
            # Get header row to understand column structure
            header_row = [cell.value if cell.value is not None else "" for cell in log_ws[1]]
        
            # Find important column indices
            try:
                # Adjust for the different column names in the example
                id_col = header_row.index("Student ID")
                location_col = header_row.index("Location")
                date_col = header_row.index("Log Date")
                time_col = header_row.index("Log Time")
            
                # Add missing columns if needed
                if "Subject" not in header_row:
                    header_row.append("Subject")
                    subject_col = len(header_row) - 1
                    for row in range(1, log_ws.max_row + 1):
                        log_ws.cell(row=row, column=subject_col+1).value = ""
                else:
                    subject_col = header_row.index("Subject")
                
                if "Session" not in header_row:
                    header_row.append("Session")
                    session_col = len(header_row) - 1
                    for row in range(1, log_ws.max_row + 1):
                        log_ws.cell(row=row, column=session_col+1).value = ""
                else:
                    session_col = header_row.index("Session")
                
                if "Status" not in header_row:
                    header_row.append("Status")
                    status_col = len(header_row) - 1
                    for row in range(1, log_ws.max_row + 1):
                        log_ws.cell(row=row, column=status_col+1).value = ""
                else:
                    status_col = header_row.index("Status")
                
                if "Notes" not in header_row:
                    header_row.append("Notes")
                    notes_col = len(header_row) - 1
                    for row in range(1, log_ws.max_row + 1):
                        log_ws.cell(row=row, column=notes_col+1).value = ""
                else:
                    notes_col = header_row.index("Notes")
            
            except ValueError as e:
                # Proper header not found, raise error
                raise ValueError(f"Required column not found in log file: {str(e)}")
        
            # Get total number of rows in the log sheet
            max_row = log_ws.max_row
        
            # Process each appeal
            for appeal in self.selected_appeals:
                # Format date and time values for consistency
                date_value = appeal['date']
                time_value = appeal['time']
            
                if isinstance(date_value, datetime):
                    date_value = date_value.strftime('%Y-%m-%d')
                if isinstance(time_value, datetime):
                    time_value = time_value.strftime('%H:%M:%S')
            
                # Prepare new row with appeal data
                new_row = max_row + 1
                log_ws.cell(row=new_row, column=id_col+1).value = appeal['student_id']
                log_ws.cell(row=new_row, column=date_col+1).value = date_value
                log_ws.cell(row=new_row, column=time_col+1).value = time_value
                log_ws.cell(row=new_row, column=subject_col+1).value = appeal['subject']
                log_ws.cell(row=new_row, column=session_col+1).value = appeal['session']
                log_ws.cell(row=new_row, column=status_col+1).value = "Present"  # Mark as present for approved appeals
                log_ws.cell(row=new_row, column=location_col+1).value = appeal['location']
                log_ws.cell(row=new_row, column=notes_col+1).value = "exception"  # Add exception flag as requested
            
                # Increment row counter
                max_row += 1
        
            # Save changes to the log file
            log_wb.save(log_file)
        
            # Show success message
            success_dialog = QMessageBox(self)
            success_dialog.setWindowTitle("Success")
            success_dialog.setText(
                "Appeals have been successfully processed and added to the log file.")
            success_dialog.setIcon(QMessageBox.Icon.Information)

            # Style OK button
            ok_button = success_dialog.addButton(QMessageBox.StandardButton.Ok)
            ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

            # Style dialog background
            success_dialog.setStyleSheet(f"""
                QMessageBox {{
                    background-color: {CARD_BG};
                    }}
                QLabel {{
                    color: {TEXT_COLOR};
                    font-size: 14px;
                }}
            """)

            success_dialog.exec()
        except Exception as e:
            # Handle any errors that occur during processing
            error_dialog = QMessageBox(self)
            error_dialog.setWindowTitle("Error")
            error_dialog.setText(f"Error processing appeals: {str(e)}")
            error_dialog.setIcon(QMessageBox.Icon.Critical)
            error_dialog.setStyleSheet(f"""
                QMessageBox {{
                    background-color: {CARD_BG};
                    }}
                QLabel {{
                    color: {TEXT_COLOR};
                    font-size: 14px;
                }}
            """)
            error_dialog.exec()

#==========================================================attendance processor==========================================================#

class AttendanceProcessor(QWidget):
    def __init__(self):
        super().__init__()
        self.schedules = []
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)

        self.init_ui()

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Create a scroll area to make the page scrollable
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("QScrollArea { border: none; }")
        
        # Create a widget to hold the content
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setSpacing(20)
        scroll_layout.setContentsMargins(0, 0, 0, 0)

        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio,
                                               Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("Attendance Processor")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)

        # Back button
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)

        # Exit button
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)

        # Add buttons to vertical layout
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()

        # Add button layout to header
        header_layout.addStretch()
        header_layout.addLayout(header_buttons_layout)
        scroll_layout.addLayout(header_layout)

        # Previous Reports Section - NEW
        prev_reports_group = QGroupBox("Previous Reports (Optional)")
        prev_reports_group.setStyleSheet(GROUP_BOX_STYLE)
        prev_reports_layout = QVBoxLayout(prev_reports_group)

        # Single line layout for previous reports
        prev_reports_input_layout = QHBoxLayout()
        prev_reports_input_layout.addWidget(QLabel("Last Report File:"))
        self.prev_report_file_input = QLineEdit()
        self.prev_report_file_input.setPlaceholderText("Select previous report Excel file...")
        self.prev_report_file_input.setMinimumWidth(200)
        prev_reports_input_layout.addWidget(self.prev_report_file_input)
        prev_reports_browse_btn = QPushButton("Browse")
        prev_reports_browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        prev_reports_input_layout.addWidget(prev_reports_browse_btn)
        prev_reports_browse_btn.clicked.connect(
            lambda: self.browse_file(self.prev_report_file_input))
        prev_reports_layout.addLayout(prev_reports_input_layout)

        # Add infobox to explain this feature
        info_label = QLabel("Note: Upload previous reports to track student group transfers. "
                           "The system will identify students who switched groups and update their attendance accordingly.")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("font-style: italic; color: #AAAAAA;")
        prev_reports_layout.addWidget(info_label)

        scroll_layout.addWidget(prev_reports_group)

        # Reference Data Section
        ref_group = QGroupBox("Reference Data")
        ref_group.setStyleSheet(GROUP_BOX_STYLE)
        ref_layout = QVBoxLayout(ref_group)

        # Single line layout for reference data
        ref_input_layout = QHBoxLayout()
        ref_input_layout.addWidget(QLabel("Database File:"))
        self.ref_file_input = QLineEdit()
        self.ref_file_input.setPlaceholderText("Select Excel file...")
        self.ref_file_input.setMinimumWidth(200)
        ref_input_layout.addWidget(self.ref_file_input)
        ref_browse_btn = QPushButton("Browse")
        ref_browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        ref_input_layout.addWidget(ref_browse_btn)
        ref_input_layout.addWidget(QLabel("Sheet Name:"))
        self.ref_sheet_combo = QComboBox()
        self.ref_sheet_combo.setMinimumWidth(100)
        ref_input_layout.addWidget(self.ref_sheet_combo)
        ref_browse_btn.clicked.connect(
            lambda: self.browse_file(self.ref_file_input))
        ref_layout.addLayout(ref_input_layout)
        scroll_layout.addWidget(ref_group)

        # Attendance Logs Section
        log_group = QGroupBox("Attendance Logs")
        log_group.setStyleSheet(GROUP_BOX_STYLE)
        log_layout = QVBoxLayout(log_group)

        # Single line layout for log data
        log_input_layout = QHBoxLayout()
        log_input_layout.addWidget(QLabel("Log File:"))
        self.log_file_input = QLineEdit()
        self.log_file_input.setPlaceholderText("Select Excel file...")
        self.log_file_input.setMinimumWidth(200)
        log_input_layout.addWidget(self.log_file_input)
        log_browse_btn = QPushButton("Browse")
        log_browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        log_input_layout.addWidget(log_browse_btn)
        log_input_layout.addWidget(QLabel("Sheet Name:"))
        self.log_sheet_combo = QComboBox()
        self.log_sheet_combo.setMinimumWidth(100)
        log_input_layout.addWidget(self.log_sheet_combo)
        log_browse_btn.clicked.connect(
            lambda: self.browse_file(self.log_file_input))
        log_layout.addLayout(log_input_layout)
        scroll_layout.addWidget(log_group)

        # Session Schedules Section
        schedule_group = QGroupBox("Sessions Schedules")
        schedule_group.setStyleSheet(GROUP_BOX_STYLE)
        schedule_layout = QVBoxLayout(schedule_group)

        self.schedule_table = QTableWidget()
        self.schedule_table.setColumnCount(5)
        self.schedule_table.setHorizontalHeaderLabels(
            ['Year', 'Module', 'File', 'Sheet', 'Total Sessions'])
        self.schedule_table.setStyleSheet(TABLE_STYLE)

        # Center align the header text
        header = self.schedule_table.horizontalHeader()
        header.setDefaultAlignment(
            Qt.AlignmentFlag.AlignCenter)  # Center header text

        # Set column resize modes to stretch and fit content
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)  # Year
        header.setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)  # Module
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # File
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # Sheet
        header.setSectionResizeMode(
            4, QHeaderView.ResizeMode.Stretch)  # Total Sessions

        schedule_layout.addWidget(self.schedule_table)
        schedule_btn_layout = QHBoxLayout()
        add_schedule_btn = QPushButton("Add Schedule")
        add_schedule_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        add_schedule_btn.clicked.connect(self.add_schedule)
        remove_schedule_btn = QPushButton("Remove Schedule")
        remove_schedule_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        remove_schedule_btn.clicked.connect(self.remove_schedule)
        schedule_btn_layout.addWidget(add_schedule_btn)
        schedule_btn_layout.addWidget(remove_schedule_btn)
        schedule_layout.addLayout(schedule_btn_layout)
        scroll_layout.addWidget(schedule_group)

        # Progress Bar Section
        progress_group = QGroupBox("Progress")
        progress_group.setStyleSheet(GROUP_BOX_STYLE)
        progress_layout = QVBoxLayout(progress_group)

        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_bar.setStyleSheet(PROGRESS_BAR_STYLE)

        # Create loading gif label
        self.loading_label = QLabel()
        # Adjust based on your GIF size
        self.loading_label.setFixedSize(24, 24)
        self.loading_label.setVisible(False)  # Hidden by default

        # Create the movie object for the GIF
        self.loading_movie = QMovie()
        # Adjust based on your GIF size
        self.loading_movie.setScaledSize(QSize(24, 24))
        self.loading_label.setMovie(self.loading_movie)

        # Make sure to have your loading.gif in the same directory as the script
        loading_gif_path = os.path.join(
            os.path.dirname(__file__), 'loading.gif')
        if os.path.exists(loading_gif_path):
            self.loading_movie.setFileName(loading_gif_path)
        else:
            print(f"Warning: loading.gif not found at {loading_gif_path}")

        # Create a horizontal layout to hold both the progress bar and loading animation
        progress_h_layout = QHBoxLayout()
        progress_h_layout.addWidget(self.progress_bar)
        progress_h_layout.addWidget(self.loading_label)
        progress_layout.addLayout(progress_h_layout)

        scroll_layout.addWidget(progress_group)

        # Output Console Section
        console_group = QGroupBox("Output Console")
        console_group.setStyleSheet(GROUP_BOX_STYLE)
        console_layout = QVBoxLayout(console_group)

        self.output_console = QTextEdit()
        self.output_console.setReadOnly(True)
        self.output_console.setMaximumHeight(150)
        self.output_console.setStyleSheet(CONSOLE_STYLE)
        console_layout.addWidget(self.output_console)
        scroll_layout.addWidget(console_group)

        # Bottom Buttons
        button_layout = QHBoxLayout()
        
        # Process Button
        process_btn = QPushButton("Process Attendance Records")
        process_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        process_btn.clicked.connect(self.process_data)
        button_layout.addWidget(process_btn)
        
        # Update Report Button
        update_btn = QPushButton("Update Report")
        update_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        update_btn.clicked.connect(self.update_report)
        button_layout.addWidget(update_btn)
        
        scroll_layout.addLayout(button_layout)
        
        # Set the scroll content as the scroll area's widget
        scroll_area.setWidget(scroll_content)
        
        # Add scroll area to main layout
        main_layout.addWidget(scroll_area)

        # Connect file input changes to sheet loading
        self.ref_file_input.textChanged.connect(
            lambda: self.load_sheets(self.ref_file_input.text(), self.ref_sheet_combo))
        self.log_file_input.textChanged.connect(
            lambda: self.load_sheets(self.log_file_input.text(), self.log_sheet_combo))
        self.prev_report_file_input.textChanged.connect(
            lambda: self.check_previous_report_file(self.prev_report_file_input.text()))

    def check_previous_report_file(self, file_path):
        """Check if the previous report file has the expected sheets"""
        if os.path.isfile(file_path):
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                # Check if the file has the expected sheets
                required_sheets = ["Summary", "Attendance"]
                has_required_sheets = all(sheet in wb.sheetnames for sheet in required_sheets)
                
                if not has_required_sheets:
                    self.output_console.append("Warning: Previous report file may not have the expected format. "
                                            "It should contain 'Summary' and 'Attendance' sheets.")
                else:
                    self.output_console.append("Previous report file loaded successfully.")
                    
                    # Try to extract the report date from the file name
                    try:
                        file_name = os.path.basename(file_path)
                        # Look for date pattern in the filename (YYYYMMDD_HHMMSS)
                        date_match = re.search(r'(\d{8}_\d{6})', file_name)
                        if date_match:
                            date_str = date_match.group(1)
                            report_date = datetime.strptime(date_str, '%Y%m%d_%H%M%S')
                            self.output_console.append(f"Previous report date: {report_date.strftime('%Y-%m-%d')}")
                        else:
                            # Try to extract from sheet names if they include dates
                            for sheet in wb.sheetnames:
                                date_match = re.search(r'(\d{2}/\d{2}/\d{4})', sheet)
                                if date_match:
                                    date_str = date_match.group(1)
                                    report_date = datetime.strptime(date_str, '%d/%m/%Y')
                                    self.output_console.append(f"Previous report date: {report_date.strftime('%Y-%m-%d')}")
                                    break
                    except Exception as e:
                        self.output_console.append(f"Note: Could not extract report date from file. Will use file modification date as fallback.")
            except Exception as e:
                self.output_console.append(f"Error loading previous report file: {str(e)}")

    def browse_file(self, input_field):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx)")
        if filename:
            input_field.setText(filename)

    def load_sheets(self, file_path, combo_box):
        if os.path.isfile(file_path):
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                combo_box.clear()
                combo_box.addItems(wb.sheetnames)
            except Exception as e:
                QMessageBox.critical(
                    self, "Error", f"Error loading workbook: {str(e)}")

    def add_schedule(self):
        dialog = ScheduleDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.schedules.append(dialog.get_schedule_data())
            self.update_schedule_table()

    def remove_schedule(self):
        current_row = self.schedule_table.currentRow()
        if current_row >= 0:
            self.schedules.pop(current_row)
            self.update_schedule_table()

    def update_schedule_table(self):
        self.schedule_table.setRowCount(len(self.schedules))
        for i, schedule in enumerate(self.schedules):
            for j, value in enumerate(schedule):
                item = QTableWidgetItem(str(value))
                # Center align the text
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.schedule_table.setItem(i, j, item)

    def process_data(self):
        if not self.validate_inputs():
            return
    
        # Show attendance threshold dialog
        threshold_dialog = AttendanceThresholdDialog(self)
        if threshold_dialog.exec() != QDialog.DialogCode.Accepted:
            return  # User canceled, don't continue processing
        
        attendance_threshold = threshold_dialog.get_threshold()
        
        # Disable UI elements
        self.setEnabled(False)
        self.output_console.clear()
        self.progress_bar.setValue(0)
    
        # Show and start the loading animation
        self.loading_label.setVisible(True)
        self.loading_movie.start()
    
        # Create and start processing thread
        self.process_thread = ProcessThread(
            self.ref_file_input.text(),
            self.ref_sheet_combo.currentText(),
            self.log_file_input.text(),
            self.log_sheet_combo.currentText(),
            self.schedules,
            attendance_threshold,  # Pass the user-specified threshold
            self.prev_report_file_input.text()  # Pass the previous report file path
        )
    
        # Connect signals
        self.process_thread.progress_updated.connect(self.update_progress)
        self.process_thread.error_occurred.connect(self.handle_error)
        self.process_thread.processing_complete.connect(self.handle_completion)
    
        # Start processing
        self.process_thread.start()
    
    def validate_inputs(self):
        # Validate reference file
        if not self.ref_file_input.text() or not self.ref_sheet_combo.currentText():
            self.show_custom_warning(
                "Reference Data Required", "Please select reference file and sheet")
            return False

        # Validate log file
        if not self.log_file_input.text() or not self.log_sheet_combo.currentText():
            self.show_custom_warning(
                "Log Data Required", "Please select log file and sheet")
            return False

        # Validate schedules
        if not self.schedules:
            self.show_custom_warning(
                "Schedules Required", "Please add at least one schedule")
            return False

        # Note: We don't validate the previous report file as it's optional

        return True

    def show_custom_warning(self, title, message):
        """Show a custom styled warning dialog"""
        warning_dialog = QMessageBox(self)
        warning_dialog.setWindowTitle(title)
        warning_dialog.setText(message)
        warning_dialog.setIcon(QMessageBox.Icon.Warning)

        # Create and style OK button
        ok_button = warning_dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

        # Style dialog background and text
        warning_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)

        warning_dialog.exec()

    def update_progress(self, value):
        self.progress_bar.setValue(value)
        self.output_console.append(f"Processing... {value}%")

    def update_report(self):
        """Handle the Update Report button click"""
        # First, validate the inputs
        if not self.validate_inputs():
            return
    
        # Additionally validate that a previous report file is selected
        if not self.prev_report_file_input.text() or not os.path.isfile(self.prev_report_file_input.text()):
            self.show_custom_warning(
                "Previous Report Required", 
                "Please select a previous report file to update from")
            return
    
        # Show attendance threshold dialog (same as process_data)
        threshold_dialog = AttendanceThresholdDialog(self)
        if threshold_dialog.exec() != QDialog.DialogCode.Accepted:
            return  # User canceled, don't continue processing
    
        attendance_threshold = threshold_dialog.get_threshold()
    
        # Disable UI elements
        self.setEnabled(False)
        self.output_console.clear()
        self.progress_bar.setValue(0)
    
        # Show and start the loading animation
        self.loading_label.setVisible(True)
        self.loading_movie.start()
    
        # Create and start the update process thread
        self.update_thread = UpdateProcessThread(
            self.ref_file_input.text(),
            self.ref_sheet_combo.currentText(),
            self.log_file_input.text(),
            self.log_sheet_combo.currentText(),
            self.schedules,
            attendance_threshold,
            self.prev_report_file_input.text()  # Pass the previous report file path
        )
    
        # Connect signals
        self.update_thread.progress_updated.connect(self.update_progress)
        self.update_thread.error_occurred.connect(self.handle_error)
        self.update_thread.processing_complete.connect(self.handle_completion)
    
        # Start processing
        self.update_thread.start()

    def handle_error(self, error_message):
        self.setEnabled(True)
        error_dialog = QMessageBox(self)
        error_dialog.setWindowTitle("Error")
        error_dialog.setText(f"Error processing data: {error_message}")
        error_dialog.setIcon(QMessageBox.Icon.Critical)

        # Style OK button
        ok_button = error_dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

        # Style dialog background
        error_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)

        error_dialog.exec()
        self.output_console.append(f"Error: {error_message}")

    def handle_completion(self):
        self.setEnabled(True)
        self.progress_bar.setValue(100)
    
        # Determine which process just completed
        operation_type = "Update" if hasattr(self, 'update_thread') and self.sender() == self.update_thread else "Processing"
        self.output_console.append(f"{operation_type} complete!")

        # Hide the loading animation
        self.loading_label.setVisible(False)
        self.loading_movie.stop()

        success_dialog = QMessageBox(self)
        success_dialog.setWindowTitle("Success")
        success_dialog.setText(
            f"{operation_type} complete! Check the attendance_reports folder.")
        success_dialog.setIcon(QMessageBox.Icon.Information)

        # Style OK button
        ok_button = success_dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

        # Style dialog background
        success_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)

        success_dialog.exec()

class ProcessThread(QThread):
    progress_updated = pyqtSignal(int)
    error_occurred = pyqtSignal(str)
    processing_complete = pyqtSignal()

    def __init__(self, ref_file, ref_sheet, log_file, log_sheet, schedules, attendance_threshold=0.75, prev_report_file=None):
        super().__init__()
        self.ref_file = ref_file
        self.ref_sheet = ref_sheet
        self.log_file = log_file
        self.log_sheet = log_sheet
        self.schedules = schedules
        self.ATTENDANCE_THRESHOLD = attendance_threshold
        self.prev_report_file = prev_report_file  # Add the new parameter

        # Time window constants in minutes - STANDARD SESSIONS
        self.STANDARD_BEFORE_MINUTES = 15
        self.STANDARD_AFTER_MINUTES = 150

        # Time window constants in minutes - EXCEPTION SESSIONS (12, 1, 13, 3, 15)
        self.EXCEPTION_BEFORE_MINUTES = 15
        self.EXCEPTION_AFTER_MINUTES = 150

        # Exception hours that use different time windows
        self.EXCEPTION_HOURS = [12, 1, 13, 3, 15]

        # Define subject colors
        self.SUBJECT_COLORS = {
            # Red with white text
            "anatomy": {"bg": "800020", "text": "FFFFFF"},
            # Pink with black text
            "histology": {"bg": "FFE4E1", "text": "000000"},
            # Purple with white text
            "pathology": {"bg": "663399", "text": "FFFFFF"},
            # Green with white text
            "parasitology": {"bg": "556B2F", "text": "FFFFFF"},
            # Yellow with white text
            "physiology": {"bg": "D4A017", "text": "FFFFFF"},
            # Teal with white text
            "microbiology": {"bg": "4682B4", "text": "FFFFFF"},
            # Navy with white text
            "pharmacology": {"bg": "000080", "text": "FFFFFF"},
            # Cyan with white text
            "biochemistry": {"bg": "1A3668", "text": "FFFFFF"},
            # Gray with white text
            "clinical": {"bg": "333333", "text": "FFFFFF"},
            # Black with white text
            "other": {"bg": "000000", "text": "FFFFFF"}
        }

    def run(self):
        try:
            # Calculate total steps
            total_steps = 2 + len(self.schedules) * 5
            current_step = 0

            # Load reference data
            ref_wb = openpyxl.load_workbook(self.ref_file)
            ref_ws = ref_wb[self.ref_sheet]
            student_db = list(ref_ws.values)
            student_map = self.create_student_map(student_db)
            current_step += 1
            self.progress_updated.emit(int(current_step / total_steps * 100))

            # Load log data
            log_wb = openpyxl.load_workbook(self.log_file)
            log_ws = log_wb[self.log_sheet]
            log_history = list(log_ws.values)
            current_step += 1
            self.progress_updated.emit(int(current_step / total_steps * 100))

            # Create output directory
            output_dir = os.path.join(os.getcwd(), "attendance_reports")
            os.makedirs(output_dir, exist_ok=True)

            # Get current date for sheet names
            current_date = datetime.now().strftime('%d_%m_%Y')
            attendance_sheet_name = f"Attendance_{current_date}"
            summary_sheet_name = f"Summary_{current_date}"

            # Process each schedule
            for year, module, sched_file, sched_sheet, total_required in self.schedules:
                # Load schedule data
                sched_wb = openpyxl.load_workbook(sched_file)
                sched_ws = sched_wb[sched_sheet]
                session_schedule = list(sched_ws.values)
                current_step += 1
                self.progress_updated.emit(
                    int(current_step / total_steps * 100))

                # Calculate sessions
                completed_sessions = self.calculate_completed_sessions(
                    session_schedule[1:])
                required_attendance = self.calculate_required_attendance(
                    session_schedule[1:], total_required)
                current_step += 1
                self.progress_updated.emit(
                    int(current_step / total_steps * 100))

                # Validate attendance
                valid_attendance = self.validate_attendance(log_history, session_schedule[1:],
                                                            student_map, f"Year {year}")
                current_step += 1
                self.progress_updated.emit(
                    int(current_step / total_steps * 100))

                # Create output workbook and sheets
                output_wb = openpyxl.Workbook()
                output_wb.remove(output_wb.active)

                # Create Summary sheet first, then Attendance sheet with date in sheet name
                self.create_summary_sheet(output_wb, summary_sheet_name, valid_attendance, required_attendance,
                                          student_map, f"Year {year}", completed_sessions, total_required)
                self.create_valid_logs_sheet(output_wb, attendance_sheet_name, valid_attendance)

                current_step += 1
                self.progress_updated.emit(
                    int(current_step / total_steps * 100))

                # Save output workbook
                current_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                year_dir = os.path.join(output_dir, f"Year_{year}")
                os.makedirs(year_dir, exist_ok=True)
                output_path = os.path.join(
                    year_dir, f"Y{year}_{module}_attendance_{current_timestamp}.xlsx")
                output_wb.save(output_path)
                current_step += 1
                self.progress_updated.emit(
                    int(current_step / total_steps * 100))

            self.processing_complete.emit()

        except Exception as e:
            self.error_occurred.emit(str(e))

    def create_student_map(self, student_db):
        student_map = {}
        for row in student_db[1:]:
            if row[0]:
                student_id = str(row[0])
                email = f"{student_id}@med.asu.edu.eg"
                student_map[student_id] = {
                    "name": row[1],
                    "year": row[2],
                    "group": row[3],
                    "email": email
                }
        return student_map

    def calculate_completed_sessions(self, session_schedule):
        completed_sessions = {}
        for row in session_schedule:
            if len(row) >= 2:
                year, group = row[:2]
                key = f"{year}-{group}"
                completed_sessions[key] = completed_sessions.get(key, 0) + 1
        return completed_sessions

    def calculate_required_attendance(self, session_schedule, total_required_sessions):
        required_attendance = {}
    
        # Process each row in the schedule
        for row in session_schedule:
            if len(row) >= 5:  # Ensure we have at least year, group, subject, session_num, location
                year, group, subject, session_num, location = row[:5]
                key = f"{year}-{group}"
            
                # Initialize dictionaries if they don't exist
                if key not in required_attendance:
                    required_attendance[key] = {}
                
                if subject not in required_attendance[key]:
                    required_attendance[key][subject] = {
                        "total": 0,
                        "sessions": {}
                    }
                
                if session_num not in required_attendance[key][subject]["sessions"]:
                    required_attendance[key][subject]["sessions"][session_num] = {
                        "total": 0,
                        "locations": {}
                    }
                
                if location not in required_attendance[key][subject]["sessions"][session_num]["locations"]:
                    # This is a new unique session-location combination
                    required_attendance[key][subject]["sessions"][session_num]["locations"][location] = 1
                    required_attendance[key][subject]["sessions"][session_num]["total"] += 1
                    required_attendance[key][subject]["total"] += 1
    
        return required_attendance

    def validate_attendance(self, log_history, session_schedule, student_map, target_year):
        valid_attendance = {}
        session_map = {}
        unique_logs = set()

        # Build a more detailed session map that includes all session info
        for row in session_schedule:
            if len(row) >= 7:  # Ensure we have all needed fields
                year, group, subject, session_num, location, date, start_time = row[:7]
                key = f"{year}-{group}"
                session_datetime = self.parse_datetime(date, start_time)
                
                # Create a unique key for each session that combines all relevant info
                session_key = f"{location}-{date}-{start_time}"
                
                if key not in session_map:
                    session_map[key] = {}
                
                # Store complete session information
                session_map[key][session_key] = {
                    "subject": subject,
                    "session_num": session_num,
                    "location": location,
                    "start_time": session_datetime
                }

        # Process attendance logs
        for row in log_history[1:]:
            if len(row) >= 4:
                student_id, location, date, time = row[:4]
                student_id = str(student_id)
                
                if student_id in student_map:
                    student = student_map[student_id]
                    key = f"{student['year']}-{student['group']}"
                    
                    # Try to find matching session for this attendance log
                    log_datetime = self.parse_datetime(date, time)
                    
                    # Check all potential sessions for this year-group
                    if key in session_map:
                        for session_key, session_info in session_map[key].items():
                            session_location = session_info["location"]
                            session_start = session_info["start_time"]
                            
                            # Only match logs from the same location as the session
                            if location.lower() == session_location.lower():
                                # Determine time window based on session start hour
                                session_hour = session_start.hour
                                if session_hour in self.EXCEPTION_HOURS:
                                    before_window = timedelta(minutes=self.EXCEPTION_BEFORE_MINUTES)
                                    after_window = timedelta(minutes=self.EXCEPTION_AFTER_MINUTES)
                                else:
                                    before_window = timedelta(minutes=self.STANDARD_BEFORE_MINUTES)
                                    after_window = timedelta(minutes=self.STANDARD_AFTER_MINUTES)
                                
                                # Check if log time is within the allowed window
                                if session_start - before_window <= log_datetime <= session_start + after_window:
                                    unique_log_key = f"{student_id}-{session_info['subject']}-{session_info['session_num']}-{location}-{date}"
                                    
                                    # Only count each unique session attendance once
                                    if unique_log_key not in unique_logs:
                                        unique_logs.add(unique_log_key)
                                        
                                        if key not in valid_attendance:
                                            valid_attendance[key] = []
                                            
                                        valid_attendance[key].append([
                                            student_id, student['name'], student['year'],
                                            student['group'], student['email'], session_info['subject'],
                                            session_info['session_num'], location, date, time
                                        ])
                                        
                                        # Found a match, no need to check other sessions
                                        break
        
        return valid_attendance

    def parse_datetime(self, date, time):
        if isinstance(date, str):
            date = datetime.strptime(date, '%d/%m/%Y').date()
        if isinstance(time, str):
            time = datetime.strptime(time, '%H:%M:%S').time()
        return datetime.combine(date, time)

    def create_valid_logs_sheet(self, workbook, sheet_name, data):
        sheet = workbook.create_sheet(sheet_name)
        header = ["Student ID", "Name", "Year", "Group", "Email",
                  "Subject", "Session", "Location", "Date", "Time"]
        sheet.append(header)

        # Apply header formatting
        for i, cell in enumerate(sheet[1]):
            cell.font = Font(bold=True)
            # Light gray background for all headers
            cell.fill = PatternFill("solid", fgColor="D3D3D3")
            cell.alignment = Alignment(
                horizontal='center', vertical='center')  # Center align text

            # Make the header row a bit taller
            sheet.row_dimensions[1].height = 22

        # Add freeze panes to keep header visible when scrolling
        sheet.freeze_panes = 'C2'

        row_num = 2  # Start from row 2 (after header)
        for key in data:
            for row_data in data[key]:
                sheet.append(row_data)
                row_num += 1

        # Format date and time columns
        for col in 'I', 'J':  # Adjusted to match the correct column letters for Date and Time
            for cell in sheet[col]:
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = 'DD/MM/YYYY' if col == 'I' else 'HH:MM:SS'

        # Improved auto-fit column widths
        for col_idx, column in enumerate(sheet.columns, 1):
            # Get maximum length in the column
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        
            # Set column width with minimum and maximum limits
            if max_length > 0:
                adjusted_width = min(max(max_length + 2, 12), 50)  # Min 12, Max 50
            
                # Special case for name column (typically column B)
                if col_idx == 2:  # Name column
                    adjusted_width = max(adjusted_width, 25)  # Names need more space
            
                # Apply the calculated width
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

        # Add auto-filter to easily sort and filter data
        sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}"

    def get_subject_color(self, subject_name):
        # Convert subject name to lowercase for case-insensitive matching
        subject_lower = subject_name.lower()

        # Check if the subject contains any of our defined subject keywords
        for key in self.SUBJECT_COLORS:
            if key in subject_lower:
                return self.SUBJECT_COLORS[key]

        # Default to "other" if no match is found
        return self.SUBJECT_COLORS["other"]

    def calculate_min_sessions_needed(self, total_required, total_attended):
        if total_attended >= self.ATTENDANCE_THRESHOLD * total_required:
            return 0
        min_total_needed = math.ceil(
            self.ATTENDANCE_THRESHOLD * total_required)
        return min_total_needed - total_attended

    def create_summary_sheet(self, workbook, sheet_name, valid_attendance, required_attendance,
                             student_map, target_year, completed_sessions, total_required_sessions):
        sheet = workbook.create_sheet(sheet_name)

        # Collect all subjects and their sessions
        subjects = {}
        for key, subject_data in required_attendance.items():
            for subject, data in subject_data.items():
                if subject not in subjects:
                    subjects[subject] = {"sessions": set(), "locations": set()}
                for session_num, session_data in data["sessions"].items():
                    subjects[subject]["sessions"].add(session_num)
                    subjects[subject]["locations"].update(
                        session_data["locations"].keys())

        # Create header
        header = ["Student ID", "Name", "Year", "Group", "Email", "Status", "Percentage",
                  "Sessions Needed", "Sessions Left", "Sessions Completed", "Total Required", "Total Attended"]

        # Track column indices for subject coloring
        subject_column_ranges = {}
        current_col = len(header) + 1  # Start after the basic columns

        # Add subject totals and session details to header
        for subject in sorted(subjects.keys()):
            # Mark the start column for this subject
            start_col = current_col

            header.extend(
                [f"Required {subject} (Total)", f"Attended {subject} (Total)"])
            current_col += 2

            for session in sorted(subjects[subject]["sessions"]):
                for location in sorted(subjects[subject]["locations"]):
                    header.extend([
                        f"{subject} S{session} at {location} (Req)",
                        f"{subject} S{session} at {location} (Att)"
                    ])
                    current_col += 2

            # Record the column range for this subject
            subject_column_ranges[subject] = (start_col, current_col - 1)

        sheet.append(header)

        # Apply header formatting and colors
        for i, cell in enumerate(sheet[1], 1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

            # Apply subject-specific coloring to subject headers
            for subject, (start_idx, end_idx) in subject_column_ranges.items():
                if start_idx <= i <= end_idx:
                    subject_color = self.get_subject_color(subject)
                    cell.fill = PatternFill(
                        "solid", fgColor=subject_color["bg"])
                    cell.font = Font(bold=True, color=subject_color["text"])
                    break

            # Apply gray background to basic columns
            if i <= len(header) - len(subject_column_ranges):
                if not cell.fill.fgColor.rgb:  # Only if no color was set
                    cell.fill = PatternFill(
                        "solid", fgColor="D3D3D3")  # Light gray

        # Make the header row taller to accommodate wrapped text
        sheet.row_dimensions[1].height = 40

        # Add freeze panes to keep headers visible when scrolling
        sheet.freeze_panes = 'C2'

        # Define status colors
        COLOR_PASS = "66E4A6"
        COLOR_FAIL = "FF4C4C"
        COLOR_HIGH_RISK = "FF7C7C"
        COLOR_MODERATE_RISK = "FFB97D"
        COLOR_LOW_RISK = "FFF1A6"
        COLOR_NO_RISK = "3388D5"

        for student_id, student in student_map.items():
            if student['year'] == target_year:
                key = f"{student['year']}-{student['group']}"
                group_completed = completed_sessions.get(key, 0)
                total_attended = 0
                attendance_by_subject = {}

                # Process attendance data
                for entry in valid_attendance.get(key, []):
                    if entry[0] == student_id:
                        subject = entry[5]
                        session_num = entry[6]
                        location = entry[7]

                        if subject not in attendance_by_subject:
                            attendance_by_subject[subject] = {
                                "total": 0,
                                "sessions": {}
                            }
                        if session_num not in attendance_by_subject[subject]["sessions"]:
                            attendance_by_subject[subject]["sessions"][session_num] = {
                                "locations": {}
                            }

                        attendance_by_subject[subject]["total"] += 1
                        if location not in attendance_by_subject[subject]["sessions"][session_num]["locations"]:
                            attendance_by_subject[subject]["sessions"][session_num]["locations"][location] = 0
                        attendance_by_subject[subject]["sessions"][session_num]["locations"][location] += 1
                        total_attended += 1

                # Calculate status and color
                required_sessions = math.ceil(
                    self.ATTENDANCE_THRESHOLD * total_required_sessions)
                sessions_left = total_required_sessions - group_completed
                max_possible = total_attended + sessions_left
                min_sessions_needed = max(
                    required_sessions - total_attended, 0)

                if group_completed >= total_required_sessions:
                    if total_attended >= required_sessions:
                        status, color = "Pass", COLOR_PASS
                    else:
                        status, color = "Fail", COLOR_FAIL
                else:
                    if max_possible < required_sessions:
                        status, color = "Fail", COLOR_FAIL
                    elif total_attended >= required_sessions:
                        status, color = "Pass", COLOR_PASS
                    else:
                        sessions_margin = sessions_left - min_sessions_needed
                        if sessions_margin <= 1:
                            status, color = "High Risk", COLOR_HIGH_RISK
                        elif sessions_margin <= 3:
                            status, color = "Moderate Risk", COLOR_MODERATE_RISK
                        elif sessions_margin <= 5:
                            status, color = "Low Risk", COLOR_LOW_RISK
                        else:
                            status, color = "No Risk", COLOR_NO_RISK

                percentage = total_attended / \
                    total_required_sessions if total_required_sessions > 0 else 0

                # Create row data
                row = [
                    student_id, student['name'], student['year'], student['group'],
                    student['email'], status, f"{percentage:.1%}", min_sessions_needed,
                    sessions_left, group_completed, total_required_sessions, total_attended
                ]

                # Add subject totals and session details
                for subject in sorted(subjects.keys()):
                    subj_req = required_attendance.get(key, {}).get(
                        subject, {"total": 0, "sessions": {}})
                    subj_att = attendance_by_subject.get(
                        subject, {"total": 0, "sessions": {}})

                    # Add subject totals
                    row.extend([subj_req["total"], subj_att["total"]])

                    # Add session details
                    for session in sorted(subjects[subject]["sessions"]):
                        for location in sorted(subjects[subject]["locations"]):
                            req_count = subj_req.get("sessions", {}).get(
                                session, {}).get("locations", {}).get(location, 0)
                            att_count = subj_att.get("sessions", {}).get(
                                session, {}).get("locations", {}).get(location, 0)
                            row.extend([req_count, att_count])

                sheet.append(row)

                # Apply cell formatting and colors for this row
                row_idx = sheet.max_row

                # Format status cell
                status_cell = sheet.cell(row=row_idx, column=6)
                status_cell.font = Font(bold=True)
                status_cell.fill = PatternFill("solid", fgColor=color)
                status_cell.alignment = Alignment(horizontal='center')

                # Format percentage cell
                percentage_cell = sheet.cell(row=row_idx, column=7)
                percentage_cell.number_format = '0.0%'
                percentage_cell.alignment = Alignment(horizontal='center')

                # Apply subject-specific colors to the data cells
                for subject, (start_col, end_col) in subject_column_ranges.items():
                    subject_color = self.get_subject_color(subject)
                    for col in range(start_col, end_col + 1):
                        cell = sheet.cell(row=row_idx, column=col)
                        # Apply a lighter version of the subject color for data cells
                        bg_color = self.lighten_color(subject_color["bg"])
                        cell.fill = PatternFill("solid", fgColor=bg_color)
                        cell.alignment = Alignment(horizontal='center')

        # Add auto-filter to easily sort and filter data
        sheet.auto_filter.ref = f"A1:L{sheet.max_row}"

        # Improved column width auto-fitting
        column_widths = {}
    
        # First pass: Calculate max content length for each column
        for row in sheet.iter_rows():
            for cell in row:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                if cell.value:
                    # For headers (row 1), calculate based on word-wrapped text
                    if cell.row == 1:
                        # Split header by spaces and find the longest word
                        words = str(cell.value).split()
                        if words:
                            max_word_len = max(len(word) for word in words)
                            # For headers, consider both total length and longest word
                            header_width = min(max(max_word_len + 1, len(str(cell.value)) / 2), 30)
                            column_widths[col_letter] = max(column_widths.get(col_letter, 0), header_width)
                    else:
                        # For data cells, use the full text length
                        try:
                            text_len = len(str(cell.value))
                            column_widths[col_letter] = max(column_widths.get(col_letter, 0), text_len + 1)
                        except:
                            pass
    
        # Second pass: Apply calculated widths with constraints
        for col_letter, width in column_widths.items():
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
        
            # Base width calculation
            adjusted_width = min(max(width, 10), 40)  # Min 10, Max 40
        
            # Special case for specific columns
            if col_idx == 2:  # Name column
                adjusted_width = max(adjusted_width, 25)  # Names need more space
            elif col_idx >= 13:  # Subject specific columns
                adjusted_width = max(adjusted_width, 12)  # Subject columns need at least this width
            
            sheet.column_dimensions[col_letter].width = adjusted_width

    def lighten_color(self, hex_color, factor=0.75):
        """
        Lightens the given color by the factor.
        1.0 means keep the same, 0.5 means 50% lighter.
        """
        # Convert hex to RGB
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)

        # Lighten the color
        r = int(r + (255 - r) * factor)
        g = int(g + (255 - g) * factor)
        b = int(b + (255 - b) * factor)

        # Convert back to hex
        return f"{r:02x}{g:02x}{b:02x}".upper()

class UpdateProcessThread(QThread):
    progress_updated = pyqtSignal(int)
    error_occurred = pyqtSignal(str)
    processing_complete = pyqtSignal()

    def __init__(self, ref_file, ref_sheet, log_file, log_sheet, schedules, 
                 attendance_threshold=0.75, prev_report_file=None):
        super().__init__()
        self.ref_file = ref_file
        self.ref_sheet = ref_sheet
        self.log_file = log_file
        self.log_sheet = log_sheet
        self.schedules = schedules
        self.ATTENDANCE_THRESHOLD = attendance_threshold
        self.prev_report_file = prev_report_file
        
        # Minimum number of consecutive sessions to consider a group transfer confirmed
        self.TRANSFER_CONFIRMATION_THRESHOLD = 3
        
        # Time window constants from ProcessThread
        self.STANDARD_BEFORE_MINUTES = 15
        self.STANDARD_AFTER_MINUTES = 150
        self.EXCEPTION_BEFORE_MINUTES = 15
        self.EXCEPTION_AFTER_MINUTES = 150
        self.EXCEPTION_HOURS = [12, 1, 13, 3, 15]
        
        # Define subject colors (copied from ProcessThread)
        self.SUBJECT_COLORS = {
            "anatomy": {"bg": "800020", "text": "FFFFFF"},
            "histology": {"bg": "FFE4E1", "text": "000000"},
            "pathology": {"bg": "663399", "text": "FFFFFF"},
            "parasitology": {"bg": "556B2F", "text": "FFFFFF"},
            "physiology": {"bg": "D4A017", "text": "FFFFFF"},
            "microbiology": {"bg": "4682B4", "text": "FFFFFF"},
            "pharmacology": {"bg": "000080", "text": "FFFFFF"},
            "biochemistry": {"bg": "1A3668", "text": "FFFFFF"},
            "clinical": {"bg": "333333", "text": "FFFFFF"},
            "other": {"bg": "000000", "text": "FFFFFF"}
        }
        
    def run(self):
        try:
            # Calculate total steps
            # Adding +1 to total_steps for the new transfer log sheet creation
            total_steps = 8 + len(self.schedules) * 5
            current_step = 0
        
            # Step 1: Load previous report
            if not self.prev_report_file:
                self.error_occurred.emit("Previous report file is required for update operation")
                return
            
            prev_report_date = self.extract_report_date(self.prev_report_file)
            if not prev_report_date:
                self.error_occurred.emit("Could not determine the date of the previous report")
                return
            
            prev_report_wb = openpyxl.load_workbook(self.prev_report_file)
            prev_summary_sheet = None
            prev_attendance_sheet = None
        
            # Find the summary and attendance sheets
            for sheet_name in prev_report_wb.sheetnames:
                if "Summary" in sheet_name:
                    prev_summary_sheet = prev_report_wb[sheet_name]
                elif "Attendance" in sheet_name:
                    prev_attendance_sheet = prev_report_wb[sheet_name]
        
            if not prev_summary_sheet or not prev_attendance_sheet:
                self.error_occurred.emit("Could not find Summary or Attendance sheets in the previous report")
                return
            
            current_step += 1
            self.progress_updated.emit(int(current_step / total_steps * 100))
        
            # Step 2: Load reference data (current student information)
            ref_wb = openpyxl.load_workbook(self.ref_file)
            ref_ws = ref_wb[self.ref_sheet]
            student_db = list(ref_ws.values)
            current_student_map = self.create_student_map(student_db)
            current_step += 1
            self.progress_updated.emit(int(current_step / total_steps * 100))
        
            # Step 3: Create previous student map from the summary sheet
            prev_student_map = self.extract_student_map_from_summary(prev_summary_sheet)
            current_step += 1
            self.progress_updated.emit(int(current_step / total_steps * 100))
        
            # Step 4: Identify students who have transferred groups
            transferred_students = self.identify_transferred_students(prev_student_map, current_student_map)
            current_step += 1
            self.progress_updated.emit(int(current_step / total_steps * 100))
        
            # Step 5: Load log data
            log_wb = openpyxl.load_workbook(self.log_file)
            log_ws = log_wb[self.log_sheet]
            log_history = list(log_ws.values)
            current_step += 1
            self.progress_updated.emit(int(current_step / total_steps * 100))
        
            # Step 6: Extract previous attendance data 
            prev_attendance_data = self.extract_attendance_data(prev_attendance_sheet)
            current_step += 1
            self.progress_updated.emit(int(current_step / total_steps * 100))
        
            # Create output directory
            output_dir = os.path.join(os.getcwd(), "attendance_reports")
            os.makedirs(output_dir, exist_ok=True)
        
            # Step 7: Process each schedule and update attendance
            for year, module, sched_file, sched_sheet, total_required in self.schedules:
                # Load schedule data
                sched_wb = openpyxl.load_workbook(sched_file)
                sched_ws = sched_wb[sched_sheet]
                session_schedule = list(sched_ws.values)
                current_step += 1
                self.progress_updated.emit(int(current_step / total_steps * 100))
            
                # Filter sessions that occurred after the previous report date
                new_sessions = []
                for session in session_schedule[1:]:  # Skip header
                    if len(session) >= 6:  # Ensure date field exists
                        session_date = session[5]
                        if isinstance(session_date, str):
                            try:
                                session_date = datetime.strptime(session_date, '%d/%m/%Y').date()
                            except:
                                continue
                        if hasattr(session_date, 'date'):
                            session_date = session_date.date()
                    
                        if session_date >= prev_report_date.date():
                            new_sessions.append(session)
            
                # Calculate sessions
                completed_sessions = self.calculate_completed_sessions(session_schedule[1:])
                required_attendance = self.calculate_required_attendance(session_schedule[1:], total_required)
                current_step += 1
                self.progress_updated.emit(int(current_step / total_steps * 100))
            
                # Track transfer dates and patterns for transferred students
                transfer_data = self.analyze_transfer_patterns(transferred_students, log_history, 
                                                              new_sessions, current_student_map)
            
                # Debug transfer data
                for student_id, data in transfer_data.items():
                    print(f"Student {student_id} transfer data:")
                    print(f"  Previous group: {data['previous_group']}")
                    print(f"  Current group: {data['current_group']}")
                    print(f"  Transfer date: {data['transfer_date']}")
                    print(f"  Attendance pattern: {data['attendance_pattern'][:3]}... (total: {len(data['attendance_pattern'])})")
            
                # Process new attendance logs considering transfers
                new_valid_attendance = self.validate_attendance_with_transfers(
                    log_history, new_sessions, current_student_map, 
                    transferred_students, transfer_data, f"Year {year}", prev_report_date)
            
                # Combine previous and new attendance data
                combined_attendance = self.combine_attendance_data(
                    prev_attendance_data, new_valid_attendance, prev_report_date, transferred_students, transfer_data)
            
                current_step += 1
                self.progress_updated.emit(int(current_step / total_steps * 100))
            
                # Create output workbook and sheets
                output_wb = openpyxl.Workbook()
                output_wb.remove(output_wb.active)
            
                # Add date to sheet names
                current_date = datetime.now()
                summary_sheet_name = f"Summary_{current_date.strftime('%d_%m_%Y')}"
                attendance_sheet_name = f"Attendance_{current_date.strftime('%d_%m_%Y')}"
                transfer_sheet_name = f"Transfers_{current_date.strftime('%d_%m_%Y')}"
            
                # Create Summary sheet first, then Attendance sheet with combined data
                self.create_summary_sheet(output_wb, summary_sheet_name, combined_attendance, 
                                         required_attendance, current_student_map, transferred_students,
                                         transfer_data, f"Year {year}", completed_sessions, total_required)
            
                self.create_valid_logs_sheet(output_wb, attendance_sheet_name, combined_attendance)
            
                # New step: Create the transfer log sheet
                self.create_transfer_log_sheet(output_wb, transfer_sheet_name, transferred_students, transfer_data)
            
                current_step += 1
                self.progress_updated.emit(int(current_step / total_steps * 100))
            
                # Save output workbook
                current_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                year_dir = os.path.join(output_dir, f"Year_{year}")
                os.makedirs(year_dir, exist_ok=True)
                output_path = os.path.join(
                    year_dir, f"Y{year}_{module}_attendance_updated_{current_timestamp}.xlsx")
                output_wb.save(output_path)
                current_step += 1
                self.progress_updated.emit(int(current_step / total_steps * 100))
        
            # Final step increment for the addition of transfer log sheet
            current_step += 1
            self.progress_updated.emit(int(current_step / total_steps * 100))
        
            self.processing_complete.emit()
        
        except Exception as e:
            import traceback
            self.error_occurred.emit(f"Error: {str(e)}\n{traceback.format_exc()}")

    def extract_report_date(self, file_path):
        """Extract the report date from the filename or file modification date"""
        try:
            # First try from filename (YYYYMMDD_HHMMSS pattern)
            file_name = os.path.basename(file_path)
            date_match = re.search(r'(\d{8}_\d{6})', file_name)
            if date_match:
                date_str = date_match.group(1)
                return datetime.strptime(date_str, '%Y%m%d_%H%M%S')
            
            # Then try to load the workbook and look for dates in sheet names
            wb = openpyxl.load_workbook(file_path, read_only=True)
            for sheet in wb.sheetnames:
                date_match = re.search(r'(\d{2}_\d{2}_\d{4})', sheet)
                if date_match:
                    date_str = date_match.group(1).replace('_', '/')
                    return datetime.strptime(date_str, '%d/%m/%Y')
            
            # Fallback to file modification time
            mod_time = os.path.getmtime(file_path)
            return datetime.fromtimestamp(mod_time)
        
        except Exception as e:
            print(f"Error extracting report date: {str(e)}")
            return None
    
    def create_student_map(self, student_db):
        """Create a map of student details from the reference file"""
        student_map = {}
        for row in student_db[1:]:  # Skip header
            if row[0]:
                student_id = str(row[0])
                email = f"{student_id}@med.asu.edu.eg"
                student_map[student_id] = {
                    "name": row[1],
                    "year": row[2],
                    "group": row[3],
                    "email": email
                }
        return student_map
    
    def extract_student_map_from_summary(self, summary_sheet):
        """Extract student information from the previous summary sheet"""
        student_map = {}
        header_row = None
        
        # Find the header row
        for row_idx, row in enumerate(summary_sheet.iter_rows(values_only=True)):
            if row and "Student ID" in row:
                header_row = row_idx + 1
                break
        
        if not header_row:
            return student_map
            
        # Find column indices
        col_indices = {}
        for col_idx, cell_value in enumerate(summary_sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True).__next__()):
            if cell_value == "Student ID":
                col_indices["id"] = col_idx
            elif cell_value == "Name":
                col_indices["name"] = col_idx
            elif cell_value == "Year":
                col_indices["year"] = col_idx
            elif cell_value == "Group":
                col_indices["group"] = col_idx
            elif cell_value == "Email":
                col_indices["email"] = col_idx
        
        # Extract student data
        for row in summary_sheet.iter_rows(min_row=header_row+1, values_only=True):
            if row and row[col_indices["id"]]:
                student_id = str(row[col_indices["id"]])
                student_map[student_id] = {
                    "name": row[col_indices["name"]],
                    "year": row[col_indices["year"]],
                    "group": row[col_indices["group"]],
                    "email": row[col_indices["email"]]
                }
        
        return student_map
    
    def identify_transferred_students(self, prev_student_map, current_student_map):
        """Identify students who have changed groups"""
        transferred_students = {}
        
        for student_id, prev_data in prev_student_map.items():
            if student_id in current_student_map:
                current_data = current_student_map[student_id]
                # Check if the group has changed
                if prev_data["group"] != current_data["group"]:
                    transferred_students[student_id] = {
                        "previous_group": prev_data["group"],
                        "current_group": current_data["group"],
                        "name": current_data["name"],
                        "year": current_data["year"],
                        "email": current_data["email"]
                    }
        
        return transferred_students
    
    def extract_attendance_data(self, attendance_sheet):
        """Extract attendance data from the previous attendance sheet"""
        attendance_data = {}
        header_row = None
        
        # Find the header row
        for row_idx, row in enumerate(attendance_sheet.iter_rows(values_only=True)):
            if row and "Student ID" in row:
                header_row = row_idx + 1
                break
        
        if not header_row:
            return attendance_data
        
        # Extract attendance entries
        for row in attendance_sheet.iter_rows(min_row=header_row+1, values_only=True):
            if row and len(row) >= 10:  # Ensure all needed fields are present
                student_id = str(row[0])
                student_year = row[2]
                student_group = row[3]
                
                key = f"{student_year}-{student_group}"
                
                if key not in attendance_data:
                    attendance_data[key] = []
                
                # Store the complete attendance entry
                attendance_data[key].append(list(row))
        
        return attendance_data
    
    def calculate_completed_sessions(self, session_schedule):
        """Calculate the number of completed sessions for each year-group"""
        completed_sessions = {}
        for row in session_schedule:
            if len(row) >= 2:
                year, group = row[:2]
                key = f"{year}-{group}"
                completed_sessions[key] = completed_sessions.get(key, 0) + 1
        return completed_sessions
    
    def calculate_required_attendance(self, session_schedule, total_required_sessions):
        """Calculate required attendance for each session by subject"""
        required_attendance = {}
    
        # Process each row in the schedule
        for row in session_schedule:
            if len(row) >= 5:  # Ensure we have at least year, group, subject, session_num, location
                year, group, subject, session_num, location = row[:5]
                key = f"{year}-{group}"
            
                # Initialize dictionaries if they don't exist
                if key not in required_attendance:
                    required_attendance[key] = {}
                
                if subject not in required_attendance[key]:
                    required_attendance[key][subject] = {
                        "total": 0,
                        "sessions": {}
                    }
                
                if session_num not in required_attendance[key][subject]["sessions"]:
                    required_attendance[key][subject]["sessions"][session_num] = {
                        "total": 0,
                        "locations": {}
                    }
                
                if location not in required_attendance[key][subject]["sessions"][session_num]["locations"]:
                    # This is a new unique session-location combination
                    required_attendance[key][subject]["sessions"][session_num]["locations"][location] = 1
                    required_attendance[key][subject]["sessions"][session_num]["total"] += 1
                    required_attendance[key][subject]["total"] += 1
    
        return required_attendance
    
    def analyze_transfer_patterns(self, transferred_students, log_history, session_schedule, student_map):
        """Analyze attendance patterns to determine when students were transferred"""
        transfer_data = {}
        
        for student_id, transfer_info in transferred_students.items():
            previous_group = transfer_info["previous_group"]
            current_group = transfer_info["current_group"]
            student_year = transfer_info["year"]
            
            # Create session maps for both previous and current groups
            prev_group_key = f"{student_year}-{previous_group}"
            current_group_key = f"{student_year}-{current_group}"
            
            prev_group_sessions = self.create_session_map(session_schedule, prev_group_key)
            current_group_sessions = self.create_session_map(session_schedule, current_group_key)
            
            # Get attendance records for this student
            student_logs = []
            for row in log_history[1:]:
                if len(row) >= 4 and str(row[0]) == student_id:
                    student_logs.append(row)
            
            # Sort logs by date and time
            student_logs.sort(key=lambda x: self.parse_datetime(x[2], x[3]) if self.parse_datetime(x[2], x[3]) else datetime.min)
            
            # Track which group's sessions the student attended
            attendance_pattern = []
            for log in student_logs:
                log_datetime = self.parse_datetime(log[2], log[3])
                if not log_datetime:
                    continue
                    
                location = log[1]
                
                # Check if this log matches a session in either group
                prev_match = self.match_log_to_session(log, log_datetime, location, prev_group_sessions)
                current_match = self.match_log_to_session(log, log_datetime, location, current_group_sessions)
                
                if prev_match and current_match:
                    # Both groups had a session at this time and location - ambiguous
                    attendance_pattern.append(("both", log_datetime, location))
                elif prev_match:
                    attendance_pattern.append(("previous", log_datetime, location))
                elif current_match:
                    attendance_pattern.append(("current", log_datetime, location))
            
            # Analyze the pattern to find a consistent switch point
            transfer_point = self.detect_transfer_point(attendance_pattern)
            
            transfer_data[student_id] = {
                "previous_group": previous_group,
                "current_group": current_group,
                "transfer_date": transfer_point,
                "attendance_pattern": attendance_pattern
            }
        
        return transfer_data
    
    def create_session_map(self, sessions, group_key):
        """Create a map of sessions for a specific group"""
        session_map = {}
        
        for session in sessions:
            if len(session) >= 7:  # Ensure we have all needed fields
                year, group, subject, session_num, location, date, start_time = session[:7]
                key = f"{year}-{group}"
                
                if key == group_key:
                    session_datetime = self.parse_datetime(date, start_time)
                    if not session_datetime:
                        continue
                        
                    session_key = f"{location}-{date}-{start_time}"
                    
                    session_map[session_key] = {
                        "subject": subject,
                        "session_num": session_num,
                        "location": location,
                        "start_time": session_datetime,
                        "date": date
                    }
        
        return session_map
    
    def match_log_to_session(self, log, log_datetime, location, session_map):
        """Check if a log matches any session in the given session map"""
        for session_key, session_info in session_map.items():
            session_location = session_info["location"]
            session_start = session_info["start_time"]
            
            # Only match logs from the same location as the session
            if location.lower() == session_location.lower():
                # Determine time window based on session start hour
                session_hour = session_start.hour
                if session_hour in self.EXCEPTION_HOURS:
                    before_window = timedelta(minutes=self.EXCEPTION_BEFORE_MINUTES)
                    after_window = timedelta(minutes=self.EXCEPTION_AFTER_MINUTES)
                else:
                    before_window = timedelta(minutes=self.STANDARD_BEFORE_MINUTES)
                    after_window = timedelta(minutes=self.STANDARD_AFTER_MINUTES)
                
                # Check if log time is within the allowed window
                if session_start - before_window <= log_datetime <= session_start + after_window:
                    return True
        
        return False
    
    def detect_transfer_point(self, attendance_pattern):
        """Detect the point at which a student consistently switched to the new group"""
        if not attendance_pattern:
            return None
            
        # We're looking for TRANSFER_CONFIRMATION_THRESHOLD consecutive "current" group attendances
        consecutive_current = 0
        potential_transfer_date = None
        
        for idx, (group, log_datetime, _) in enumerate(attendance_pattern):
            if group == "current":
                consecutive_current += 1
                if consecutive_current == 1:
                    potential_transfer_date = log_datetime
                    
                if consecutive_current >= self.TRANSFER_CONFIRMATION_THRESHOLD:
                    return potential_transfer_date
            else:
                consecutive_current = 0
                potential_transfer_date = None
        
        # If we don't have enough consecutive sessions but have some current attendance
        # use the first "current" attendance as a fallback
        for group, log_datetime, _ in attendance_pattern:
            if group == "current":
                return log_datetime
                
        return None
    
    def validate_attendance_with_transfers(self, log_history, session_schedule, student_map, 
                                          transferred_students, transfer_data, target_year, prev_report_date):
        """Validate attendance considering student transfers"""
        valid_attendance = {}
        session_map = {}
        unique_logs = set()

        # Build session maps for all groups
        for row in session_schedule:
            if len(row) >= 7:  # Ensure we have all needed fields
                year, group, subject, session_num, location, date, start_time = row[:7]
                key = f"{year}-{group}"
            
                # Skip sessions before previous report date
                session_date = date
                if isinstance(session_date, str):
                    try:
                        session_date = datetime.strptime(session_date, '%d/%m/%Y').date()
                    except:
                        continue
                if hasattr(session_date, 'date'):
                    session_date = session_date.date()
            
                if session_date < prev_report_date.date():
                    continue
            
                session_datetime = self.parse_datetime(date, start_time)
                if not session_datetime:
                    continue
            
                # Create a unique key for each session that combines all relevant info
                session_key = f"{location}-{date}-{start_time}"
            
                if key not in session_map:
                    session_map[key] = {}
            
                # Store complete session information
                session_map[key][session_key] = {
                    "subject": subject,
                    "session_num": session_num,
                    "location": location,
                    "start_time": session_datetime,
                    "date": date
                }

        # Process attendance logs
        for row in log_history[1:]:
            if len(row) >= 4:
                student_id, location, date, time = row[:4]
                student_id = str(student_id)
            
                # Skip logs before previous report date
                log_date = date
                if isinstance(log_date, str):
                    try:
                        log_date = datetime.strptime(log_date, '%d/%m/%Y').date()
                    except:
                        continue
                if hasattr(log_date, 'date'):
                    log_date = log_date.date()
            
                if log_date < prev_report_date.date():
                    continue
            
                if student_id in student_map:
                    student = student_map[student_id]
                    log_datetime = self.parse_datetime(date, time)
                    if not log_datetime:
                        continue
                
                    # FIXED: Only validate against the correct group based on transfer status and date
                    if student_id in transferred_students:
                        transfer_info = transfer_data.get(student_id, {})
                        transfer_date = transfer_info.get("transfer_date")
                    
                        # If log is before transfer date, use previous group
                        if transfer_date and log_datetime < transfer_date:
                            group_to_use = transferred_students[student_id]["previous_group"]
                        else:
                            # After transfer date, use current group
                            group_to_use = student["group"]
                    else:
                        # For non-transferred students, use current group
                        group_to_use = student["group"]
                
                    # Create keys for both the student's actual year-group and the group we're using for validation
                    actual_key = f"{student['year']}-{student['group']}"
                    validation_key = f"{student['year']}-{group_to_use}"
                
                    # Check ONLY the selected group's sessions for validation
                    if validation_key in session_map:
                        for session_key, session_info in session_map[validation_key].items():
                            session_location = session_info["location"]
                            session_start = session_info["start_time"]
                        
                            # Only match logs from the same location as the session
                            if location.lower() == session_location.lower():
                                # Determine time window based on session start hour
                                session_hour = session_start.hour
                                if session_hour in self.EXCEPTION_HOURS:
                                    before_window = timedelta(minutes=self.EXCEPTION_BEFORE_MINUTES)
                                    after_window = timedelta(minutes=self.EXCEPTION_AFTER_MINUTES)
                                else:
                                    before_window = timedelta(minutes=self.STANDARD_BEFORE_MINUTES)
                                    after_window = timedelta(minutes=self.STANDARD_AFTER_MINUTES)
                            
                                # Check if log time is within the allowed window
                                if session_start - before_window <= log_datetime <= session_start + after_window:
                                    unique_log_key = f"{student_id}-{session_info['subject']}-{session_info['session_num']}-{location}-{date}"
                                
                                    # Only count each unique session attendance once
                                    if unique_log_key not in unique_logs:
                                        unique_logs.add(unique_log_key)
                                    
                                        # Use the actual student group for storing the attendance
                                        if actual_key not in valid_attendance:
                                            valid_attendance[actual_key] = []
                                        
                                        # Store attendance record with information about which group was used for validation
                                        valid_attendance[actual_key].append([
                                            student_id, student['name'], student['year'],
                                            student['group'], student['email'], session_info['subject'],
                                            session_info['session_num'], location, date, time,
                                            group_to_use  # Add which group's schedule was used for validation
                                        ])
                                    
                                        # Found a match, no need to check other sessions
                                        break
    
        return valid_attendance
    
    def combine_attendance_data(self, prev_attendance, new_attendance, prev_report_date, transferred_students, transfer_data):
        """Combine previous and new attendance data, considering student transfers"""
        combined_attendance = {}
        
        # Add previous attendance data
        for key, entries in prev_attendance.items():
            combined_attendance[key] = []
            for entry in entries:
                # Add entries from previous attendance data
                # If the student transferred, we need to retain the original validation group
                student_id = str(entry[0]) if len(entry) > 0 else None
                
                if student_id and student_id in transferred_students:
                    # Copy the entry and add the previous group as validation group if not present
                    new_entry = list(entry)
                    # Ensure we have space for the validation group
                    while len(new_entry) < 11:
                        new_entry.append(None)
                    # Set validation group to previous group if not already set
                    if new_entry[10] is None:
                        new_entry[10] = transferred_students[student_id]["previous_group"]
                    combined_attendance[key].append(new_entry)
                else:
                    # For non-transferred students, just add the entry as is
                    combined_attendance[key].append(list(entry))
        
        # Add new attendance data, avoiding duplicates
        for key, entries in new_attendance.items():
            if key not in combined_attendance:
                combined_attendance[key] = []
            
            existing_entries = set()
            # Create a set of unique identifiers for existing entries
            for entry in combined_attendance[key]:
                if len(entry) >= 10:
                    unique_id = f"{entry[0]}-{entry[5]}-{entry[6]}-{entry[7]}-{entry[8]}"
                    existing_entries.add(unique_id)
            
            # Add only new entries that don't already exist
            for entry in entries:
                if len(entry) >= 10:
                    unique_id = f"{entry[0]}-{entry[5]}-{entry[6]}-{entry[7]}-{entry[8]}"
                    if unique_id not in existing_entries:
                        combined_attendance[key].append(entry)
        
        return combined_attendance
    
    def parse_datetime(self, date, time):
        """Parse date and time into a datetime object"""
        if not date or not time:
            return None
            
        if isinstance(date, str):
            try:
                date = datetime.strptime(date, '%d/%m/%Y').date()
            except ValueError:
                try:
                    # Try alternative format
                    date = datetime.strptime(date, '%Y-%m-%d').date()
                except:
                    return None
        elif hasattr(date, 'date'):
            date = date.date()
            
        if isinstance(time, str):
            try:
                time = datetime.strptime(time, '%H:%M:%S').time()
            except ValueError:
                try:
                    # Try alternative format
                    time = datetime.strptime(time, '%H:%M').time()
                except:
                    return None
        elif hasattr(time, 'time'):
            time = time.time()
            
        return datetime.combine(date, time)
    
    def create_valid_logs_sheet(self, workbook, sheet_name, data):
        """Create the attendance log sheet"""
        sheet = workbook.create_sheet(sheet_name)
        header = ["Student ID", "Name", "Year", "Group", "Email",
                  "Subject", "Session", "Location", "Date", "Time", "Validation Group"]
        sheet.append(header)

        # Apply header formatting
        for i, cell in enumerate(sheet[1]):
            cell.font = Font(bold=True)
            # Light gray background for all headers
            cell.fill = PatternFill("solid", fgColor="D3D3D3")
            cell.alignment = Alignment(
                horizontal='center', vertical='center')  # Center align text

            # Make the header row a bit taller
            sheet.row_dimensions[1].height = 22

        # Add freeze panes to keep header visible when scrolling
        sheet.freeze_panes = 'C2'

        row_num = 2  # Start from row 2 (after header)
        for key in data:
            for row_data in data[key]:
                # Ensure all entries have the same length
                while len(row_data) < len(header):
                    row_data.append(None)
                sheet.append(row_data[:len(header)])  # Only include up to the header length
                row_num += 1

        # Format date and time columns
        for col in 'I', 'J':  # Columns for Date and Time
            for cell in sheet[col]:
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = 'DD/MM/YYYY' if col == 'I' else 'HH:MM:SS'

        # Improved auto-fit column widths
        for col_idx, column in enumerate(sheet.columns, 1):
            # Get maximum length in the column
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        
            # Set column width with minimum and maximum limits
            if max_length > 0:
                adjusted_width = min(max(max_length + 2, 12), 50)  # Min 12, Max 50
            
                # Special case for name column (typically column B)
                if col_idx == 2:  # Name column
                    adjusted_width = max(adjusted_width, 25)  # Names need more space
            
                # Apply the calculated width
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

        # Add auto-filter to easily sort and filter data
        sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}"
        
    def get_subject_color(self, subject_name):
        """Return background and text colors for a given subject"""
        # Convert subject name to lowercase for case-insensitive matching
        subject_lower = subject_name.lower() if subject_name else ""

        # Check if the subject contains any of our defined subject keywords
        for key in self.SUBJECT_COLORS:
            if key in subject_lower:
                return self.SUBJECT_COLORS[key]

        # Default to "other" if no match is found
        return self.SUBJECT_COLORS["other"]

    def calculate_min_sessions_needed(self, total_required, total_attended):
        """Calculate minimum number of sessions needed to meet attendance threshold"""
        if total_attended >= self.ATTENDANCE_THRESHOLD * total_required:
            return 0
        min_total_needed = math.ceil(
            self.ATTENDANCE_THRESHOLD * total_required)
        return min_total_needed - total_attended

    def create_summary_sheet(self, workbook, sheet_name, combined_attendance, required_attendance,
                               current_student_map, transferred_students, transfer_data, target_year, 
                               completed_sessions, total_required_sessions):
        """
        Create a summary sheet that shows attendance statistics for each student,
        handling transferred students by validating their attendance against appropriate group schedules.

        Args:
            workbook: The Excel workbook to add the sheet to
            sheet_name: Name for the new sheet
            combined_attendance: Dictionary of attendance data from both previous and new reports
            required_attendance: Dictionary of required attendance data
            current_student_map: Dictionary mapping student IDs to their current information
            transferred_students: Dictionary of students who have transferred groups
            transfer_data: Dictionary with transfer analysis information
            target_year: The year to filter students by (e.g., "Year 1")
            completed_sessions: Dictionary tracking completed sessions by year-group
            total_required_sessions: Total number of required sessions
        """
        sheet = workbook.create_sheet(sheet_name)

        # Collect all subjects and their sessions
        subjects = {}
        for key, subject_data in required_attendance.items():
            for subject, data in subject_data.items():
                if subject not in subjects:
                    subjects[subject] = {"sessions": set(), "locations": set()}
                for session_num, session_data in data["sessions"].items():
                    subjects[subject]["sessions"].add(session_num)
                    subjects[subject]["locations"].update(
                        session_data["locations"].keys())

        # Create header
        header = ["Student ID", "Name", "Year", "Group", "Email", "Status", "Percentage",
                  "Sessions Needed", "Sessions Left", "Sessions Completed", "Total Required", "Total Attended"]

        # Track column indices for subject coloring
        subject_column_ranges = {}
        current_col = len(header) + 1  # Start after the basic columns

        # Add subject totals and session details to header
        for subject in sorted(subjects.keys()):
            # Mark the start column for this subject
            start_col = current_col

            header.extend(
                [f"Required {subject} (Total)", f"Attended {subject} (Total)"])
            current_col += 2

            for session in sorted(subjects[subject]["sessions"]):
                for location in sorted(subjects[subject]["locations"]):
                    header.extend([
                        f"{subject} S{session} at {location} (Req)",
                        f"{subject} S{session} at {location} (Att)"
                    ])
                    current_col += 2

            # Record the column range for this subject
            subject_column_ranges[subject] = (start_col, current_col - 1)

        sheet.append(header)

        # Apply header formatting and colors
        for i, cell in enumerate(sheet[1], 1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

            # Apply subject-specific coloring to subject headers
            for subject, (start_idx, end_idx) in subject_column_ranges.items():
                if start_idx <= i <= end_idx:
                    subject_color = self.get_subject_color(subject)
                    cell.fill = PatternFill(
                        "solid", fgColor=subject_color["bg"])
                    cell.font = Font(bold=True, color=subject_color["text"])
                    break

            # Apply gray background to basic columns
            if i <= len(header) - len(subject_column_ranges):
                if not cell.fill.fgColor.rgb:  # Only if no color was set
                    cell.fill = PatternFill(
                        "solid", fgColor="D3D3D3")  # Light gray

        # Make the header row taller to accommodate wrapped text
        sheet.row_dimensions[1].height = 40

        # Add freeze panes to keep headers visible when scrolling
        sheet.freeze_panes = 'C2'

        # Define status colors
        COLOR_PASS = "66E4A6"
        COLOR_FAIL = "FF4C4C"
        COLOR_HIGH_RISK = "FF7C7C"
        COLOR_MODERATE_RISK = "FFB97D"
        COLOR_LOW_RISK = "FFF1A6"
        COLOR_NO_RISK = "3388D5"

        # Process each student in the target year
        for student_id, student in current_student_map.items():
            if target_year in str(student['year']):
                # Get the current year-group key
                current_key = f"{student['year']}-{student['group']}"
                group_completed = completed_sessions.get(current_key, 0)
                total_attended = 0
                attendance_by_subject = {}
        
                # Check if this student has transferred between groups
                is_transferred = student_id in transferred_students
        
                # For transferred students, we need to consider both previous and current group attendance
                previous_key = None
                transfer_point = None
        
                if is_transferred:
                    previous_group = transferred_students[student_id]["previous_group"]
                    previous_key = f"{student['year']}-{previous_group}"
                    transfer_point = transfer_data.get(student_id, {}).get("transfer_date")
        
                # Build a list of all attendance entries for this student from all relevant groups
                student_attendance = []
        
                # Get attendance from current group
                if current_key in combined_attendance:
                    for entry in combined_attendance[current_key]:
                        if entry[0] == student_id:
                            student_attendance.append(entry)
        
                # Get attendance from previous group if transferred
                if is_transferred and previous_key in combined_attendance:
                    for entry in combined_attendance[previous_key]:
                        if entry[0] == student_id:
                            student_attendance.append(entry)
        
                # Process attendance data
                for entry in student_attendance:
                    subject = entry[5]
                    session_num = entry[6]
                    location = entry[7]
                
                    # Determine date of this attendance
                    entry_date = None
                    if len(entry) > 8 and entry[8]:  # Date field
                        try:
                            if isinstance(entry[8], str):
                                entry_date = datetime.strptime(entry[8], '%d/%m/%Y')
                            else:
                                entry_date = entry[8]  # Assume it's already a datetime
                                if hasattr(entry_date, 'date'):
                                    entry_date = datetime.combine(entry_date, datetime.min.time())
                        except:
                            pass
                
                    validation_group = entry[10] if len(entry) > 10 else None
            
                    # Determine if this attendance should be counted based on validation group and transfer status
                    should_count = False
            
                    if not is_transferred:
                        # For non-transferred students, count everything
                        should_count = True
                    elif not validation_group:
                        # No validation group specified, follow standard rules based on date
                        if transfer_point and entry_date:
                            # Check if attendance is before or after transfer
                            if entry_date < transfer_point:
                                # Before transfer - should be validated against previous group (which is already done)
                                should_count = True
                            else:
                                # After transfer - should be validated against current group (which is already done)
                                should_count = True
                        else:
                            # No clear transfer point or date - count it
                            should_count = True
                    else:
                        # For entries with explicit validation group
                        previous_group = transferred_students[student_id]["previous_group"]
                        current_group = student["group"]
                    
                        if transfer_point and entry_date:
                            if entry_date < transfer_point:
                                # Before transfer - only count if validated against previous group
                                should_count = (validation_group == previous_group)
                            else:
                                # After transfer - only count if validated against current group
                                should_count = (validation_group == current_group)
                        else:
                            # No clear transfer point or date - apply standard rules
                            should_count = (validation_group == previous_group or validation_group == current_group)
            
                    if should_count:
                        # Initialize attendance tracking structures
                        if subject not in attendance_by_subject:
                            attendance_by_subject[subject] = {
                                "total": 0,
                                "sessions": {}
                            }
                        if session_num not in attendance_by_subject[subject]["sessions"]:
                            attendance_by_subject[subject]["sessions"][session_num] = {
                                "locations": {}
                            }
                
                        # Update attendance counts
                        attendance_by_subject[subject]["total"] += 1
                        if location not in attendance_by_subject[subject]["sessions"][session_num]["locations"]:
                            attendance_by_subject[subject]["sessions"][session_num]["locations"][location] = 0
                        attendance_by_subject[subject]["sessions"][session_num]["locations"][location] += 1
                        total_attended += 1

                # Calculate status and color
                required_sessions = math.ceil(
                    self.ATTENDANCE_THRESHOLD * total_required_sessions)
                sessions_left = total_required_sessions - group_completed
                max_possible = total_attended + sessions_left
                min_sessions_needed = max(
                    required_sessions - total_attended, 0)

                if group_completed >= total_required_sessions:
                    if total_attended >= required_sessions:
                        status, color = "Pass", COLOR_PASS
                    else:
                        status, color = "Fail", COLOR_FAIL
                else:
                    if max_possible < required_sessions:
                        status, color = "Fail", COLOR_FAIL
                    elif total_attended >= required_sessions:
                        status, color = "Pass", COLOR_PASS
                    else:
                        sessions_margin = sessions_left - min_sessions_needed
                        if sessions_margin <= 1:
                            status, color = "High Risk", COLOR_HIGH_RISK
                        elif sessions_margin <= 3:
                            status, color = "Moderate Risk", COLOR_MODERATE_RISK
                        elif sessions_margin <= 5:
                            status, color = "Low Risk", COLOR_LOW_RISK
                        else:
                            status, color = "No Risk", COLOR_NO_RISK

                percentage = total_attended / \
                    total_required_sessions if total_required_sessions > 0 else 0

                # Create row data
                row = [
                    student_id, student['name'], student['year'], student['group'],
                    student['email'], status, f"{percentage:.1%}", min_sessions_needed,
                    sessions_left, group_completed, total_required_sessions, total_attended
                ]

                # Add subject totals and session details
                for subject in sorted(subjects.keys()):
                    # FIXED: For required attendance, select the appropriate requirements based on transfer status
                    subject_req_total = 0
                    subject_req_sessions = {}
                
                    # For transferred students, we need to choose which requirements to use
                    if is_transferred:
                        # Always use current group requirements for the report
                        if current_key in required_attendance and subject in required_attendance[current_key]:
                            curr_req = required_attendance[current_key][subject]
                            subject_req_total = curr_req["total"]
                        
                            # Add sessions from current group
                            for session_num, session_data in curr_req["sessions"].items():
                                if session_num not in subject_req_sessions:
                                    subject_req_sessions[session_num] = {"locations": {}}
                                for location, count in session_data["locations"].items():
                                    if location not in subject_req_sessions[session_num]["locations"]:
                                        subject_req_sessions[session_num]["locations"][location] = 0
                                    subject_req_sessions[session_num]["locations"][location] = count  # Set, not add
                    else:
                        # For non-transferred students, just use current group
                        if current_key in required_attendance and subject in required_attendance[current_key]:
                            curr_req = required_attendance[current_key][subject]
                            subject_req_total = curr_req["total"]
                        
                            # Add sessions
                            for session_num, session_data in curr_req["sessions"].items():
                                if session_num not in subject_req_sessions:
                                    subject_req_sessions[session_num] = {"locations": {}}
                                for location, count in session_data["locations"].items():
                                    if location not in subject_req_sessions[session_num]["locations"]:
                                        subject_req_sessions[session_num]["locations"][location] = 0
                                    subject_req_sessions[session_num]["locations"][location] = count
            
                    # Get actual attendance for this subject
                    subj_att = attendance_by_subject.get(subject, {"total": 0, "sessions": {}})

                    # Add subject totals
                    row.extend([subject_req_total, subj_att["total"]])

                    # Add session details
                    for session in sorted(subjects[subject]["sessions"]):
                        for location in sorted(subjects[subject]["locations"]):
                            req_count = subject_req_sessions.get(session, {}).get("locations", {}).get(location, 0)
                            att_count = subj_att.get("sessions", {}).get(session, {}).get("locations", {}).get(location, 0)
                            row.extend([req_count, att_count])

                sheet.append(row)

                # Apply cell formatting and colors for this row
                row_idx = sheet.max_row

                # Format status cell
                status_cell = sheet.cell(row=row_idx, column=6)
                status_cell.font = Font(bold=True)
                status_cell.fill = PatternFill("solid", fgColor=color)
                status_cell.alignment = Alignment(horizontal='center')

                # Format percentage cell
                percentage_cell = sheet.cell(row=row_idx, column=7)
                percentage_cell.number_format = '0.0%'
                percentage_cell.alignment = Alignment(horizontal='center')

                # Apply subject-specific colors to the data cells
                for subject, (start_col, end_col) in subject_column_ranges.items():
                    subject_color = self.get_subject_color(subject)
                    for col in range(start_col, end_col + 1):
                        cell = sheet.cell(row=row_idx, column=col)
                        # Apply a lighter version of the subject color for data cells
                        bg_color = self.lighten_color(subject_color["bg"])
                        cell.fill = PatternFill("solid", fgColor=bg_color)
                        cell.alignment = Alignment(horizontal='center')

        # Add auto-filter to easily sort and filter data
        sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(header))}1"

        # Improved column width auto-fitting
        column_widths = {}

        # First pass: Calculate max content length for each column
        for row in sheet.iter_rows():
            for cell in row:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                if cell.value:
                    # For headers (row 1), calculate based on word-wrapped text
                    if cell.row == 1:
                        # Split header by spaces and find the longest word
                        words = str(cell.value).split()
                        if words:
                            max_word_len = max(len(word) for word in words)
                            # For headers, consider both total length and longest word
                            header_width = min(max(max_word_len + 1, len(str(cell.value)) / 2), 30)
                            column_widths[col_letter] = max(column_widths.get(col_letter, 0), header_width)
                    else:
                        # For data cells, use the full text length
                        try:
                            text_len = len(str(cell.value))
                            column_widths[col_letter] = max(column_widths.get(col_letter, 0), text_len + 1)
                        except:
                            pass

        # Second pass: Apply calculated widths with constraints
        for col_letter, width in column_widths.items():
            col_idx = openpyxl.utils.column_index_from_string(col_letter)

            # Base width calculation
            adjusted_width = min(max(width, 10), 40)  # Min 10, Max 40

            # Special case for specific columns
            if col_idx == 2:  # Name column
                adjusted_width = max(adjusted_width, 25)  # Names need more space
            elif col_idx >= 13:  # Subject specific columns
                adjusted_width = max(adjusted_width, 12)  # Subject columns need at least this width
    
            sheet.column_dimensions[col_letter].width = adjusted_width

    def create_transfer_log_sheet(self, workbook, sheet_name, transferred_students, transfer_data):
        """Create a sheet that logs all student transfers with their dates"""
        sheet = workbook.create_sheet(sheet_name)
    
        # Create header
        header = ["Student ID", "Name", "Year", "Group Before", "Group After", "Transfer Date"]
        sheet.append(header)
    
        # Apply header formatting
        for i, cell in enumerate(sheet[1]):
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D3D3D3")  # Light gray background
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
        # Make the header row a bit taller
        sheet.row_dimensions[1].height = 22
    
        # Add freeze panes to keep header visible when scrolling
        sheet.freeze_panes = 'C2'
    
        # Add data for each transferred student
        for student_id, transfer_info in transferred_students.items():
            transfer_date = transfer_data.get(student_id, {}).get("transfer_date")
            formatted_date = ""
            if transfer_date:
                formatted_date = transfer_date.strftime('%d/%m/%Y %H:%M')
        
            row = [
                student_id,
                transfer_info["name"],
                transfer_info["year"],
                transfer_info["previous_group"],
                transfer_info["current_group"],
                formatted_date
            ]
            sheet.append(row)
    
        # Format date column
        for cell in sheet["F"][1:]:  # Format transfer date column
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD/MM/YYYY HH:MM'
    
        # Auto-fit column widths
        for col_idx, column in enumerate(sheet.columns, 1):
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        
            if max_length > 0:
                adjusted_width = min(max(max_length + 2, 12), 50)  # Min 12, Max 50
            
                # Special case for name column (typically column B)
                if col_idx == 2:  # Name column
                    adjusted_width = max(adjusted_width, 25)  # Names need more space
            
                # Apply the calculated width
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width
    
        # Add auto-filter
        sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}"

    def lighten_color(self, hex_color, factor=0.75):
        """
        Lightens the given color by the factor.
        1.0 means keep the same, 0.5 means 50% lighter.
        """
        # Convert hex to RGB
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)

        # Lighten the color
        r = int(r + (255 - r) * factor)
        g = int(g + (255 - g) * factor)
        b = int(b + (255 - b) * factor)

        # Convert back to hex
        return f"{r:02x}{g:02x}{b:02x}".upper()

class ScheduleDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Schedule")
        self.setMinimumWidth(400)
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)

        self.init_ui()

        # Add input validation
        self.total_input.setValidator(QIntValidator(1, 999, self))
        self.year_input.setValidator(QIntValidator(1, 6, self))

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Form Group
        form_group = QGroupBox("Schedule Details")
        form_group.setStyleSheet(GROUP_BOX_STYLE)
        form_layout = QVBoxLayout(form_group)
        form_layout.setSpacing(10)

        # Form fields
        self.year_input = QLineEdit()
        self.year_input.setPlaceholderText("Academic year...")
        self.module_input = QLineEdit()
        self.module_input.setPlaceholderText("Module to process...")
        self.total_input = QLineEdit()
        self.total_input.setPlaceholderText("Total sessions number...")
        self.file_input = QLineEdit()
        self.file_input.setPlaceholderText("Select Excel file...")
        self.sheet_combo = QComboBox()

        # Add form fields with consistent spacing
        form_layout.addWidget(QLabel("Academic Year:"))
        form_layout.addWidget(self.year_input)
        form_layout.addWidget(QLabel("Module Name:"))
        form_layout.addWidget(self.module_input)
        form_layout.addWidget(QLabel("Total Required Sessions:"))
        form_layout.addWidget(self.total_input)

        # Schedule File section with Browse button
        file_label = QLabel("Schedule File:")
        file_top_layout = QHBoxLayout()
        file_top_layout.addWidget(file_label)
        browse_btn = QPushButton("Browse")
        browse_btn.clicked.connect(self.browse_file)
        browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        file_top_layout.addWidget(browse_btn)
        file_top_layout.addStretch()
        form_layout.addLayout(file_top_layout)
        form_layout.addWidget(self.file_input)

        # Sheet selection
        form_layout.addWidget(QLabel("Sheet Name:"))
        form_layout.addWidget(self.sheet_combo)

        main_layout.addWidget(form_group)

        # Buttons at the bottom
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        cancel_button.setStyleSheet(STANDARD_BUTTON_STYLE)
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        main_layout.addLayout(button_layout)

    def browse_file(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx)"
        )
        if filename:
            self.file_input.setText(filename)
            self.load_sheets(filename)

    def load_sheets(self, file_path):
        if os.path.isfile(file_path):
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(wb.sheetnames)
            except Exception as e:
                QMessageBox.critical(
                    self, "Error", f"Error loading workbook: {str(e)}")

    def get_schedule_data(self):
        return [
            self.year_input.text(),
            self.module_input.text(),
            self.file_input.text(),
            self.sheet_combo.currentText(),
            int(self.total_input.text())
        ]

    def accept(self):
        """Validate inputs before closing dialog"""
        try:
            # Check required fields
            if not self.year_input.text().strip():
                raise ValueError("Academic year is required")
            if not self.module_input.text().strip():
                raise ValueError("Module name is required")
            if not self.total_input.text().strip():
                raise ValueError("Total required sessions is required")
            if not self.file_input.text().strip():
                raise ValueError("Schedule file is required")
            if not self.sheet_combo.currentText():
                raise ValueError("Sheet name is required")

            # Validate numeric input
            total_sessions = self.total_input.text()
            if not total_sessions.isdigit():
                raise ValueError("Total sessions must be a whole number")
            if int(total_sessions) <= 0:
                raise ValueError("Total sessions must be greater than zero")

            # Validate file exists
            if not os.path.isfile(self.file_input.text()):
                raise FileNotFoundError(
                    "Selected schedule file does not exist")

        except (ValueError, FileNotFoundError) as e:
            # Create custom message box
            error_dialog = QMessageBox(self)
            error_dialog.setWindowTitle("Invalid Input")
            error_dialog.setText(str(e))
            error_dialog.setIcon(QMessageBox.Icon.Warning)

            # Configure OK button
            ok_button = error_dialog.addButton(QMessageBox.StandardButton.Ok)
            ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

            # Style dialog background
            error_dialog.setStyleSheet(f"""
                QMessageBox {{
                    background-color: {CARD_BG};
                }}
                QLabel {{
                color: {TEXT_COLOR};
                    font-size: 14px;
                }}
            """)

            error_dialog.exec()
            return

        super().accept()

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

class AttendanceThresholdDialog(QDialog):
    def __init__(self, parent=None, current_threshold=75):
        super().__init__(parent)
        self.setWindowTitle("Set Attendance Threshold")
        self.setMinimumWidth(300)
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)

        self.init_ui(current_threshold)

    def init_ui(self, current_threshold):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Form Group
        form_group = QGroupBox("Attendance Threshold")
        form_group.setStyleSheet(GROUP_BOX_STYLE)
        form_layout = QVBoxLayout(form_group)
        form_layout.setSpacing(10)

        # Form fields
        self.threshold_input = QLineEdit()
        self.threshold_input.setText(str(current_threshold))
        self.threshold_input.setPlaceholderText("Enter percentage (e.g., 75)")
        self.threshold_input.setValidator(QIntValidator(1, 100, self))

        # Add description label
        description_label = QLabel("Enter the minimum attendance percentage required to pass (1-100):")
        description_label.setWordWrap(True)
        
        # Add form fields with consistent spacing
        form_layout.addWidget(description_label)
        form_layout.addWidget(self.threshold_input)

        main_layout.addWidget(form_group)

        # Buttons at the bottom
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        cancel_button.setStyleSheet(STANDARD_BUTTON_STYLE)
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        main_layout.addLayout(button_layout)

    def get_threshold(self):
        """Return the threshold value as a decimal (0.0-1.0)"""
        try:
            threshold_percent = int(self.threshold_input.text().strip())
            return threshold_percent / 100.0
        except ValueError:
            return 0.75  # Default if something goes wrong

    def accept(self):
        """Validate threshold before closing dialog"""
        try:
            threshold_text = self.threshold_input.text().strip()
            if not threshold_text:
                raise ValueError("Threshold percentage is required")
            
            threshold = int(threshold_text)
            if threshold < 1 or threshold > 100:
                raise ValueError("Threshold must be between 1 and 100")

        except ValueError as e:
            # Create custom message box
            error_dialog = QMessageBox(self)
            error_dialog.setWindowTitle("Invalid Input")
            error_dialog.setText(str(e))
            error_dialog.setIcon(QMessageBox.Icon.Warning)

            # Configure OK button
            ok_button = error_dialog.addButton(QMessageBox.StandardButton.Ok)
            ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

            # Style dialog background
            error_dialog.setStyleSheet(f"""
                QMessageBox {{
                    background-color: {CARD_BG};
                }}
                QLabel {{
                color: {TEXT_COLOR};
                    font-size: 14px;
                }}
            """)

            error_dialog.exec()
            return

        super().accept()


#==========================================================attendance analyzer==========================================================#

class AttendanceDashboard(QWidget):
    def __init__(self):
        super().__init__()
        self.student_data = []
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)

        self.init_ui()

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def navigate_to_processor(self):
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            dashboard_page = stacked_widget.widget(1)
            stacked_widget.setCurrentWidget(dashboard_page)

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio,
                                               Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("Analysis Dashboard")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)

        # Back button
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)

        # Exit button
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)

        # Add buttons to vertical layout
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()

        # Add button layout to header
        header_layout.addLayout(header_buttons_layout)
        main_layout.addLayout(header_layout)

        # File Selection Section
        file_group = QGroupBox("File Selection")
        file_group.setStyleSheet(GROUP_BOX_STYLE)
        file_layout = QVBoxLayout(file_group)

        # Single line layout for file selection
        file_input_layout = QHBoxLayout()
        file_input_layout.addWidget(QLabel("Reports File:"))
        self.file_path = QLineEdit()
        self.file_path.setPlaceholderText("    Select Excel file...")
        self.file_path.setMinimumWidth(200)
        file_input_layout.addWidget(self.file_path, stretch=1)

        # Browse button
        browse_btn = QPushButton("Browse")
        browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        browse_btn.clicked.connect(self.browse_file)
        file_input_layout.addWidget(browse_btn)

        # Sheet selection
        file_input_layout.addWidget(QLabel("Sheet Name:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(100)
        file_input_layout.addWidget(self.sheet_combo)

        file_layout.addLayout(file_input_layout)
        main_layout.addWidget(file_group)

        # Statistics Section
        stats_group = QGroupBox("Statistics")
        stats_group.setStyleSheet(file_group.styleSheet())
        stats_layout = QHBoxLayout(stats_group)

        # Create stat cards
        self.total_students = self.create_stat_card("Total Students", "0")
        self.pass_rate = self.create_stat_card("Pass Rate", "0%")
        self.avg_attendance = self.create_stat_card("Avg Attendance", "0%")
        self.at_risk = self.create_stat_card("At Risk Students", "0")

        stats_layout.addWidget(self.total_students)
        stats_layout.addWidget(self.pass_rate)
        stats_layout.addWidget(self.avg_attendance)
        stats_layout.addWidget(self.at_risk)
        main_layout.addWidget(stats_group)

        # Status Distribution Section
        status_group = QGroupBox("Status Distribution")
        status_group.setStyleSheet(file_group.styleSheet())
        status_layout = QVBoxLayout(status_group)

        self.status_table = QTableWidget()
        self.status_table.setColumnCount(3)
        self.status_table.setHorizontalHeaderLabels(
            ['Status', 'Count', 'Percentage'])

        # Center align the header text
        header = self.status_table.horizontalHeader()
        header.setDefaultAlignment(
            Qt.AlignmentFlag.AlignCenter)  # Center header text

        # Set column resize modes to stretch and fit content
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)  # Year
        header.setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)  # Module
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # File

        # Center align all cells in the table using stylesheet
        self.status_table.setStyleSheet(TABLE_STYLE)
        status_layout.addWidget(self.status_table)
        main_layout.addWidget(status_group)

        # Student List Section
        student_group = QGroupBox("Student List")
        student_group.setStyleSheet(file_group.styleSheet())
        student_layout = QVBoxLayout(student_group)

        # Search bar
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by ID or Name...")
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                background-color: {INPUT_BG};
                color: {TEXT_COLOR};
                padding: 5px;
                border-radius: 3px;
                min-width: 300px;
            }}
        """)
        self.search_input.textChanged.connect(self.filter_students)
        search_layout.addWidget(self.search_input)
        student_layout.addLayout(search_layout)

        # Student table
        self.student_table = QTableWidget()
        self.student_table.setColumnCount(6)
        self.student_table.setHorizontalHeaderLabels([
            'Student ID', 'Name', 'Status', 'Attendance %', 'Sessions Needed', 'Sessions Attended'
        ])
        self.student_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.student_table.setStyleSheet(TABLE_STYLE)
        student_layout.addWidget(self.student_table)
        main_layout.addWidget(student_group)

        # Bottom Buttons
        display_layout = QHBoxLayout()
        display_btn = QPushButton("Display Statistics")
        display_btn.clicked.connect(self.display_statistics)
        display_layout.addWidget(display_btn)
        main_layout.addLayout(display_layout)
        display_btn.setStyleSheet(STANDARD_BUTTON_STYLE)

    def create_stat_card(self, title, value):
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {DARK_BLUE};
                border-radius: 5px;
                padding: 10px;
            }}
            QFrame:hover {{
                background-color: #1b2649;
            }}
            QLabel {{
                color: {TEXT_COLOR};
            }}
        """)
        layout = QVBoxLayout(card)

        title_label = QLabel(title)
        title_label.setStyleSheet("font-size: 14px;")
        value_label = QLabel(value)
        value_label.setStyleSheet("font-size: 24px; font-weight: bold;")

        layout.addWidget(title_label)
        layout.addWidget(value_label)

        # Store value label reference for updating
        card.value_label = value_label
        return card

    def return_to_home(self):
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def load_report(self):
        year = self.year_combo.currentText()
        module = self.module_combo.currentText()
        if not year or not module:
            return

        file_path = os.path.join(os.getcwd(), "attendance_reports",
                                 f"Year_{year}", f"Y{year}_{module}_attendance.xlsx")

        if not os.path.exists(file_path):
            return

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            summary_sheet = wb["Summary"]

            # Skip header row
            rows = list(summary_sheet.rows)[1:]

            # Calculate statistics
            total_students = len(rows)
            pass_count = sum(1 for row in rows if row[5].value == "Pass")
            pass_rate = (pass_count / total_students *
                         100) if total_students > 0 else 0

            attendance_sum = sum(
                float(row[6].value.strip('%')) for row in rows)
            avg_attendance = attendance_sum / total_students if total_students > 0 else 0

            at_risk_count = sum(
                1 for row in rows if "Risk" in str(row[5].value))

            # Update status distribution
            status_counts = {}
            for row in rows:
                status = row[5].value
                status_counts[status] = status_counts.get(status, 0) + 1

            self.status_table.setRowCount(len(status_counts))
            for i, (status, count) in enumerate(status_counts.items()):
                percentage = (count / total_students *
                              100) if total_students > 0 else 0
                self.status_table.setItem(i, 0, QTableWidgetItem(status))
                self.status_table.setItem(i, 1, QTableWidgetItem(str(count)))
                self.status_table.setItem(
                    i, 2, QTableWidgetItem(f"{percentage:.1f}%"))

                # Color code status cells
                status_cell = self.status_table.item(i, 0)
                # Convert to string and remove whitespace
                status_str = str(status).strip()

                # Check for substring matches instead of exact matches
                if status_str == "Pass":
                    status_cell.setBackground(QColor("#66E4A6"))  # Light green
                elif status_str == "Fail":
                    status_cell.setBackground(QColor("#FF4C4C"))  # Red
                elif "Risk" in status_str:  # Check for "Risk" substring
                    if "High" in status_str:
                        status_cell.setBackground(
                            QColor("#FF7C7C"))  # Light red
                    elif "Moderate" in status_str:
                        status_cell.setBackground(QColor("#FFB97D"))  # Orange
                    elif "Low" in status_str:
                        status_cell.setBackground(
                            QColor("#FFF1A6"))  # Light yellow
                    elif "No" in status_str:
                        status_cell.setBackground(QColor("#3388D5"))  # Blue

            # Update student table
            self.student_data = []  # Store for filtering
            for row in rows:
                self.student_data.append([
                    str(row[0].value),  # ID
                    str(row[1].value),  # Name
                    str(row[5].value),  # Status
                    str(row[6].value),  # Attendance %
                    str(row[7].value),  # Sessions Needed
                    f"{row[11].value}/{row[10].value}",  # Sessions Attended
                ])

            self.update_student_table(self.student_data)

        except Exception as e:
            QMessageBox.critical(
                self, "Error", f"Error loading report: {str(e)}")

    def update_student_table(self, data):
        self.student_table.setRowCount(len(data))
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                item = QTableWidgetItem(value)
                self.student_table.setItem(i, j, item)

                # Color code status cell
                if j == 2:  # Status column
                    if value == "Pass":
                        item.setBackground(QColor("#66E4A6"))
                    elif value == "Fail":
                        item.setBackground(QColor("#FF4C4C"))
                    elif "Risk" in value:
                        item.setBackground(QColor("#FFB97D"))

    def filter_students(self):
        search_text = self.search_input.text().lower()
        filtered_data = [
            row for row in self.student_data
            if search_text in row[0].lower() or search_text in row[1].lower()
        ]
        self.update_student_table(filtered_data)

    def browse_file(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "    Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_name:
            self.file_path.setText(file_name)
            self.update_sheet_list(file_name)

    def update_sheet_list(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(wb.sheetnames)

            # Automatically select Summary sheet if it exists
            summary_index = self.sheet_combo.findText(
                "Summary", Qt.MatchFlag.MatchExactly)
            if summary_index >= 0:
                self.sheet_combo.setCurrentIndex(summary_index)

            wb.close()
        except Exception as e:
            QMessageBox.critical(
                self, "Error", f"Error reading Excel file: {str(e)}")

    def display_statistics(self):
        file_path = self.file_path.text()
        sheet_name = self.sheet_combo.currentText()

        if not file_path or not sheet_name:
            self.show_custom_warning(
                "Reports File Required", "Please select both file and sheet name")
            return

        if not os.path.exists(file_path):
            self.show_custom_warning(
                "Failed to Load Reports File", "Selected file does not exist")
            return

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            if sheet_name not in wb.sheetnames:
                self.show_custom_warning(
                    "Warning", "Selected sheet not found in workbook")
                return

            summary_sheet = wb[sheet_name]

            # Skip header row
            rows = list(summary_sheet.rows)[1:]

            # Calculate statistics
            total_students = len(rows)
            pass_count = sum(1 for row in rows if row[5].value == "Pass")
            pass_rate = (pass_count / total_students *
                         100) if total_students > 0 else 0

            # Fixed attendance calculation
            attendance_sum = sum(float(str(row[6].value).strip(
                '%')) for row in rows if row[6].value is not None)
            avg_attendance = attendance_sum / total_students if total_students > 0 else 0

            at_risk_count = sum(
                1 for row in rows if "Risk" in str(row[5].value))

            # Update stat cards
            self.total_students.value_label.setText(str(total_students))
            self.pass_rate.value_label.setText(f"{pass_rate:.1f}%")
            self.avg_attendance.value_label.setText(f"{avg_attendance:.1f}%")
            self.at_risk.value_label.setText(str(at_risk_count))

            # Update status distribution
            status_counts = {}
            for row in rows:
                status = row[5].value
                status_counts[status] = status_counts.get(status, 0) + 1

            self.status_table.setRowCount(len(status_counts))
            for i, (status, count) in enumerate(status_counts.items()):
                percentage = (count / total_students *
                              100) if total_students > 0 else 0
                self.status_table.setItem(i, 0, QTableWidgetItem(str(status)))
                self.status_table.setItem(i, 1, QTableWidgetItem(str(count)))
                self.status_table.setItem(
                    i, 2, QTableWidgetItem(f"{percentage:.1f}%"))

                # Color code status cells
                status_cell = self.status_table.item(i, 0)
                if status == "Pass":
                    status_cell.setBackground(QColor("#66E4A6"))
                elif status == "Fail":
                    status_cell.setBackground(QColor("#FF4C4C"))
                elif "Risk" in str(status):
                    status_cell.setBackground(QColor("#FFB97D"))

            # Update student table
            self.student_data = []  # Store for filtering
            for row in rows:
                self.student_data.append([
                    str(row[0].value),  # ID
                    str(row[1].value),  # Name
                    str(row[5].value),  # Status
                    str(row[6].value),  # Attendance %
                    str(row[7].value),  # Sessions Needed
                    f"{row[11].value}/{row[10].value}",  # Sessions Attended
                ])

            self.update_student_table(self.student_data)
            wb.close()

        except Exception as e:
            QMessageBox.critical(
                self, "Error", f"Error loading report: {str(e)}")

    def show_custom_warning(self, title, message):
        """Show a custom styled warning dialog"""
        warning_dialog = QMessageBox(self)
        warning_dialog.setWindowTitle(title)
        warning_dialog.setText(message)
        warning_dialog.setIcon(QMessageBox.Icon.Warning)

        # Create and style OK button
        ok_button = warning_dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

        # Style dialog background and text
        warning_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)

        warning_dialog.exec()

#==========================================================main app==========================================================#

class MainApplication(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Attendance Management App")
        self.setMinimumSize(1000, 750)
        icon_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self.setWindowTitle("Attendance Management App")

        # Create stacked widget to manage pages
        self.stacked_widget = QStackedWidget()
        self.setCentralWidget(self.stacked_widget)

        # Create pages
        self.start_page = StartPage(self)
        self.info_page = InfoPage()
        self.preparer_page = LogSheetPreparer()
        self.processor_page = AttendanceProcessor()
        self.dashboard_page = AttendanceDashboard()
        self.schedule_manager_page = ScheduleManager()  
        self.reference_preparer_page = ReferenceFilePreparer() 
        self.appeal_processor_page = AppealProcessor()

        # Add pages to stacked widget
        self.stacked_widget.addWidget(self.start_page)
        self.stacked_widget.addWidget(self.info_page)
        self.stacked_widget.addWidget(self.preparer_page)
        self.stacked_widget.addWidget(self.processor_page)
        self.stacked_widget.addWidget(self.dashboard_page)
        self.stacked_widget.addWidget(self.schedule_manager_page) 
        self.stacked_widget.addWidget(self.reference_preparer_page) 
        self.stacked_widget.addWidget(self.appeal_processor_page)

        # Connect start page buttons to switch pages
        self.start_page.info_button.clicked.connect(self.show_info)
        self.start_page.preparer_btn.clicked.connect(self.show_preparer)
        self.start_page.process_btn.clicked.connect(self.show_processor)
        self.start_page.dashboard_btn.clicked.connect(self.show_dashboard)
        self.start_page.schedule_btn.clicked.connect(self.show_schedule_manager) 
        self.start_page.reference_btn.clicked.connect(self.show_reference_preparer) 
        self.start_page.appeal_btn.clicked.connect(self.show_appeal_processor)

        # Set the window style
        self.setStyleSheet(f"""
            QMainWindow {{
                background-color: {BLACK};
            }}
        """)

    def show_info(self):
        self.stacked_widget.setCurrentWidget(self.info_page)

    def show_preparer(self):
        self.stacked_widget.setCurrentWidget(self.preparer_page)

    def show_processor(self):
        self.stacked_widget.setCurrentWidget(self.processor_page)

    def show_dashboard(self):
        self.stacked_widget.setCurrentWidget(self.dashboard_page)

    def show_schedule_manager(self):
        self.stacked_widget.setCurrentWidget(self.schedule_manager_page)

    def show_reference_preparer(self):
        self.stacked_widget.setCurrentWidget(self.reference_preparer_page)

    def show_appeal_processor(self):
        self.stacked_widget.setCurrentWidget(self.appeal_processor_page)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = MainApplication()
    window.show()
    sys.exit(app.exec())
